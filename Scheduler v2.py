import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import calendar
import datetime
import random
import json
import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import deepcopy

# --- Configuration & Constants ---
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

APP_NAME = "Clinic Physician Scheduler Pro"
STATE_FILE = "physician_state.json"

# A robust palette of distinct colors for the picker
COLOR_PALETTE = [
    "#FF0000", "#00FF00", "#0000FF", "#FFFF00", "#00FFFF", "#FF00FF", 
    "#C0C0C0", "#800000", "#808000", "#008000", "#800080", "#008080", 
    "#000080", "#FF4500", "#DA70D6", "#FA8072", "#20B2AA", "#778899", 
    "#B0C4DE", "#FFFFE0", "#FFD700", "#ADFF2F", "#7FFFD4", "#FF69B4",
    "#CD5C5C", "#4B0082", "#F0E68C", "#E6E6FA", "#FFF0F5", "#7CFC00"
]

DEFAULT_ROSTER = [
    ("Gandhi", 8), ("Wesley", 8), ("Khaja", 2), ("Rendon", 2),
    ("Reymunde", 2), ("Dash", 2), ("Govindu", 2), ("Lee", 2),
    ("Huq", 2), ("Koney", 2), ("Aisenberg", 2), ("Bhandari", 2)
]

# --- Helper Functions ---

def get_app_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def parse_date_input(text, year, month):
    if not text.strip():
        return []
    parts = [p.strip() for p in text.split(',')]
    result = []
    for p in parts:
        p = p.upper()
        try:
            day_str = ''.join(filter(str.isdigit, p))
            if not day_str: continue
            day = int(day_str)
            _, last_day = calendar.monthrange(year, month)
            if day < 1 or day > last_day: continue

            if "AM" in p:
                result.append({'day': day, 'type': 'AM'})
            elif "PM" in p:
                result.append({'day': day, 'type': 'PM'})
            else:
                result.append({'day': day, 'type': 'AM'})
                result.append({'day': day, 'type': 'PM'})
        except ValueError:
            continue
    return result

# --- Classes ---

class Physician:
    def __init__(self, name, target, active=True, half_month="All", preferred="", avoid="", override="", color="#FFFFFF"):
        self.id = str(random.getrandbits(32))
        self.name = name
        self.target = int(target)
        self.active = active
        self.half_month = half_month # "All", "1st", "2nd"
        self.preferred_str = preferred
        self.avoid_str = avoid
        self.override_str = override
        self.color = color
        
        self.assigned_shifts = [] 
    
    def to_dict(self):
        return {
            "name": self.name,
            "target": self.target,
            "active": self.active,
            "half_month": self.half_month,
            "preferred": self.preferred_str,
            "avoid": self.avoid_str,
            "override": self.override_str,
            "color": self.color
        }

    @classmethod
    def from_dict(cls, data):
        return cls(
            name=data.get("name", ""),
            target=data.get("target", 0),
            active=data.get("active", True),
            half_month=data.get("half_month", "All"),
            preferred=data.get("preferred", ""),
            avoid=data.get("avoid", ""),
            override=data.get("override", ""),
            color=data.get("color", "#FFFFFF")
        )

class SchedulerLogic:
    def __init__(self, physicians, year, month, daily_needs):
        self.physicians = [p for p in physicians if p.active]
        self.year = year
        self.month = month
        self.daily_needs = daily_needs
        self.schedule = {} 
        self.logs = []
        self.warnings = []
        
        _, self.last_day = calendar.monthrange(year, month)
        for d in range(1, self.last_day + 1):
            self.schedule[d] = {'AM': [], 'PM': []}
            
    def log(self, msg):
        self.logs.append(msg)

    def is_weekend(self, day):
        weekday = calendar.weekday(self.year, self.month, day)
        return weekday >= 5

    def get_valid_days(self, physician):
        valid = []
        for d in range(1, self.last_day + 1):
            if self.is_weekend(d): continue
            if physician.half_month == "1st" and d > 15: continue
            if physician.half_month == "2nd" and d <= 15: continue
            valid.append(d)
        return valid

    def can_assign(self, physician, day, shift_type, force=False):
        if not force:
            avoids = parse_date_input(physician.avoid_str, self.year, self.month)
            for avoid in avoids:
                if avoid['day'] == day and avoid['type'] == shift_type:
                    return False
        
        for assigned_d, assigned_t in physician.assigned_shifts:
            if assigned_d == day and assigned_t == shift_type:
                return False
        return True

    def run(self):
        for p in self.physicians:
            p.assigned_shifts = []

        # 1. Overrides
        for p in self.physicians:
            overrides = parse_date_input(p.override_str, self.year, self.month)
            for req in overrides:
                day, s_type = req['day'], req['type']
                if day > self.last_day: continue
                self.schedule[day][s_type].append(p.name)
                p.assigned_shifts.append((day, s_type))

        # Helper
        def get_open_slots():
            slots = []
            for d in range(1, self.last_day + 1):
                if self.is_weekend(d): continue
                needed_am = self.daily_needs.get(d, {}).get('AM', 0)
                current_am = len(self.schedule[d]['AM'])
                for _ in range(needed_am - current_am): slots.append((d, 'AM'))
                
                needed_pm = self.daily_needs.get(d, {}).get('PM', 0)
                current_pm = len(self.schedule[d]['PM'])
                for _ in range(needed_pm - current_pm): slots.append((d, 'PM'))
            return slots

        # 2. Standard Fill
        for p in self.physicians:
            needed = p.target - len(p.assigned_shifts)
            if needed <= 0: continue
            valid_days = self.get_valid_days(p)
            random.shuffle(valid_days)
            
            # Preferred
            prefs = parse_date_input(p.preferred_str, self.year, self.month)
            for pf in prefs:
                if needed <= 0: break
                d, t = pf['day'], pf['type']
                req = self.daily_needs.get(d, {}).get(t, 0)
                cur = len(self.schedule[d][t])
                if cur < req and self.can_assign(p, d, t):
                    self.schedule[d][t].append(p.name)
                    p.assigned_shifts.append((d, t))
                    needed -= 1

            # Random fill
            attempts = 0
            while needed > 0 and attempts < 150:
                attempts += 1
                open_slots = get_open_slots()
                p_slots = [s for s in open_slots if s[0] in valid_days and self.can_assign(p, s[0], s[1])]
                if not p_slots: break
                pick = random.choice(p_slots)
                self.schedule[pick[0]][pick[1]].append(p.name)
                p.assigned_shifts.append(pick)
                needed -= 1

        # 3. Round Robin
        open_slots = get_open_slots()
        if open_slots:
            rr_idx = 0 
            max_loops = len(open_slots) * len(self.physicians) * 3
            loop_count = 0
            while open_slots and loop_count < max_loops:
                d, t = open_slots[0]
                assigned = False
                start_idx = rr_idx % len(self.physicians)
                for i in range(len(self.physicians)):
                    p_idx = (start_idx + i) % len(self.physicians)
                    p = self.physicians[p_idx]
                    valid_days = self.get_valid_days(p)
                    if d in valid_days and self.can_assign(p, d, t):
                        self.schedule[d][t].append(p.name)
                        p.assigned_shifts.append((d, t))
                        assigned = True
                        rr_idx = p_idx + 1 
                        break
                if not assigned:
                    self.warnings.append(f"CRITICAL: Unfillable slot Day {d} {t}")
                    open_slots.pop(0)
                else:
                    open_slots = get_open_slots()
                loop_count += 1

class ColorPicker(ctk.CTkToplevel):
    def __init__(self, parent, current_color, taken_colors, on_select):
        super().__init__(parent)
        self.title("Select Color")
        self.geometry("400x400")
        self.attributes("-topmost", True)
        self.on_select = on_select
        self.taken_colors = taken_colors
        
        lbl = ctk.CTkLabel(self, text="Choose a color for this physician:", font=("Arial", 14))
        lbl.pack(pady=10)
        
        grid = ctk.CTkFrame(self)
        grid.pack(expand=True, padx=20, pady=20)
        
        # Ensure palette is large enough
        palette = list(COLOR_PALETTE)
        while len(palette) < len(taken_colors) + 5:
            # Generate random colors if we run out
            r = lambda: random.randint(0,255)
            palette.append('#%02X%02X%02X' % (r(),r(),r()))

        r, c = 0, 0
        for color in palette:
            # Determine logic: if taken, maybe mark it? 
            # Req: Allow selection, but swap ownership if taken.
            btn = ctk.CTkButton(grid, text="", width=40, height=40, fg_color=color, hover_color=color,
                                command=lambda col=color: self.select(col))
            btn.grid(row=r, column=c, padx=5, pady=5)
            c += 1
            if c > 4:
                c = 0
                r += 1

    def select(self, color):
        self.on_select(color)
        self.destroy()

class AppUI(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1400x900")
        
        self.physicians = []
        self.history = [] 
        self.schedule_results = None
        self.scheduler_logic = None
        
        self.tabview = ctk.CTkTabview(self)
        self.tabview.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.tab_emp = self.tabview.add("1. Physicians")
        self.tab_cal = self.tabview.add("2. Needs")
        self.tab_out = self.tabview.add("3. Schedule")
        self.tab_col = self.tabview.add("3.5 Color Schedule")
        self.tab_imp = self.tabview.add("4. Import/Export")
        self.tab_hlp = self.tabview.add("5. Instructions")
        
        self.setup_tab1()
        self.setup_tab2()
        self.setup_tab3()
        self.setup_tab_colored()
        self.setup_tab4()
        self.setup_tab5()
        
        self.load_state_from_disk()
        
    def save_snapshot(self):
        snap = [p.to_dict() for p in self.physicians]
        self.history.append(snap)
        if len(self.history) > 10: self.history.pop(0)

    def undo(self):
        if not self.history: return
        prev_state = self.history.pop()
        self.physicians = [Physician.from_dict(d) for d in prev_state]
        self.refresh_physician_list()

    def get_resource_path(self):
        return os.path.join(get_app_path(), STATE_FILE)

    def save_state_to_disk(self):
        self.update_physician_objects_from_ui()
        data = { "physicians": [p.to_dict() for p in self.physicians] }
        try:
            with open(self.get_resource_path(), 'w') as f: json.dump(data, f, indent=4)
        except: pass

    def load_state_from_disk(self):
        path = self.get_resource_path()
        if os.path.exists(path):
            try:
                with open(path, 'r') as f:
                    data = json.load(f)
                    self.physicians = [Physician.from_dict(x) for x in data.get("physicians", [])]
                self.refresh_physician_list()
                return
            except: pass
        self.reset_to_defaults()

    # --- TAB 1 ---
    def setup_tab1(self):
        ctrl_frame = ctk.CTkFrame(self.tab_emp)
        ctrl_frame.pack(fill="x", padx=5, pady=5)
        
        ctk.CTkButton(ctrl_frame, text="Reset Defaults", fg_color="#555555", command=self.reset_to_defaults).pack(side="left", padx=5)
        ctk.CTkButton(ctrl_frame, text="Undo", fg_color="#555555", command=self.undo).pack(side="left", padx=5)
        ctk.CTkButton(ctrl_frame, text="Add Physician", command=self.add_blank_physician).pack(side="left", padx=5)
        ctk.CTkButton(ctrl_frame, text="Clear All Dates", fg_color="#b30000", command=self.clear_all_dates).pack(side="right", padx=20)

        # Header with GRID
        self.header_frame = ctk.CTkFrame(self.tab_emp, height=30)
        self.header_frame.pack(fill="x", padx=5, pady=(5,0))
        
        headers = ["Order", "Active", "Name", "Target", "1st Half", "2nd Half", "Preferred", "Avoid", "Override", "Color", "Del"]
        # Weights for resizing
        self.header_frame.grid_columnconfigure(2, weight=1) # Name
        self.header_frame.grid_columnconfigure(6, weight=1) # Pref
        self.header_frame.grid_columnconfigure(7, weight=1) # Avoid
        
        for i, h in enumerate(headers):
            lbl = ctk.CTkLabel(self.header_frame, text=h, font=("Arial", 12, "bold"))
            lbl.grid(row=0, column=i, padx=2, sticky="ew")

        self.p_scroll = ctk.CTkScrollableFrame(self.tab_emp)
        self.p_scroll.pack(fill="both", expand=True, padx=5, pady=5)
        self.p_scroll.grid_columnconfigure(2, weight=1)
        self.p_scroll.grid_columnconfigure(6, weight=1)
        self.p_scroll.grid_columnconfigure(7, weight=1)

        self.p_rows = []

    def reset_to_defaults(self):
        self.save_snapshot()
        self.physicians = []
        for i, (name, target) in enumerate(DEFAULT_ROSTER):
            color = COLOR_PALETTE[i % len(COLOR_PALETTE)]
            self.physicians.append(Physician(name, target, color=color))
        self.refresh_physician_list()

    def add_blank_physician(self):
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        # Find unused color
        used = [p.color for p in self.physicians]
        avail = [c for c in COLOR_PALETTE if c not in used]
        col = avail[0] if avail else "#FFFFFF"
        self.physicians.append(Physician("New Doc", 0, color=col))
        self.refresh_physician_list()

    def clear_all_dates(self):
        if not messagebox.askyesno("Confirm", "Clear Override, Preferred, and Avoid dates for ALL physicians?"):
            return
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        for p in self.physicians:
            p.override_str = ""
            p.preferred_str = ""
            p.avoid_str = ""
        self.refresh_physician_list()

    def move_row(self, index, direction):
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        if direction == -1 and index > 0:
            self.physicians[index], self.physicians[index-1] = self.physicians[index-1], self.physicians[index]
        elif direction == 1 and index < len(self.physicians) - 1:
            self.physicians[index], self.physicians[index+1] = self.physicians[index+1], self.physicians[index]
        self.refresh_physician_list()

    def delete_row(self, index):
        if not messagebox.askyesno("Confirm", "Delete this physician?"): return
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        self.physicians.pop(index)
        self.refresh_physician_list()

    def open_color_picker(self, idx):
        self.update_physician_objects_from_ui()
        current_phys = self.physicians[idx]
        used_map = {p.color: p for p in self.physicians if p != current_phys}
        
        def on_color_chosen(new_color):
            if new_color in used_map:
                # Conflict!
                victim = used_map[new_color]
                # Assign victim a random unused color
                all_used = [p.color for p in self.physicians]
                avail = [c for c in COLOR_PALETTE if c not in all_used and c != new_color]
                if not avail: 
                    # Generate random
                    r = lambda: random.randint(0,255)
                    new_victim_col = '#%02X%02X%02X' % (r(),r(),r())
                else:
                    new_victim_col = avail[0]
                victim.color = new_victim_col
            
            current_phys.color = new_color
            self.refresh_physician_list()

        ColorPicker(self, current_phys.color, list(used_map.keys()), on_color_chosen)

    def toggle_half_month(self, idx, mode):
        # Mutual exclusion
        row_widgets = self.p_rows[idx]
        if mode == "1st":
            if row_widgets['chk_1st'].get():
                row_widgets['chk_2nd'].set(False)
        else:
            if row_widgets['chk_2nd'].get():
                row_widgets['chk_1st'].set(False)

    def toggle_override(self, index, btn_ref):
        curr_val = self.physicians[index].override_str
        dialog = ctk.CTkInputDialog(text="Override Dates (e.g. 12 AM):", title="Overrides")
        res = dialog.get_input()
        if res is not None:
            self.physicians[index].override_str = res
            if res.strip():
                btn_ref.configure(text="SET", fg_color="#b30000")
            else:
                btn_ref.configure(text="Set", fg_color="transparent", border_width=1)
            self.save_state_to_disk()

    def refresh_physician_list(self):
        for widget in self.p_scroll.winfo_children(): widget.destroy()
        self.p_rows = []

        for i, p in enumerate(self.physicians):
            # 1. Order
            f_ord = ctk.CTkFrame(self.p_scroll, fg_color="transparent")
            f_ord.grid(row=i, column=0, padx=2, pady=2)
            ctk.CTkButton(f_ord, text="▲", width=20, command=lambda x=i: self.move_row(x, -1)).pack(side="left")
            ctk.CTkButton(f_ord, text="▼", width=20, command=lambda x=i: self.move_row(x, 1)).pack(side="left")
            
            # 2. Active
            var_active = ctk.BooleanVar(value=p.active)
            ctk.CTkCheckBox(self.p_scroll, text="", variable=var_active, width=20).grid(row=i, column=1, padx=2)
            
            # 3. Name
            ent_name = ctk.CTkEntry(self.p_scroll)
            ent_name.insert(0, p.name)
            ent_name.grid(row=i, column=2, padx=2, sticky="ew")
            
            # 4. Target
            ent_tgt = ctk.CTkEntry(self.p_scroll, width=50)
            ent_tgt.insert(0, str(p.target))
            ent_tgt.grid(row=i, column=3, padx=2)
            
            # 5. Split (Checkboxes)
            # Logic: If half_month == "1st", check 1st. If "2nd", check 2nd. Else none.
            var_1st = ctk.BooleanVar(value=(p.half_month=="1st"))
            var_2nd = ctk.BooleanVar(value=(p.half_month=="2nd"))
            
            c1 = ctk.CTkCheckBox(self.p_scroll, text="", variable=var_1st, width=20, 
                                 command=lambda x=i: self.toggle_half_month(x, "1st"))
            c1.grid(row=i, column=4, padx=5)
            
            c2 = ctk.CTkCheckBox(self.p_scroll, text="", variable=var_2nd, width=20,
                                 command=lambda x=i: self.toggle_half_month(x, "2nd"))
            c2.grid(row=i, column=5, padx=5)

            # 6. Preferred
            ent_pref = ctk.CTkEntry(self.p_scroll)
            ent_pref.insert(0, p.preferred_str)
            ent_pref.grid(row=i, column=6, padx=2, sticky="ew")
            
            # 7. Avoid
            ent_avoid = ctk.CTkEntry(self.p_scroll)
            ent_avoid.insert(0, p.avoid_str)
            ent_avoid.grid(row=i, column=7, padx=2, sticky="ew")
            
            # 8. Override
            btn_over = ctk.CTkButton(self.p_scroll, text="SET" if p.override_str else "Set", width=50,
                                     fg_color="#b30000" if p.override_str else "transparent",
                                     border_color="red", border_width=1)
            btn_over.configure(command=lambda x=i, b=btn_over: self.toggle_override(x, b))
            btn_over.grid(row=i, column=8, padx=2)
            
            # 9. Color Picker
            # Translucent/Grey button that shows color
            btn_col = ctk.CTkButton(self.p_scroll, text="", width=30, height=20,
                                    fg_color=p.color, border_color="gray", border_width=1,
                                    command=lambda x=i: self.open_color_picker(x))
            btn_col.grid(row=i, column=9, padx=5)

            # 10. Delete (moved right)
            btn_del = ctk.CTkButton(self.p_scroll, text="X", width=30, fg_color="#b30000", 
                                    command=lambda x=i: self.delete_row(x))
            btn_del.grid(row=i, column=10, padx=15)
            
            self.p_rows.append({
                "active": var_active, "name": ent_name, "target": ent_tgt,
                "chk_1st": var_1st, "chk_2nd": var_2nd,
                "pref": ent_pref, "avoid": ent_avoid, "obj_id": p.id
            })
        self.save_state_to_disk()

    def update_physician_objects_from_ui(self):
        new_list = []
        pmap = {p.id: p for p in self.physicians}
        for row in self.p_rows:
            orig = pmap.get(row['obj_id'])
            name = row['name'].get()
            try: tgt = int(row['target'].get())
            except: tgt = 0
            
            hm = "All"
            if row['chk_1st'].get(): hm = "1st"
            elif row['chk_2nd'].get(): hm = "2nd"
            
            p = Physician(
                name=name, target=tgt, active=row['active'].get(),
                half_month=hm, preferred=row['pref'].get(), avoid=row['avoid'].get(),
                override=orig.override_str if orig else "",
                color=orig.color if orig else "#FFFFFF"
            )
            p.id = row['obj_id']
            new_list.append(p)
        self.physicians = new_list

    # --- TAB 2 ---
    def setup_tab2(self):
        ctrl = ctk.CTkFrame(self.tab_cal)
        ctrl.pack(pady=10)
        now = datetime.datetime.now()
        months = [str(i) for i in range(1, 13)]
        years = [str(now.year + i) for i in range(0, 5)]
        self.var_month = ctk.StringVar(value=str(now.month))
        self.var_year = ctk.StringVar(value=str(now.year))
        
        ctk.CTkLabel(ctrl, text="Month:").pack(side="left", padx=5)
        ctk.CTkComboBox(ctrl, values=months, variable=self.var_month, width=60, command=self.build_needs_grid).pack(side="left")
        ctk.CTkLabel(ctrl, text="Year:").pack(side="left", padx=5)
        ctk.CTkComboBox(ctrl, values=years, variable=self.var_year, width=70, command=self.build_needs_grid).pack(side="left")
        ctk.CTkButton(ctrl, text="Reset to Standards", command=self.reset_needs_std).pack(side="left", padx=20)

        self.needs_frame = ctk.CTkScrollableFrame(self.tab_cal)
        self.needs_frame.pack(fill="both", expand=True, padx=10, pady=10)
        self.needs_widgets = {}
        self.build_needs_grid()

    def build_needs_grid(self, _=None):
        for w in self.needs_frame.winfo_children(): w.destroy()
        self.needs_widgets = {}
        y = int(self.var_year.get())
        m = int(self.var_month.get())
        
        # Headers: Mon-Sun
        days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Action"]
        for i, d in enumerate(days):
            ctk.CTkLabel(self.needs_frame, text=d, font=("Arial", 14, "bold")).grid(row=0, column=i, padx=10, pady=5)
            
        cal = calendar.monthcalendar(y, m)
        r = 1
        
        # Default needs
        DEFAULT_MAP = {
            0: (1, 0), 1: (2, 0), 2: (0, 1), 3: (2, 1), 4: (1, 0), 5: (0, 0), 6: (0, 0)
        }

        for week_idx, week in enumerate(cal):
            week_widgets = []
            
            for day_idx, day_num in enumerate(week):
                # day_idx: 0=Mon, 6=Sun
                if day_num == 0:
                    ctk.CTkLabel(self.needs_frame, text="").grid(row=r, column=day_idx)
                    continue
                
                # Setup Frame
                is_weekend = (day_idx >= 5)
                bg_col = "#333333" if is_weekend else "transparent"
                cell = ctk.CTkFrame(self.needs_frame, border_width=1, border_color="gray", fg_color=bg_col)
                cell.grid(row=r, column=day_idx, padx=5, pady=5, sticky="nsew")
                
                ctk.CTkLabel(cell, text=str(day_num), font=("Arial", 12, "bold")).pack(pady=2)
                
                if not is_weekend:
                    def_am, def_pm = DEFAULT_MAP.get(day_idx, (0,0))
                    
                    # AM
                    f_am = ctk.CTkFrame(cell, fg_color="transparent")
                    f_am.pack(fill="x", padx=2, pady=1)
                    ctk.CTkLabel(f_am, text="AM", width=25).pack(side="left")
                    cmb_am = ctk.CTkComboBox(f_am, values=["0", "1", "2"], width=55, command=lambda v, w=None: self.color_combo(v, w))
                    cmb_am.set(str(def_am))
                    cmb_am.pack(side="left")
                    self.color_combo(str(def_am), cmb_am)
                    cmb_am.configure(command=lambda v, w=cmb_am: self.color_combo(v, w))
                    
                    # PM
                    f_pm = ctk.CTkFrame(cell, fg_color="transparent")
                    f_pm.pack(fill="x", padx=2, pady=1)
                    ctk.CTkLabel(f_pm, text="PM", width=25).pack(side="left")
                    cmb_pm = ctk.CTkComboBox(f_pm, values=["0", "1", "2"], width=55, command=lambda v, w=None: self.color_combo(v, w))
                    cmb_pm.set(str(def_pm))
                    cmb_pm.pack(side="left")
                    self.color_combo(str(def_pm), cmb_pm)
                    cmb_pm.configure(command=lambda v, w=cmb_pm: self.color_combo(v, w))
                    
                    self.needs_widgets[(day_num, 'AM')] = cmb_am
                    self.needs_widgets[(day_num, 'PM')] = cmb_pm
                    week_widgets.append(cmb_am)
                    week_widgets.append(cmb_pm)
                else:
                    ctk.CTkLabel(cell, text="Closed", text_color="gray").pack()

            # Close Week Button
            btn_close = ctk.CTkButton(self.needs_frame, text="Close\nWeek", width=60, fg_color="#800000",
                                      command=lambda w=week_widgets: self.close_week(w))
            btn_close.grid(row=r, column=7, padx=5)
            r += 1

    def color_combo(self, val, widget):
        if val == "0": widget.configure(fg_color="#550000", button_color="#550000")
        else: widget.configure(fg_color="#1f6aa5", button_color="#1f6aa5")

    def reset_needs_std(self):
        self.build_needs_grid()

    def close_week(self, widgets):
        for w in widgets:
            w.set("0")
            self.color_combo("0", w)

    def get_needs_data(self):
        data = {}
        for (day, slot), widget in self.needs_widgets.items():
            if day not in data: data[day] = {}
            try: data[day][slot] = int(widget.get())
            except: data[day][slot] = 0
        return data

    # --- TAB 3 & 3.5 ---
    def setup_tab3(self):
        self.setup_sched_tab(self.tab_out, colored_text=False)
    
    def setup_tab_colored(self):
        self.setup_sched_tab(self.tab_col, colored_text=True)

    def setup_sched_tab(self, parent_tab, colored_text):
        top = ctk.CTkFrame(parent_tab)
        top.pack(fill="x", padx=10, pady=5)
        
        btn = ctk.CTkButton(top, text="REGENERATE SCHEDULE", font=("Arial", 14, "bold"), 
                            fg_color="#006400", height=40, command=self.generate_schedule)
        btn.pack(fill="x")
        
        container = ctk.CTkFrame(parent_tab, fg_color="transparent")
        container.pack(fill="both", expand=True)
        
        # Calendar Area
        cal_frame = ctk.CTkScrollableFrame(container)
        cal_frame.pack(side="left", fill="both", expand=True, padx=5)
        
        # Configure columns to be uniform
        for i in range(5): 
            cal_frame.grid_columnconfigure(i, weight=1, uniform="days")
        
        if not colored_text:
            self.out_cal_frame = cal_frame
        else:
            self.col_cal_frame = cal_frame
            
        # Stats (Only on Tab 3 to save space/logic, or duplicate? Duplicate is fine)
        sidebar = ctk.CTkFrame(container, width=320)
        sidebar.pack(side="right", fill="y", padx=5)
        
        ctk.CTkLabel(sidebar, text="Statistics", font=("Arial", 14, "bold")).pack(pady=5)
        
        # Grid for stats
        stats_frame = ctk.CTkFrame(sidebar, fg_color="transparent")
        stats_frame.pack(fill="x", padx=5)
        
        # Store refs
        if not colored_text:
            self.stats_grid = stats_frame
            self.stats_label = ctk.CTkLabel(top, text="Status: Waiting", text_color="gray")
            self.stats_label.pack(pady=5)
            
            ctk.CTkLabel(sidebar, text="Issues / Warnings", font=("Arial", 14, "bold"), text_color="orange").pack(pady=10)
            self.warn_text = ctk.CTkTextbox(sidebar, height=200, text_color="orange")
            self.warn_text.pack(fill="both", expand=True, padx=5, pady=5)
        else:
            # Just a placeholder on the color tab side
            ctk.CTkLabel(stats_frame, text="(See Tab 3 for Details)").pack()

    def generate_schedule(self):
        self.update_physician_objects_from_ui()
        self.save_state_to_disk()
        y = int(self.var_year.get())
        m = int(self.var_month.get())
        needs = self.get_needs_data()
        
        logic = SchedulerLogic(self.physicians, y, m, needs)
        logic.run()
        self.scheduler_logic = logic
        
        # Render Both Tabs
        self.render_calendar_logic(self.out_cal_frame, logic, use_colors=False)
        self.render_calendar_logic(self.col_cal_frame, logic, use_colors=True)
        self.render_stats(logic)

    def render_calendar_logic(self, frame, logic, use_colors):
        for w in frame.winfo_children(): w.destroy()
        y, m = logic.year, logic.month
        
        days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
        for i, d in enumerate(days):
            ctk.CTkLabel(frame, text=d, font=("Arial", 14, "bold")).grid(row=0, column=i, padx=5, pady=5, sticky="ew")
            
        cal = calendar.monthcalendar(y, m)
        r = 1
        
        # Color Map
        c_map = {p.name: p.color for p in logic.physicians}

        for week in cal:
            has_weekday = False
            for day_idx, day_num in enumerate(week):
                if day_idx >= 5: continue 
                has_weekday = True
                
                # Frame for Day
                cell = ctk.CTkFrame(frame, border_width=1, border_color="#555555", height=150)
                cell.grid(row=r, column=day_idx, padx=2, pady=2, sticky="nsew")
                cell.grid_propagate(False) # Enforce size
                
                if day_num == 0:
                    cell.configure(fg_color="transparent", border_width=0)
                    continue
                
                # Header
                ctk.CTkLabel(cell, text=str(day_num), font=("Arial", 10, "bold"), text_color="gray").pack(anchor="ne", padx=2)
                
                # Logic
                day_sched = logic.schedule.get(day_num, {'AM': [], 'PM': []})
                needs_am = logic.daily_needs.get(day_num, {}).get('AM', 0)
                needs_pm = logic.daily_needs.get(day_num, {}).get('PM', 0)
                
                # AM Half
                f_am = ctk.CTkFrame(cell, fg_color="transparent")
                f_am.pack(fill="both", expand=True, padx=2)
                if needs_am > 0:
                    for name in day_sched['AM']:
                        col = c_map.get(name, "white") if use_colors else "#aaddff"
                        ctk.CTkLabel(f_am, text=name, text_color=col, font=("Arial", 12, "bold")).pack(anchor="w")
                    if len(day_sched['AM']) < needs_am:
                        ctk.CTkLabel(f_am, text="[OPEN]", text_color="red", font=("Arial", 10)).pack(anchor="w")
                
                # Divider
                ctk.CTkFrame(cell, height=1, fg_color="gray").pack(fill="x", padx=5, pady=1)

                # PM Half
                f_pm = ctk.CTkFrame(cell, fg_color="transparent")
                f_pm.pack(fill="both", expand=True, padx=2)
                if needs_pm > 0:
                    for name in day_sched['PM']:
                        col = c_map.get(name, "white") if use_colors else "#ffccaa"
                        ctk.CTkLabel(f_pm, text=name, text_color=col, font=("Arial", 12, "bold")).pack(anchor="w")
                    if len(day_sched['PM']) < needs_pm:
                        ctk.CTkLabel(f_pm, text="[OPEN]", text_color="red", font=("Arial", 10)).pack(anchor="w")

            if has_weekday: r += 1

    def render_stats(self, logic):
        # Clear stats grid
        for w in self.stats_grid.winfo_children(): w.destroy()
        
        # Headers
        headers = ["Name", "Target", "Actual", "Net"]
        for i, h in enumerate(headers):
            ctk.CTkLabel(self.stats_grid, text=h, font=("Arial", 12, "bold")).grid(row=0, column=i, padx=5, pady=2, sticky="w")
            
        r = 1
        total_slots = 0
        filled_slots = 0
        for d, reqs in logic.daily_needs.items():
            total_slots += reqs.get('AM', 0) + reqs.get('PM', 0)
            filled_slots += len(logic.schedule[d]['AM']) + len(logic.schedule[d]['PM'])
        
        if filled_slots < total_slots:
            self.stats_label.configure(text=f"MISSING SHIFTS: {filled_slots}/{total_slots}", text_color="red")
        else:
            self.stats_label.configure(text="ALL SHIFTS FILLED ✔", text_color="#00ff00")

        for p in logic.physicians:
            act = len(p.assigned_shifts)
            net = act - p.target
            
            # Net Color Logic
            if net > 0: net_col = "yellow"; net_str = f"+{net}"
            elif net < 0: net_col = "#00BFFF"; net_str = f"{net}" # Deep Sky Blue
            else: net_col = "#00FF00"; net_str = "0"
            
            ctk.CTkLabel(self.stats_grid, text=p.name).grid(row=r, column=0, sticky="w", padx=5)
            ctk.CTkLabel(self.stats_grid, text=str(p.target)).grid(row=r, column=1, sticky="w", padx=5)
            ctk.CTkLabel(self.stats_grid, text=str(act)).grid(row=r, column=2, sticky="w", padx=5)
            ctk.CTkLabel(self.stats_grid, text=net_str, text_color=net_col).grid(row=r, column=3, sticky="w", padx=5)
            r += 1

        self.warn_text.delete("0.0", "end")
        if not all([len(s['AM'])+len(s['PM']) == logic.daily_needs.get(d,{}).get('AM',0)+logic.daily_needs.get(d,{}).get('PM',0) for d, s in logic.schedule.items()]):
             self.warn_text.insert("end", "CRITICAL: Not all shifts are filled.\n")
        
        for w in logic.warnings:
            self.warn_text.insert("end", f"- {w}\n")

    # --- TAB 4 & 5 (Same logic) ---
    def setup_tab4(self):
        f = ctk.CTkFrame(self.tab_imp)
        f.pack(expand=True)
        ctk.CTkLabel(f, text="Excel Operations", font=("Arial", 20)).pack(pady=20)
        ctk.CTkButton(f, text="Export Schedule to Excel", width=200, height=50, command=self.export_excel).pack(pady=10)
        ctk.CTkButton(f, text="Import State from Excel", width=200, height=50, command=self.import_excel).pack(pady=10)

    def export_excel(self):
        if not self.scheduler_logic: return
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not filename: return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Schedule"
            bold = Font(bold=True)
            center = Alignment(horizontal="center", vertical="top", wrap_text=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            fill_header = PatternFill("solid", fgColor="DDDDDD")
            
            days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
            for i, d in enumerate(days):
                cell = ws.cell(row=1, column=i+1, value=d)
                cell.font = bold; cell.fill = fill_header; cell.alignment = center
                ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = 25
            
            logic = self.scheduler_logic
            cal = calendar.monthcalendar(logic.year, logic.month)
            curr_row = 2
            
            for week in cal:
                has_weekday = False
                for day_idx, day_num in enumerate(week):
                    if day_idx >= 5: continue
                    has_weekday = True
                    cell = ws.cell(row=curr_row, column=day_idx+1)
                    cell.border = border; cell.alignment = center
                    if day_num == 0: continue
                    
                    ams = logic.schedule.get(day_num, {}).get('AM', [])
                    pms = logic.schedule.get(day_num, {}).get('PM', [])
                    content = f"{day_num}\nAM: {', '.join(ams)}\n\n\nPM: {', '.join(pms)}\n\n\n"
                    cell.value = content
                if has_weekday: curr_row += 1

            ws_set = wb.create_sheet("Settings")
            self.update_physician_objects_from_ui()
            data = [p.to_dict() for p in self.physicians]
            ws_set.cell(row=1, column=1, value=json.dumps(data))
            wb.save(filename)
            messagebox.showinfo("Success", "Export complete.")
        except Exception as e: messagebox.showerror("Error", str(e))

    def import_excel(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not filename: return
        try:
            wb = openpyxl.load_workbook(filename)
            if "Settings" in wb.sheetnames:
                data = json.loads(wb["Settings"].cell(row=1, column=1).value)
                self.save_snapshot()
                self.physicians = [Physician.from_dict(d) for d in data]
                self.refresh_physician_list()
                messagebox.showinfo("Success", "Restored.")
        except Exception as e: messagebox.showerror("Error", str(e))

    def setup_tab5(self):
        txt = ctk.CTkTextbox(self.tab_hlp, font=("Arial", 14))
        txt.pack(fill="both", expand=True, padx=10, pady=10)
        info = """CLINIC SCHEDULER PRO - INSTRUCTIONS
        
1. PHYSICIANS (Tab 1)
   - Order: Use Up/Down arrows. Top doctors get priority in Round Robin.
   - Half Month: Check "1st" or "2nd" to restrict specific halves.
   - Color: Click the colored square to assign a unique color.
   - "Clear All Dates": Resets all override/avoid/preference fields.

2. NEEDS (Tab 2)
   - Calendar includes Weekends (Greyed out).
   - "Close Week" buttons set all needs for that row to 0.

3. SCHEDULE (Tab 3 & 3.5)
   - Tab 3: Standard colors.
   - Tab 3.5: Uses assigned physician colors.
   - Top Half = AM, Bottom Half = PM.

4. STATS
   - Net Shifts: Yellow (+), Green (0), Blue (-).
"""
        txt.insert("0.0", info)
        txt.configure(state="disabled")

if __name__ == "__main__":
    app = AppUI()
    app.mainloop()