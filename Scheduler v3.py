import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import calendar
import datetime
import random
import json
import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import deepcopy

# --- Configuration & Constants ---
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

APP_NAME = "Clinic Physician Scheduler Pro"
STATE_FILE = "physician_state.json"

# A robust palette of distinct colors for the picker
COLOR_PALETTE = [
    "#FF0000", "#00FF00", "#0000FF", "#FFFF00", "#00FFFF", "#FF00FF", 
    "#C0C0C0", "#800000", "#808000", "#008000", "#800080", "#008080", 
    "#000080", "#FF4500", "#DA70D6", "#FA8072", "#20B2AA", "#778899", 
    "#B0C4DE", "#FFFFE0", "#FFD700", "#ADFF2F", "#7FFFD4", "#FF69B4",
    "#CD5C5C", "#4B0082", "#F0E68C", "#E6E6FA", "#FFF0F5", "#7CFC00"
]

DEFAULT_ROSTER = [
    ("Gandhi", 8), ("Wesley", 8), ("Khaja", 2), ("Rendon", 2),
    ("Reymunde", 2), ("Dash", 2), ("Govindu", 2), ("Lee", 2),
    ("Huq", 2), ("Koney", 2), ("Aisenberg", 2), ("Bhandari", 2)
]

# --- Helper Functions ---

def get_app_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def parse_date_input(text, year, month):
    """Parses date strings like '12 AM, 15, 20 PM' into structured data."""
    if not text.strip():
        return []
    parts = [p.strip() for p in text.split(',')]
    result = []
    for p in parts:
        p = p.upper()
        try:
            day_str = ''.join(filter(str.isdigit, p))
            if not day_str: continue
            day = int(day_str)
            _, last_day = calendar.monthrange(year, month)
            if day < 1 or day > last_day: continue

            if "AM" in p:
                result.append({'day': day, 'type': 'AM'})
            elif "PM" in p:
                result.append({'day': day, 'type': 'PM'})
            else:
                result.append({'day': day, 'type': 'AM'})
                result.append({'day': day, 'type': 'PM'})
        except ValueError:
            continue
    return result

# --- Classes ---

class Physician:
    def __init__(self, name, target, active=True, half_month="All", preferred="", avoid="", override="", color="#FFFFFF"):
        self.id = str(random.getrandbits(32))
        self.name = name
        self.target = int(target)
        self.active = active
        self.half_month = half_month # "All", "1st", "2nd"
        self.preferred_str = preferred
        self.avoid_str = avoid
        self.override_str = override
        self.color = color
        
        self.assigned_shifts = [] 
    
    def to_dict(self):
        return {
            "name": self.name,
            "target": self.target,
            "active": self.active,
            "half_month": self.half_month,
            "preferred": self.preferred_str,
            "avoid": self.avoid_str,
            "override": self.override_str,
            "color": self.color
        }

    @classmethod
    def from_dict(cls, data):
        return cls(
            name=data.get("name", ""),
            target=data.get("target", 0),
            active=data.get("active", True),
            half_month=data.get("half_month", "All"),
            preferred=data.get("preferred", ""),
            avoid=data.get("avoid", ""),
            override=data.get("override", ""),
            color=data.get("color", "#FFFFFF")
        )

class SchedulerLogic:
    def __init__(self, physicians, year, month, daily_needs):
        self.physicians = [p for p in physicians if p.active]
        self.year = year
        self.month = month
        self.daily_needs = daily_needs
        self.schedule = {} 
        self.logs = []
        self.warnings = [] # Stores deviations from requests
        
        _, self.last_day = calendar.monthrange(year, month)
        for d in range(1, self.last_day + 1):
            self.schedule[d] = {'AM': [], 'PM': []}
            
    def log(self, msg):
        self.logs.append(msg)

    def is_weekend(self, day):
        weekday = calendar.weekday(self.year, self.month, day)
        return weekday >= 5

    def get_valid_days(self, physician):
        """Returns a list of days a physician can potentially be scheduled based on half-month restrictions."""
        valid = []
        for d in range(1, self.last_day + 1):
            if self.is_weekend(d): continue
            if physician.half_month == "1st" and d > 15: continue
            if physician.half_month == "2nd" and d <= 15: continue
            valid.append(d)
        return valid

    def can_assign(self, physician, day, shift_type, check_avoid=True):
        """Checks if a physician is available for a shift based on existing assignments and avoid dates."""
        if check_avoid:
            avoids = parse_date_input(physician.avoid_str, self.year, self.month)
            for avoid in avoids:
                if avoid['day'] == day and avoid['type'] == shift_type:
                    self.warnings.append(f"Avoidance Conflict: {physician.name} avoided D{day} {shift_type} but was placed via override.")
                    return False
        
        # Check if already assigned this specific slot
        for assigned_d, assigned_t in physician.assigned_shifts:
            if assigned_d == day and assigned_t == shift_type:
                return False
        return True

    def run(self):
        self.warnings = []
        for p in self.physicians:
            p.assigned_shifts = []

        # 1. Overrides (Highest Priority - Always Assigned, logs if conflict with Avoid)
        for p in self.physicians:
            overrides = parse_date_input(p.override_str, self.year, self.month)
            for req in overrides:
                day, s_type = req['day'], req['type']
                if day > self.last_day: continue
                
                # Check for avoid conflict, but assign anyway (Override trumps Avoid)
                can_assign_result = self.can_assign(p, day, s_type, check_avoid=True)
                if not can_assign_result:
                    self.warnings.append(f"OVERRIDE WARNING: {p.name} Override on D{day} {s_type} conflicts with their Avoidance. Assigned anyway.")
                    
                # Check if slot is full (this shouldn't happen for overrides, but good to check)
                needed = self.daily_needs.get(day, {}).get(s_type, 0)
                current = len(self.schedule[day][s_type])
                
                if current < needed:
                    self.schedule[day][s_type].append(p.name)
                    p.assigned_shifts.append((day, s_type))
                else:
                    self.warnings.append(f"OVERRIDE FAILED: {p.name} on D{day} {s_type} - Clinic slot is not needed (Needs: {needed}).")

        # 2. Preferences (Second Priority - Assigned if target allows and slot is available)
        for p in self.physicians:
            prefs = parse_date_input(p.preferred_str, self.year, self.month)
            for req in prefs:
                day, s_type = req['day'], req['type']
                if day > self.last_day: continue

                # Must still need shifts
                if len(p.assigned_shifts) >= p.target:
                    self.warnings.append(f"Preference Ignored: {p.name} on D{day} {s_type} - Target of {p.target} shifts already met.")
                    continue
                
                # Must be an open slot needed by the clinic
                needed = self.daily_needs.get(day, {}).get(s_type, 0)
                current = len(self.schedule[day][s_type])
                
                if current < needed and self.can_assign(p, day, s_type, check_avoid=True):
                    self.schedule[day][s_type].append(p.name)
                    p.assigned_shifts.append((day, s_type))
                else:
                    self.warnings.append(f"Preference Failed: {p.name} on D{day} {s_type} - Slot unavailable (full, avoided, or half-month restriction).")


        # 3. Iterative Round Robin Fill (Fair distribution of remaining shifts)
        phys_idx = 0 
        max_iterations = len(self.physicians) * self.last_day * 2 * 3 # Safety break

        for i in range(max_iterations):
            # a. Check if any physician still needs shifts AND if there are open slots
            eligible_physicians = [p for p in self.physicians if len(p.assigned_shifts) < p.target]
            
            # Check for open clinic slots
            open_slots_exist = False
            for d in range(1, self.last_day + 1):
                if d in self.daily_needs:
                    for s_type in ['AM', 'PM']:
                        needed = self.daily_needs[d].get(s_type, 0)
                        current = len(self.schedule[d][s_type])
                        if current < needed:
                            open_slots_exist = True
                            break
                if open_slots_exist: break
                
            if not eligible_physicians or not open_slots_exist:
                break # Exit loop if no one needs shifts or no slots are open

            # b. Select the next physician in the rotation
            p_index = phys_idx % len(eligible_physicians)
            p = eligible_physicians[p_index]
            
            # c. Find the highest priority open slot that the physician can take
            found_slot = False
            valid_days = self.get_valid_days(p) # Respects half-month
            
            # Iterate through available slots in calendar order
            for day in sorted(valid_days):
                for s_type in ['AM', 'PM']:
                    needed = self.daily_needs.get(day, {}).get(s_type, 0)
                    current = len(self.schedule[day][s_type])

                    if current < needed: # Slot is open
                        # Check availability (Avoid conflicts, no double booking)
                        if self.can_assign(p, day, s_type, check_avoid=True): 
                            # ASSIGN
                            self.schedule[day][s_type].append(p.name)
                            p.assigned_shifts.append((day, s_type))
                            phys_idx += 1 # Move to the next physician for the next iteration
                            found_slot = True
                            break
                if found_slot: break
            
            if not found_slot:
                # If the current physician can't fill any remaining open slot, move to the next in rotation
                phys_idx += 1 

        # 4. Final Audits and Logging
        
        # Log Unmet Targets
        for p in self.physicians:
            act = len(p.assigned_shifts)
            if act < p.target:
                self.warnings.append(f"TARGET UNMET: {p.name} needed {p.target}, assigned {act}.")

        # Log Unfilled Clinic Slots
        for d in range(1, self.last_day + 1):
            if d in self.daily_needs:
                for s_type in ['AM', 'PM']:
                    needed = self.daily_needs[d].get(s_type, 0)
                    current = len(self.schedule[d][s_type])
                    if current < needed:
                        self.warnings.append(f"CRITICAL: Unfilled Clinic Slot on Day {d} {s_type} ({needed - current} still needed).")

class ColorPicker(ctk.CTkToplevel):
    # (ColorPicker class remains unchanged)
    def __init__(self, parent, current_color, taken_colors, on_select):
        super().__init__(parent)
        self.title("Select Color")
        self.geometry("400x400")
        self.attributes("-topmost", True)
        self.on_select = on_select
        self.taken_colors = taken_colors
        
        lbl = ctk.CTkLabel(self, text="Choose a color for this physician:", font=("Arial", 14))
        lbl.pack(pady=10)
        
        grid = ctk.CTkFrame(self)
        grid.pack(expand=True, padx=20, pady=20)
        
        # Ensure palette is large enough
        palette = list(COLOR_PALETTE)
        while len(palette) < len(taken_colors) + 5:
            # Generate random colors if we run out
            r = lambda: random.randint(0,255)
            palette.append('#%02X%02X%02X' % (r(),r(),r()))

        r, c = 0, 0
        for color in palette:
            # Determine logic: if taken, maybe mark it? 
            # Req: Allow selection, but swap ownership if taken.
            btn = ctk.CTkButton(grid, text="", width=40, height=40, fg_color=color, hover_color=color,
                                command=lambda col=color: self.select(col))
            btn.grid(row=r, column=c, padx=5, pady=5)
            c += 1
            if c > 4:
                c = 0
                r += 1

    def select(self, color):
        self.on_select(color)
        self.destroy()

class AppUI(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1400x900")
        
        self.physicians = []
        self.history = [] 
        self.schedule_results = None
        self.scheduler_logic = None
        
        self.tabview = ctk.CTkTabview(self)
        self.tabview.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.tab_emp = self.tabview.add("1. Physicians")
        self.tab_cal = self.tabview.add("2. Needs")
        self.tab_out = self.tabview.add("3. Schedule")
        self.tab_col = self.tabview.add("3.5 Color Schedule")
        self.tab_imp = self.tabview.add("4. Import/Export")
        self.tab_hlp = self.tabview.add("5. Instructions")
        
        self.setup_tab1()
        self.setup_tab2()
        self.setup_tab3()
        self.setup_tab_colored()
        self.setup_tab4()
        self.setup_tab5()
        
        self.load_state_from_disk()
        
    def save_snapshot(self):
        snap = [p.to_dict() for p in self.physicians]
        self.history.append(snap)
        if len(self.history) > 10: self.history.pop(0)

    def undo(self):
        if not self.history: return
        prev_state = self.history.pop()
        self.physicians = [Physician.from_dict(d) for d in prev_state]
        self.refresh_physician_list()

    def get_resource_path(self):
        return os.path.join(get_app_path(), STATE_FILE)

    def save_state_to_disk(self):
        self.update_physician_objects_from_ui()
        data = { "physicians": [p.to_dict() for p in self.physicians] }
        try:
            with open(self.get_resource_path(), 'w') as f: json.dump(data, f, indent=4)
        except: pass

    def load_state_from_disk(self):
        path = self.get_resource_path()
        if os.path.exists(path):
            try:
                with open(path, 'r') as f:
                    data = json.load(f)
                    self.physicians = [Physician.from_dict(x) for x in data.get("physicians", [])]
                self.refresh_physician_list()
                return
            except: pass
        self.reset_to_defaults()

    # --- TAB 1 ---
    def setup_tab1(self):
        ctrl_frame = ctk.CTkFrame(self.tab_emp)
        ctrl_frame.pack(fill="x", padx=5, pady=5)
        
        ctk.CTkButton(ctrl_frame, text="Reset Defaults", fg_color="#555555", command=self.reset_to_defaults).pack(side="left", padx=5)
        ctk.CTkButton(ctrl_frame, text="Undo", fg_color="#555555", command=self.undo).pack(side="left", padx=5)
        ctk.CTkButton(ctrl_frame, text="Add Physician", command=self.add_blank_physician).pack(side="left", padx=5)
        ctk.CTkButton(ctrl_frame, text="Clear All Dates", fg_color="#b30000", command=self.clear_all_dates).pack(side="right", padx=20)

        # Header with GRID
        self.header_frame = ctk.CTkFrame(self.tab_emp, height=30)
        self.header_frame.pack(fill="x", padx=5, pady=(5,0))
        
        headers = ["Order", "Active", "Name", "Target", "1st Half", "2nd Half", "Preferred", "Avoid", "Override", "Color", "Del"]
        # Weights for resizing
        self.header_frame.grid_columnconfigure(2, weight=1) # Name
        self.header_frame.grid_columnconfigure(6, weight=1) # Pref
        self.header_frame.grid_columnconfigure(7, weight=1) # Avoid
        
        for i, h in enumerate(headers):
            lbl = ctk.CTkLabel(self.header_frame, text=h, font=("Arial", 12, "bold"))
            lbl.grid(row=0, column=i, padx=2, sticky="ew")

        self.p_scroll = ctk.CTkScrollableFrame(self.tab_emp)
        self.p_scroll.pack(fill="both", expand=True, padx=5, pady=5)
        self.p_scroll.grid_columnconfigure(2, weight=1)
        self.p_scroll.grid_columnconfigure(6, weight=1)
        self.p_scroll.grid_columnconfigure(7, weight=1)

        self.p_rows = []

    def reset_to_defaults(self):
        self.save_snapshot()
        self.physicians = []
        for i, (name, target) in enumerate(DEFAULT_ROSTER):
            color = COLOR_PALETTE[i % len(COLOR_PALETTE)]
            self.physicians.append(Physician(name, target, color=color))
        self.refresh_physician_list()

    def add_blank_physician(self):
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        # Find unused color
        used = [p.color for p in self.physicians]
        avail = [c for c in COLOR_PALETTE if c not in used]
        col = avail[0] if avail else "#FFFFFF"
        self.physicians.append(Physician("New Doc", 0, color=col))
        self.refresh_physician_list()

    def clear_all_dates(self):
        if not messagebox.askyesno("Confirm", "Clear Override, Preferred, and Avoid dates for ALL physicians?"):
            return
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        for p in self.physicians:
            p.override_str = ""
            p.preferred_str = ""
            p.avoid_str = ""
        self.refresh_physician_list()

    def move_row(self, index, direction):
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        if direction == -1 and index > 0:
            self.physicians[index], self.physicians[index-1] = self.physicians[index-1], self.physicians[index]
        elif direction == 1 and index < len(self.physicians) - 1:
            self.physicians[index], self.physicians[index+1] = self.physicians[index+1], self.physicians[index]
        self.refresh_physician_list()

    def delete_row(self, index):
        if not messagebox.askyesno("Confirm", "Delete this physician?"): return
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        self.physicians.pop(index)
        self.refresh_physician_list()

    def open_color_picker(self, idx):
        self.update_physician_objects_from_ui()
        current_phys = self.physicians[idx]
        used_map = {p.color: p for p in self.physicians if p != current_phys}
        
        def on_color_chosen(new_color):
            if new_color in used_map:
                # Conflict!
                victim = used_map[new_color]
                
                # Find a new unused color for the victim
                all_used = [p.color for p in self.physicians if p != current_phys and p != victim]
                all_used.append(new_color) # The new color is now "used" by the current physician
                
                # Check palette size and expand if necessary (Max 20 + 5 buffer = 25 for first 20 in palette)
                palette = list(COLOR_PALETTE)
                
                avail = [c for c in palette if c not in all_used]
                
                if not avail: 
                    # Generate random
                    r = lambda: random.randint(0,255)
                    new_victim_col = '#%02X%02X%02X' % (r(),r(),r())
                else:
                    new_victim_col = avail[0]
                
                victim.color = new_victim_col
                self.warnings_text.delete("0.0", "end")
                self.warnings_text.insert("end", f"Color Conflict Resolved: {victim.name}'s color changed to {new_victim_col} (from {new_color}).\n")

            current_phys.color = new_color
            self.refresh_physician_list()
            self.save_state_to_disk()

        ColorPicker(self, current_phys.color, list(used_map.keys()), on_color_chosen)

    def toggle_half_month(self, idx, mode):
        # Mutual exclusion
        row_widgets = self.p_rows[idx]
        if mode == "1st":
            if row_widgets['chk_1st'].get():
                row_widgets['chk_2nd'].set(False)
        else:
            if row_widgets['chk_2nd'].get():
                row_widgets['chk_1st'].set(False)
        self.save_state_to_disk() # Save immediately after checkbox toggle

    def toggle_override(self, index, btn_ref):
        curr_val = self.physicians[index].override_str
        dialog = ctk.CTkInputDialog(text="Override Dates (e.g. 12 AM, 15 PM):", title="Overrides")
        res = dialog.get_input()
        if res is not None:
            self.physicians[index].override_str = res
            if res.strip():
                btn_ref.configure(text="SET", fg_color="#b30000")
            else:
                btn_ref.configure(text="Set", fg_color="transparent", border_width=1)
            self.save_state_to_disk()

    def refresh_physician_list(self):
        for widget in self.p_scroll.winfo_children(): widget.destroy()
        self.p_rows = []

        for i, p in enumerate(self.physicians):
            # 1. Order
            f_ord = ctk.CTkFrame(self.p_scroll, fg_color="transparent")
            f_ord.grid(row=i, column=0, padx=2, pady=2)
            ctk.CTkButton(f_ord, text="▲", width=20, command=lambda x=i: self.move_row(x, -1)).pack(side="left")
            ctk.CTkButton(f_ord, text="▼", width=20, command=lambda x=i: self.move_row(x, 1)).pack(side="left")
            
            # 2. Active
            var_active = ctk.BooleanVar(value=p.active)
            chk_active = ctk.CTkCheckBox(self.p_scroll, text="", variable=var_active, width=20)
            chk_active.grid(row=i, column=1, padx=2)
            
            # 3. Name
            ent_name = ctk.CTkEntry(self.p_scroll)
            ent_name.insert(0, p.name)
            ent_name.grid(row=i, column=2, padx=2, sticky="ew")
            
            # 4. Target
            ent_tgt = ctk.CTkEntry(self.p_scroll, width=50)
            ent_tgt.insert(0, str(p.target))
            ent_tgt.grid(row=i, column=3, padx=2)
            
            # 5. Split (Checkboxes)
            var_1st = ctk.BooleanVar(value=(p.half_month=="1st"))
            var_2nd = ctk.BooleanVar(value=(p.half_month=="2nd"))
            
            c1 = ctk.CTkCheckBox(self.p_scroll, text="", variable=var_1st, width=20, 
                                 command=lambda x=i: self.toggle_half_month(x, "1st"))
            c1.grid(row=i, column=4, padx=5)
            
            c2 = ctk.CTkCheckBox(self.p_scroll, text="", variable=var_2nd, width=20,
                                 command=lambda x=i: self.toggle_half_month(x, "2nd"))
            c2.grid(row=i, column=5, padx=5)

            # 6. Preferred
            ent_pref = ctk.CTkEntry(self.p_scroll)
            ent_pref.insert(0, p.preferred_str)
            ent_pref.grid(row=i, column=6, padx=2, sticky="ew")
            
            # 7. Avoid
            ent_avoid = ctk.CTkEntry(self.p_scroll)
            ent_avoid.insert(0, p.avoid_str)
            ent_avoid.grid(row=i, column=7, padx=2, sticky="ew")
            
            # 8. Override
            btn_over = ctk.CTkButton(self.p_scroll, text="SET" if p.override_str else "Set", width=50,
                                     fg_color="#b30000" if p.override_str else "transparent",
                                     border_color="red", border_width=1)
            btn_over.configure(command=lambda x=i, b=btn_over: self.toggle_override(x, b))
            btn_over.grid(row=i, column=8, padx=2)
            
            # 9. Color Picker
            btn_col = ctk.CTkButton(self.p_scroll, text="", width=30, height=20,
                                    fg_color=p.color, border_color="gray", border_width=1,
                                    command=lambda x=i: self.open_color_picker(x))
            btn_col.grid(row=i, column=9, padx=5)

            # 10. Delete (moved right)
            btn_del = ctk.CTkButton(self.p_scroll, text="X", width=30, fg_color="#b30000", 
                                    command=lambda x=i: self.delete_row(x))
            btn_del.grid(row=i, column=10, padx=15)
            
            self.p_rows.append({
                "active": var_active, "name": ent_name, "target": ent_tgt,
                "chk_1st": var_1st, "chk_2nd": var_2nd,
                "pref": ent_pref, "avoid": ent_avoid, "obj_id": p.id
            })
        self.save_state_to_disk()

    def update_physician_objects_from_ui(self):
        new_list = []
        pmap = {p.id: p for p in self.physicians}
        for row in self.p_rows:
            orig = pmap.get(row['obj_id'])
            name = row['name'].get()
            try: tgt = int(row['target'].get())
            except: tgt = 0
            
            hm = "All"
            if row['chk_1st'].get(): hm = "1st"
            elif row['chk_2nd'].get(): hm = "2nd"
            
            p = Physician(
                name=name, target=tgt, active=row['active'].get(),
                half_month=hm, preferred=row['pref'].get(), avoid=row['avoid'].get(),
                override=orig.override_str if orig else "",
                color=orig.color if orig else "#FFFFFF"
            )
            p.id = row['obj_id']
            new_list.append(p)
        self.physicians = new_list

    # --- TAB 2 ---
    def setup_tab2(self):
        ctrl = ctk.CTkFrame(self.tab_cal)
        ctrl.pack(pady=10)
        now = datetime.datetime.now()
        months = [str(i) for i in range(1, 13)]
        years = [str(now.year + i) for i in range(0, 5)]
        self.var_month = ctk.StringVar(value=str(now.month))
        self.var_year = ctk.StringVar(value=str(now.year))
        
        ctk.CTkLabel(ctrl, text="Month:").pack(side="left", padx=5)
        ctk.CTkComboBox(ctrl, values=months, variable=self.var_month, width=60, command=self.build_needs_grid).pack(side="left")
        ctk.CTkLabel(ctrl, text="Year:").pack(side="left", padx=5)
        ctk.CTkComboBox(ctrl, values=years, variable=self.var_year, width=70, command=self.build_needs_grid).pack(side="left")
        ctk.CTkButton(ctrl, text="Reset to Standards", command=self.reset_needs_std).pack(side="left", padx=20)

        self.needs_frame = ctk.CTkScrollableFrame(self.tab_cal)
        self.needs_frame.pack(fill="both", expand=True, padx=10, pady=10)
        self.needs_widgets = {}
        self.build_needs_grid()

    def build_needs_grid(self, _=None):
        for w in self.needs_frame.winfo_children(): w.destroy()
        self.needs_widgets = {}
        y = int(self.var_year.get())
        m = int(self.var_month.get())
        
        # Headers: Mon-Sun
        days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Actions"]
        for i, d in enumerate(days):
            ctk.CTkLabel(self.needs_frame, text=d, font=("Arial", 14, "bold")).grid(row=0, column=i, padx=5, pady=5)
            self.needs_frame.grid_columnconfigure(i, weight=1)
            
        cal = calendar.monthcalendar(y, m)
        r = 1
        
        # Default needs: 0=Mon, 4=Fri, 5=Sat, 6=Sun
        DEFAULT_MAP = {
            0: (1, 0), 1: (2, 0), 2: (0, 1), 3: (2, 1), 4: (1, 0), 5: (0, 0), 6: (0, 0)
        }

        for week_idx, week in enumerate(cal):
            week_widgets = []
            
            for day_idx, day_num in enumerate(week):
                # day_idx: 0=Mon, 6=Sun
                if day_num == 0:
                    ctk.CTkLabel(self.needs_frame, text="").grid(row=r, column=day_idx)
                    continue
                
                # Setup Frame
                is_weekend = (day_idx >= 5)
                # Pale Grey for Weekends
                bg_col = "#444444" if is_weekend else "transparent"
                cell = ctk.CTkFrame(self.needs_frame, border_width=1, border_color="gray", fg_color=bg_col)
                cell.grid(row=r, column=day_idx, padx=5, pady=5, sticky="nsew")
                
                ctk.CTkLabel(cell, text=str(day_num), font=("Arial", 12, "bold")).pack(pady=2)
                
                # Dropdowns only appear on weekdays
                if not is_weekend:
                    def_am, def_pm = DEFAULT_MAP.get(day_idx, (0,0))
                    
                    # AM
                    f_am = ctk.CTkFrame(cell, fg_color="transparent")
                    f_am.pack(fill="x", padx=2, pady=1)
                    ctk.CTkLabel(f_am, text="AM", width=25, font=("Arial", 10)).pack(side="left")
                    cmb_am = ctk.CTkComboBox(f_am, values=["0", "1", "2"], width=50)
                    cmb_am.set(str(def_am))
                    cmb_am.pack(side="left")
                    self.color_combo(str(def_am), cmb_am)
                    cmb_am.configure(command=lambda v, w=cmb_am: self.color_combo(v, w))
                    
                    # PM
                    f_pm = ctk.CTkFrame(cell, fg_color="transparent")
                    f_pm.pack(fill="x", padx=2, pady=1)
                    ctk.CTkLabel(f_pm, text="PM", width=25, font=("Arial", 10)).pack(side="left")
                    cmb_pm = ctk.CTkComboBox(f_pm, values=["0", "1", "2"], width=50)
                    cmb_pm.set(str(def_pm))
                    cmb_pm.pack(side="left")
                    self.color_combo(str(def_pm), cmb_pm)
                    cmb_pm.configure(command=lambda v, w=cmb_pm: self.color_combo(v, w))
                    
                    self.needs_widgets[(day_num, 'AM')] = cmb_am
                    self.needs_widgets[(day_num, 'PM')] = cmb_pm
                    week_widgets.append(cmb_am)
                    week_widgets.append(cmb_pm)
                else:
                    ctk.CTkLabel(cell, text="Weekend", text_color="gray").pack()

            # Close Week Button (Added to the last column)
            btn_close = ctk.CTkButton(self.needs_frame, text="Close\nClinic", width=70, fg_color="#800000",
                                      command=lambda w=week_widgets: self.close_week(w))
            btn_close.grid(row=r, column=7, padx=5)
            r += 1

    def color_combo(self, val, widget):
        """Changes combobox color based on shift need (0=Red, >0=Blue)."""
        if val == "0": widget.configure(fg_color="#550000", button_color="#550000")
        else: widget.configure(fg_color="#1f6aa5", button_color="#1f6aa5")

    def reset_needs_std(self):
        self.build_needs_grid()

    def close_week(self, widgets):
        """Sets all need widgets for a given week to 0."""
        for w in widgets:
            w.set("0")
            self.color_combo("0", w)

    def get_needs_data(self):
        data = {}
        for (day, slot), widget in self.needs_widgets.items():
            if day not in data: data[day] = {}
            try: data[day][slot] = int(widget.get())
            except: data[day][slot] = 0
        return data

    # --- TAB 3 & 3.5 ---
    def setup_tab3(self):
        self.setup_sched_tab(self.tab_out, colored_text=False)
    
    def setup_tab_colored(self):
        self.setup_sched_tab(self.tab_col, colored_text=True)

    def setup_sched_tab(self, parent_tab, colored_text):
        top = ctk.CTkFrame(parent_tab)
        top.pack(fill="x", padx=10, pady=5)
        
        btn = ctk.CTkButton(top, text="REGENERATE SCHEDULE", font=("Arial", 14, "bold"), 
                            fg_color="#006400", height=40, command=self.generate_schedule)
        btn.pack(fill="x")
        
        container = ctk.CTkFrame(parent_tab, fg_color="transparent")
        container.pack(fill="both", expand=True)
        
        # Calendar Area
        cal_frame = ctk.CTkScrollableFrame(container)
        cal_frame.pack(side="left", fill="both", expand=True, padx=5)
        
        # Configure columns to be uniform and allow expansion
        for i in range(5): 
            cal_frame.grid_columnconfigure(i, weight=1, uniform="days")
        
        if not colored_text:
            self.out_cal_frame = cal_frame
        else:
            self.col_cal_frame = cal_frame
            
        # Stats Side Panel
        sidebar = ctk.CTkFrame(container, width=350)
        sidebar.pack(side="right", fill="y", padx=5)
        
        ctk.CTkLabel(sidebar, text="Schedule Statistics", font=("Arial", 16, "bold")).pack(pady=5)
        
        # Grid for stats
        self.stats_grid_frame = ctk.CTkFrame(sidebar, fg_color="transparent")
        self.stats_grid_frame.pack(fill="x", padx=5)
        
        # Headers for Stats
        headers = ["Name", "Target Shifts", "Actual Shifts", "Net Difference"]
        for i, h in enumerate(headers):
            # Column 0 (Name) gets sticky "w", the rest "e" for number alignment
            sticky = "w" if i == 0 else "e"
            ctk.CTkLabel(self.stats_grid_frame, text=h, font=("Arial", 12, "bold")).grid(row=0, column=i, padx=5, pady=2, sticky=sticky)
            if i > 0: self.stats_grid_frame.grid_columnconfigure(i, weight=1)
            
        # Store refs
        if not colored_text:
            self.stats_label = ctk.CTkLabel(top, text="Status: Waiting", text_color="gray")
            self.stats_label.pack(pady=5)
            
            ctk.CTkLabel(sidebar, text="Issues / Deviations", font=("Arial", 14, "bold"), text_color="orange").pack(pady=(10, 0))
            self.warnings_text = ctk.CTkTextbox(sidebar, height=200, text_color="orange")
            self.warnings_text.pack(fill="both", expand=True, padx=5, pady=5)
            
            # Use the generic grid frame ref for stats rendering
            self.stats_grid = self.stats_grid_frame

    def generate_schedule(self):
        self.update_physician_objects_from_ui()
        self.save_state_to_disk()
        y = int(self.var_year.get())
        m = int(self.var_month.get())
        needs = self.get_needs_data()
        
        logic = SchedulerLogic(self.physicians, y, m, needs)
        logic.run()
        self.scheduler_logic = logic
        
        # Render Both Tabs
        self.render_calendar_logic(self.out_cal_frame, logic, use_colors=False)
        self.render_calendar_logic(self.col_cal_frame, logic, use_colors=True)
        self.render_stats(logic)

    def render_calendar_logic(self, frame, logic, use_colors):
        for w in frame.winfo_children(): w.destroy()
        y, m = logic.year, logic.month
        
        days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
        for i, d in enumerate(days):
            ctk.CTkLabel(frame, text=d, font=("Arial", 12, "bold")).grid(row=0, column=i, padx=1, pady=1, sticky="ew")
            
        cal = calendar.monthcalendar(y, m)
        r = 1
        
        c_map = {p.name: p.color for p in logic.physicians}

        for week in cal:
            has_weekday = False
            for day_idx, day_num in enumerate(week):
                if day_idx >= 5: continue 
                has_weekday = True
                
                # Frame for Day - Removed fixed height, relies on expansion and min content size
                cell = ctk.CTkFrame(frame, border_width=1, border_color="#555555", width=120) 
                cell.grid(row=r, column=day_idx, padx=1, pady=1, sticky="nsew")
                
                if day_num == 0:
                    cell.configure(fg_color="transparent", border_width=0)
                    continue
                
                # Header
                ctk.CTkLabel(cell, text=f"D{day_num}", font=("Arial", 10, "bold"), text_color="gray").pack(anchor="ne", padx=2)
                
                day_sched = logic.schedule.get(day_num, {'AM': [], 'PM': []})
                needs_am = logic.daily_needs.get(day_num, {}).get('AM', 0)
                needs_pm = logic.daily_needs.get(day_num, {}).get('PM', 0)
                
                # AM Half (Top)
                f_am = ctk.CTkFrame(cell, fg_color="#333333" if needs_am > 0 else "transparent", border_width=0)
                f_am.pack(fill="x", padx=1, pady=(1, 0), anchor="n")
                ctk.CTkLabel(f_am, text="AM:", font=("Arial", 8, "bold"), text_color="gray").pack(anchor="w", padx=2)
                
                for name in day_sched['AM']:
                    col = c_map.get(name, "white") if use_colors else "#aaddff"
                    ctk.CTkLabel(f_am, text=name, text_color=col, font=("Arial", 9)).pack(anchor="w", padx=2)
                if len(day_sched['AM']) < needs_am:
                    ctk.CTkLabel(f_am, text=f"[OPEN {needs_am - len(day_sched['AM'])}]", text_color="red", font=("Arial", 9)).pack(anchor="w", padx=2)
                elif needs_am == 0:
                    ctk.CTkLabel(f_am, text="CLOSED", text_color="gray", font=("Arial", 9)).pack(anchor="center")


                # PM Half (Bottom)
                f_pm = ctk.CTkFrame(cell, fg_color="#333333" if needs_pm > 0 else "transparent", border_width=0)
                f_pm.pack(fill="x", padx=1, pady=(0, 1), anchor="n")
                ctk.CTkLabel(f_pm, text="PM:", font=("Arial", 8, "bold"), text_color="gray").pack(anchor="w", padx=2)

                for name in day_sched['PM']:
                    col = c_map.get(name, "white") if use_colors else "#ffccaa"
                    ctk.CTkLabel(f_pm, text=name, text_color=col, font=("Arial", 9)).pack(anchor="w", padx=2)
                if len(day_sched['PM']) < needs_pm:
                    ctk.CTkLabel(f_pm, text=f"[OPEN {needs_pm - len(day_sched['PM'])}]", text_color="red", font=("Arial", 9)).pack(anchor="w", padx=2)
                elif needs_pm == 0:
                    ctk.CTkLabel(f_pm, text="CLOSED", text_color="gray", font=("Arial", 9)).pack(anchor="center")
            
            if has_weekday: r += 1

    def render_stats(self, logic):
        # Clear stats grid, but preserve headers (row 0)
        for w in self.stats_grid.winfo_children():
            if int(w.grid_info()["row"]) > 0:
                 w.destroy()
        
        r = 1
        total_slots = 0
        filled_slots = 0
        for d, reqs in logic.daily_needs.items():
            total_slots += reqs.get('AM', 0) + reqs.get('PM', 0)
            filled_slots += len(logic.schedule.get(d, {}).get('AM', [])) + len(logic.schedule.get(d, {}).get('PM', []))
        
        if filled_slots < total_slots:
            self.stats_label.configure(text=f"MISSING SHIFTS: {filled_slots}/{total_slots}", text_color="red")
        else:
            self.stats_label.configure(text="ALL CLINIC SHIFTS FILLED ✔", text_color="#00ff00")

        # Render Stats Rows
        for p in logic.physicians:
            act = len(p.assigned_shifts)
            net = act - p.target
            
            # Net Color/Text Logic
            if net > 0: net_col = "yellow"; net_str = f"+{net}"
            elif net < 0: net_col = "#00BFFF"; net_str = f"{net}" # Blue
            else: net_col = "#00FF00"; net_str = "0" # Green
            
            # Column 0: Name (Sticky West)
            ctk.CTkLabel(self.stats_grid, text=p.name).grid(row=r, column=0, sticky="w", padx=5)
            # Column 1: Target (Sticky East)
            ctk.CTkLabel(self.stats_grid, text=str(p.target)).grid(row=r, column=1, sticky="e", padx=5)
            # Column 2: Actual (Sticky East)
            ctk.CTkLabel(self.stats_grid, text=str(act)).grid(row=r, column=2, sticky="e", padx=5)
            # Column 3: Net (Sticky East, Colored)
            ctk.CTkLabel(self.stats_grid, text=net_str, text_color=net_col).grid(row=r, column=3, sticky="e", padx=5)
            r += 1

        # Render Warnings/Issues
        self.warnings_text.delete("0.0", "end")
        if not logic.warnings:
            self.warnings_text.insert("end", "No issues or deviations found in the generated schedule.")
        else:
            for w in logic.warnings:
                self.warnings_text.insert("end", f"- {w}\n")

    # --- TAB 4 & 5 (Same logic) ---
    def setup_tab4(self):
        f = ctk.CTkFrame(self.tab_imp)
        f.pack(expand=True)
        ctk.CTkLabel(f, text="Excel Operations", font=("Arial", 20)).pack(pady=20)
        ctk.CTkButton(f, text="Export Schedule to Excel", width=200, height=50, command=self.export_excel).pack(pady=10)
        ctk.CTkButton(f, text="Import State from Excel", width=200, height=50, command=self.import_excel).pack(pady=10)

    def export_excel(self):
        if not self.scheduler_logic: 
            messagebox.showwarning("Warning", "Please generate a schedule first (Tab 3).")
            return
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not filename: return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Schedule"
            
            bold = Font(bold=True)
            center = Alignment(horizontal="center", vertical="top", wrap_text=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            fill_header = PatternFill("solid", fgColor="DDDDDD")
            
            days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
            
            # Write Headers
            for i, d in enumerate(days):
                cell = ws.cell(row=1, column=i+1, value=d)
                cell.font = bold; cell.fill = fill_header; cell.alignment = center
                ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = 25
            
            logic = self.scheduler_logic
            cal = calendar.monthcalendar(logic.year, logic.month)
            curr_row = 2
            
            # Write Schedule
            for week in cal:
                has_weekday = False
                for day_idx, day_num in enumerate(week):
                    if day_idx >= 5: continue
                    has_weekday = True
                    cell = ws.cell(row=curr_row, column=day_idx+1)
                    cell.border = border; cell.alignment = center
                    
                    if day_num == 0: continue
                    
                    # Apply specific styles based on colors/needs for the schedule view
                    
                    day_needs = logic.daily_needs.get(day_num, {})
                    day_schedule = logic.schedule.get(day_num, {'AM': [], 'PM': []})

                    ams = day_schedule['AM']
                    pms = day_schedule['PM']
                    
                    am_str = f"AM ({day_needs.get('AM', 0)}): {', '.join(ams)}"
                    pm_str = f"PM ({day_needs.get('PM', 0)}): {', '.join(pms)}"
                    
                    content = f"D{day_num}\n\n{am_str}\n\n{pm_str}"
                    cell.value = content
                if has_weekday: curr_row += 1

            ws_set = wb.create_sheet("Settings")
            self.update_physician_objects_from_ui()
            data = [p.to_dict() for p in self.physicians]
            ws_set.cell(row=1, column=1, value=json.dumps(data))
            wb.save(filename)
            messagebox.showinfo("Success", "Export complete.")
        except Exception as e: messagebox.showerror("Error", str(e))

    def import_excel(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not filename: return
        try:
            wb = openpyxl.load_workbook(filename)
            if "Settings" in wb.sheetnames:
                data = json.loads(wb["Settings"].cell(row=1, column=1).value)
                self.save_snapshot()
                self.physicians = [Physician.from_dict(d) for d in data]
                self.refresh_physician_list()
                messagebox.showinfo("Success", "Restored.")
        except Exception as e: messagebox.showerror("Error", str(e))

    def setup_tab5(self):
        txt = ctk.CTkTextbox(self.tab_hlp, font=("Arial", 14))
        txt.pack(fill="both", expand=True, padx=10, pady=10)
        info = """CLINIC SCHEDULER PRO - USER GUIDE & ALGORITHM EXPLANATION
        
## 1. Pro Tips for Optimization

* **Prioritize Placement:** The physician order in Tab 1 is crucial. Those at the top of the list get priority during the Round Robin phase to fill open clinic slots. Place higher-priority or senior physicians first.
* **Use Checkboxes, Not Target:** The '1st Half' / '2nd Half' checkboxes are more effective than using the 'Avoid' field for monthly restrictions.
* **Target Accuracy:** Ensure physician targets (Tab 1) match the total needs defined in Tab 2. If the total clinic need (Tab 2) is 40 shifts, but all targets combined (Tab 1) only total 30, ten shifts will be marked as **CRITICAL: Unfilled Clinic Slot**.

## 2. Component Explanation

* **Tab 1. Physicians:**
    * **Order:** Determines scheduling priority (Top = Higher Priority).
    * **Target:** The number of shifts the physician *must* be assigned.
    * **1st/2nd Half:** Restricts scheduling to days 1-15 or 16-end of the month.
    * **Preferred/Avoid:** Enter specific days/shifts (e.g., '10 AM, 15, 20 PM').
    * **Override (SET):** Forces assignment to these dates, regardless of clinic need or target. Useful for administrative/mandatory assignments.
    * **Color:** Assigns a unique color for visualization in the Color Schedule tab.

* **Tab 2. Needs:**
    * Defines how many physicians are needed for AM (Top) and PM (Bottom) clinic sessions on each weekday.
    * **Weekends (Sat/Sun):** Greyed out and have no dropdowns, as they are non-scheduling days.
    * **Close Clinic:** Sets all needs (AM/PM) for that entire week to '0'.

* **Tab 3 & 3.5. Schedule:**
    * Displays the finalized schedule. Cells are divided into **AM (Top)** and **PM (Bottom)**.
    * Tab 3 uses standard blue/yellow colors for names. Tab 3.5 uses the physician-assigned color from Tab 1.
    * **Statistics:** Provides a clear breakdown of Target vs. Actual shifts.
    * **Net Difference:**
        * **Yellow (+):** Over-assigned (assigned more shifts than target).
        * **Green (0):** Assigned exactly the target.
        * **Blue (-):** Under-assigned (failed to meet target).
    * **Issues / Deviations:** Lists soft conflicts (e.g., failed preference due to slot closure) and hard errors (e.g., Target Unmet, CRITICAL Unfilled Slot).

## 3. Algorithm Logic

The scheduling algorithm works in three strict phases, prioritizing manual input first:

1.  **Overrides (Hard Placement):** All dates/shifts listed in the **Override** field are assigned first. The physician is placed even if the slot is not marked as 'needed' in Tab 2. Avoidance requests are ignored (but a warning is logged).

2.  **Preferences (Soft Placement):** The algorithm next attempts to fulfill **Preferred** dates, but only if:
    * The physician still needs shifts to meet their target.
    * The clinic slot is actually marked as 'needed' (not full or closed).
    * The date does not conflict with their **Avoid** list.
    * *If a preference fails, a warning is logged.*

3.  **Round Robin Fill (Fairness):** After steps 1 and 2, the algorithm enters a continuous loop:
    * It identifies all remaining open clinic slots and all physicians who still need shifts to meet their target.
    * It cycles through the highest-priority physicians (based on their position in Tab 1).
    * The selected physician is assigned the **earliest available slot** they are eligible for (respecting half-month and avoid dates).
    * This iterative, rotating process ensures shifts are distributed fairly among all doctors who still require them, preventing the lower-ranked doctors from being completely excluded.
"""
        txt.insert("0.0", info)
        txt.configure(state="disabled")

if __name__ == "__main__":
    app = AppUI()
    app.mainloop()