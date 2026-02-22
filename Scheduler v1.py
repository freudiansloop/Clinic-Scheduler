import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import calendar
import datetime
import random
import json
import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import deepcopy

# --- Configuration & Constants ---
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

APP_NAME = "Clinic Physician Scheduler"
STATE_FILE = "physician_state.json"

DEFAULT_ROSTER = [
    ("Gandhi", 8), ("Wesley", 8), ("Khaja", 2), ("Rendon", 2),
    ("Reymunde", 2), ("Dash", 2), ("Govindu", 2), ("Lee", 2),
    ("Huq", 2), ("Koney", 2), ("Aisenberg", 2), ("Bhandari", 2)
]

DEFAULT_NEEDS = {
    0: (1, 0), # Mon
    1: (2, 0), # Tue
    2: (0, 1), # Wed
    3: (2, 1), # Thu
    4: (1, 0)  # Fri
}

# --- Helper Functions ---

def get_app_path():
    """Returns the path to the directory where the script/exe is running."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def parse_date_input(text, year, month):
    """
    Parses strings like '13', '13 AM', '13 PM' into a structural dict.
    Returns list of dicts: [{'day': 13, 'type': 'AM'}, ...]
    """
    if not text.strip():
        return []
    
    parts = [p.strip() for p in text.split(',')]
    result = []
    
    for p in parts:
        p = p.upper()
        try:
            day_str = ''.join(filter(str.isdigit, p))
            if not day_str: continue
            day = int(day_str)
            
            # Basic validation for day in month
            _, last_day = calendar.monthrange(year, month)
            if day < 1 or day > last_day:
                continue

            if "AM" in p:
                result.append({'day': day, 'type': 'AM'})
            elif "PM" in p:
                result.append({'day': day, 'type': 'PM'})
            else:
                # Both
                result.append({'day': day, 'type': 'AM'})
                result.append({'day': day, 'type': 'PM'})
        except ValueError:
            continue
    return result

# --- Classes ---

class Physician:
    def __init__(self, name, target, active=True, half_month="All", preferred="", avoid="", override=""):
        self.id = str(random.getrandbits(32)) # Unique ID for UI tracking
        self.name = name
        self.target = int(target)
        self.active = active
        self.half_month = half_month # "All", "1st", "2nd"
        self.preferred_str = preferred
        self.avoid_str = avoid
        self.override_str = override
        
        # Runtime stats (reset during generation)
        self.assigned_shifts = [] # List of tuples (day, type)
    
    def to_dict(self):
        return {
            "name": self.name,
            "target": self.target,
            "active": self.active,
            "half_month": self.half_month,
            "preferred": self.preferred_str,
            "avoid": self.avoid_str,
            "override": self.override_str
        }

    @classmethod
    def from_dict(cls, data):
        return cls(
            name=data.get("name", ""),
            target=data.get("target", 0),
            active=data.get("active", True),
            half_month=data.get("half_month", "All"),
            preferred=data.get("preferred", ""),
            avoid=data.get("avoid", ""),
            override=data.get("override", "")
        )

class SchedulerLogic:
    def __init__(self, physicians, year, month, daily_needs):
        self.physicians = [p for p in physicians if p.active]
        self.year = year
        self.month = month
        self.daily_needs = daily_needs # Dict {day_int: {'AM': req, 'PM': req}}
        self.schedule = {} # {day: {'AM': [names], 'PM': [names]}}
        self.logs = []
        self.warnings = []
        
        # Initialize empty schedule structure
        _, self.last_day = calendar.monthrange(year, month)
        for d in range(1, self.last_day + 1):
            self.schedule[d] = {'AM': [], 'PM': []}
            
    def log(self, msg):
        self.logs.append(msg)

    def is_weekend(self, day):
        # Monday is 0
        weekday = calendar.weekday(self.year, self.month, day)
        return weekday >= 5

    def get_valid_days(self, physician):
        """Return list of day integers available for this physician based on half-month rule."""
        valid = []
        for d in range(1, self.last_day + 1):
            if self.is_weekend(d): continue
            
            if physician.half_month == "1st" and d > 15: continue
            if physician.half_month == "2nd" and d <= 15: continue
            
            valid.append(d)
        return valid

    def can_assign(self, physician, day, shift_type, force=False):
        # check explicit avoids (unless forced by override)
        if not force:
            avoids = parse_date_input(physician.avoid_str, self.year, self.month)
            for avoid in avoids:
                if avoid['day'] == day and avoid['type'] == shift_type:
                    return False
        
        # Check if already working same shift (impossible) or same day (optional logic, allowing double shifts?)
        # Logic: A doctor can work AM and PM same day, but not AM twice (obviously).
        for assigned_d, assigned_t in physician.assigned_shifts:
            if assigned_d == day and assigned_t == shift_type:
                return False
        
        return True

    def run(self):
        # 1. Reset Stats
        for p in self.physicians:
            p.assigned_shifts = []

        # 2. Process Overrides (Highest Priority)
        for p in self.physicians:
            overrides = parse_date_input(p.override_str, self.year, self.month)
            for req in overrides:
                day, s_type = req['day'], req['type']
                if day > self.last_day: continue
                
                # Assign regardless of needs (it's an override)
                self.schedule[day][s_type].append(p.name)
                p.assigned_shifts.append((day, s_type))
                self.log(f"Override: {p.name} assigned to Day {day} {s_type}")

        # 3. Filling Loop (Target based)
        # We define slots needed based on Daily Needs minus current assignments
        
        # Helper to get open slots
        def get_open_slots():
            slots = []
            for d in range(1, self.last_day + 1):
                if self.is_weekend(d): continue
                
                needed_am = self.daily_needs.get(d, {}).get('AM', 0)
                current_am = len(self.schedule[d]['AM'])
                if current_am < needed_am:
                    for _ in range(needed_am - current_am):
                        slots.append((d, 'AM'))
                
                needed_pm = self.daily_needs.get(d, {}).get('PM', 0)
                current_pm = len(self.schedule[d]['PM'])
                if current_pm < needed_pm:
                    for _ in range(needed_pm - current_pm):
                        slots.append((d, 'PM'))
            return slots

        # Iteration for Standard Allocation
        for p in self.physicians:
            # Determine how many shifts still needed
            needed = p.target - len(p.assigned_shifts)
            if needed <= 0: continue

            valid_days = self.get_valid_days(p)
            random.shuffle(valid_days) # Non-deterministic
            
            # A. Try Preferred Dates first
            prefs = parse_date_input(p.preferred_str, self.year, self.month)
            for pf in prefs:
                if needed <= 0: break
                d, t = pf['day'], pf['type']
                
                # Check if slot exists and is open
                req_needed = self.daily_needs.get(d, {}).get(t, 0)
                curr_filled = len(self.schedule[d][t])
                
                if curr_filled < req_needed and self.can_assign(p, d, t):
                    self.schedule[d][t].append(p.name)
                    p.assigned_shifts.append((d, t))
                    needed -= 1
                    self.log(f"Preferred: {p.name} -> {d} {t}")

            # B. Fill remaining target with random valid slots
            # We refresh open slots list constantly or iterate differently?
            # Better strategy: Get all open slots, filter by physician validity, pick random
            attempts = 0
            while needed > 0 and attempts < 100:
                attempts += 1
                open_slots = get_open_slots()
                # Filter slots based on physician valid days and strict avoids
                p_slots = [s for s in open_slots if s[0] in valid_days and self.can_assign(p, s[0], s[1])]
                
                if not p_slots:
                    self.warnings.append(f"Could not reach target for {p.name} (Stuck at {len(p.assigned_shifts)}/{p.target})")
                    break
                
                # Pick one
                pick = random.choice(p_slots)
                self.schedule[pick[0]][pick[1]].append(p.name)
                p.assigned_shifts.append(pick)
                needed -= 1

        # 4. Overflow / Round Robin
        # If slots are still empty, force assign based on list priority
        open_slots = get_open_slots()
        if open_slots:
            self.log(f"Starting Round Robin for {len(open_slots)} holes.")
            
            # Infinite loop protector
            rr_idx = 0 
            max_loops = len(open_slots) * len(self.physicians) * 2
            loop_count = 0

            while open_slots and loop_count < max_loops:
                d, t = open_slots[0] # Take first hole
                
                # Find next doctor in priority list who can work this
                assigned = False
                start_idx = rr_idx % len(self.physicians)
                
                # Scan through doctors starting from current RR index
                for i in range(len(self.physicians)):
                    p_idx = (start_idx + i) % len(self.physicians)
                    p = self.physicians[p_idx]
                    
                    # In Round Robin, we respect Half-Month and strict Avoids, but ignore Target
                    valid_days = self.get_valid_days(p)
                    if d in valid_days and self.can_assign(p, d, t):
                        self.schedule[d][t].append(p.name)
                        p.assigned_shifts.append((d, t))
                        assigned = True
                        # Update RR index for next time to ensure fairness
                        rr_idx = p_idx + 1 
                        self.log(f"Round Robin: Assigned {p.name} to {d} {t}")
                        break
                
                if not assigned:
                    # Critical Failure for this slot - No one valid available
                    # Force assign to first active person even if Avoid? 
                    # For now, leave empty and warn
                    self.warnings.append(f"CRITICAL: No valid physician found for Hole Day {d} {t}")
                    open_slots.pop(0) # Remove to prevent inf loop
                else:
                    open_slots = get_open_slots() # Refresh
                
                loop_count += 1

        # 5. Post-Process Checks (Burnout)
        for p in self.physicians:
            # Sort shifts by day
            days = sorted(list(set([x[0] for x in p.assigned_shifts])))
            consecutive = 0
            for i in range(len(days) - 1):
                if days[i+1] == days[i] + 1:
                    consecutive += 1
                    if consecutive >= 2: # 3 days in a row (0+1=2 days, +1=3days)
                        msg = f"Burnout Warning: {p.name} is scheduled 3+ days in a row starting Day {days[i-1] if i>0 else days[i]}"
                        if msg not in self.warnings: self.warnings.append(msg)
                else:
                    consecutive = 0

class AppUI(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1300x850")
        
        # State
        self.physicians = []
        self.history = [] # For Undo
        self.schedule_results = None
        self.scheduler_logic = None
        
        # Main Layout
        self.tabview = ctk.CTkTabview(self)
        self.tabview.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.tab_emp = self.tabview.add("1. Physicians")
        self.tab_cal = self.tabview.add("2. Needs")
        self.tab_out = self.tabview.add("3. Schedule")
        self.tab_imp = self.tabview.add("4. Import/Export")
        self.tab_hlp = self.tabview.add("5. Instructions")
        
        # Setup Tabs
        self.setup_tab1()
        self.setup_tab2()
        self.setup_tab3()
        self.setup_tab4()
        self.setup_tab5()
        
        # Load initial state
        self.load_state_from_disk()
        
    def save_snapshot(self):
        """Saves current state to history for Undo."""
        # Deep copy current list
        snap = [p.to_dict() for p in self.physicians]
        self.history.append(snap)
        if len(self.history) > 10:
            self.history.pop(0)

    def undo(self):
        if not self.history:
            return
        prev_state = self.history.pop()
        self.physicians = [Physician.from_dict(d) for d in prev_state]
        self.refresh_physician_list()

    def get_resource_path(self):
        return os.path.join(get_app_path(), STATE_FILE)

    def save_state_to_disk(self):
        self.update_physician_objects_from_ui()
        data = {
            "physicians": [p.to_dict() for p in self.physicians],
            # Save calendar params if needed, for now mainly physicians
        }
        try:
            with open(self.get_resource_path(), 'w') as f:
                json.dump(data, f, indent=4)
        except Exception as e:
            print(f"Error saving state: {e}")

    def load_state_from_disk(self):
        path = self.get_resource_path()
        if os.path.exists(path):
            try:
                with open(path, 'r') as f:
                    data = json.load(f)
                    self.physicians = [Physician.from_dict(x) for x in data.get("physicians", [])]
                self.refresh_physician_list()
                return
            except:
                pass # Fail silently, load defaults
        
        # Load Defaults if no file
        self.reset_to_defaults()

    # --- TAB 1: Physicians ---
    def setup_tab1(self):
        # Header Controls
        ctrl_frame = ctk.CTkFrame(self.tab_emp)
        ctrl_frame.pack(fill="x", padx=5, pady=5)
        
        ctk.CTkButton(ctrl_frame, text="Reset to Defaults", fg_color="#555555", command=self.reset_to_defaults).pack(side="left", padx=5)
        ctk.CTkButton(ctrl_frame, text="Undo", fg_color="#555555", command=self.undo).pack(side="left", padx=5)
        ctk.CTkButton(ctrl_frame, text="Add New Row", command=self.add_blank_physician).pack(side="left", padx=5)
        
        # Headers
        header_frame = ctk.CTkFrame(self.tab_emp, height=30)
        header_frame.pack(fill="x", padx=5, pady=(5,0))
        headers = ["Order", "Active", "Name", "Target", "Month Split", "Preferred Dates", "Avoid Dates", "Override", "Delete"]
        widths = [60, 50, 150, 60, 150, 150, 150, 100, 50]
        
        for i, h in enumerate(headers):
            lbl = ctk.CTkLabel(header_frame, text=h, font=("Arial", 12, "bold"))
            # Rough grid simulation using pack logic is hard, using grid
            # But headers need to align with scrollable content. 
            # Simplified: Just labels spaced out roughly.
            pass 

        # Scrollable Area
        self.p_scroll = ctk.CTkScrollableFrame(self.tab_emp)
        self.p_scroll.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.p_rows = [] # Store widget references

    def reset_to_defaults(self):
        self.save_snapshot()
        self.physicians = []
        for name, target in DEFAULT_ROSTER:
            self.physicians.append(Physician(name, target))
        self.refresh_physician_list()

    def add_blank_physician(self):
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        self.physicians.append(Physician("New Doctor", 0))
        self.refresh_physician_list()

    def move_row(self, index, direction):
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        if direction == -1 and index > 0:
            self.physicians[index], self.physicians[index-1] = self.physicians[index-1], self.physicians[index]
        elif direction == 1 and index < len(self.physicians) - 1:
            self.physicians[index], self.physicians[index+1] = self.physicians[index+1], self.physicians[index]
        self.refresh_physician_list()

    def delete_row(self, index):
        if not messagebox.askyesno("Confirm", "Delete this physician?"):
            return
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        self.physicians.pop(index)
        self.refresh_physician_list()

    def toggle_override(self, index, btn_ref, entry_ref):
        # UI Trick: Show/Hide entry based on logic, or just styling
        # Implementation: Input dialog is cleaner for complicated string
        curr_val = self.physicians[index].override_str
        dialog = ctk.CTkInputDialog(text="Enter Override Dates (e.g. 12 AM, 15 PM):", title="Overrides")
        res = dialog.get_input()
        if res is not None:
            self.physicians[index].override_str = res
            if res.strip():
                btn_ref.configure(text="SET", fg_color="#b30000")
            else:
                btn_ref.configure(text="Override", fg_color="transparent", border_width=1, border_color="red")
            self.save_state_to_disk()

    def refresh_physician_list(self):
        # Clear existing
        for widget in self.p_scroll.winfo_children():
            widget.destroy()
        self.p_rows = []

        # Rebuild
        for i, p in enumerate(self.physicians):
            row_frame = ctk.CTkFrame(self.p_scroll, fg_color="transparent")
            row_frame.pack(fill="x", pady=2)
            
            # 1. Order Btns
            btn_up = ctk.CTkButton(row_frame, text="▲", width=25, command=lambda x=i: self.move_row(x, -1))
            btn_up.pack(side="left", padx=1)
            btn_down = ctk.CTkButton(row_frame, text="▼", width=25, command=lambda x=i: self.move_row(x, 1))
            btn_down.pack(side="left", padx=1)
            
            # 2. Active
            var_active = ctk.BooleanVar(value=p.active)
            chk = ctk.CTkCheckBox(row_frame, text="", variable=var_active, width=25)
            chk.pack(side="left", padx=5)
            
            # 3. Name
            ent_name = ctk.CTkEntry(row_frame, width=140, placeholder_text="Name")
            ent_name.insert(0, p.name)
            ent_name.pack(side="left", padx=2)
            
            # 4. Target
            ent_tgt = ctk.CTkEntry(row_frame, width=50)
            ent_tgt.insert(0, str(p.target))
            ent_tgt.pack(side="left", padx=2)
            
            # 5. Split (Combobox)
            cmb_split = ctk.CTkComboBox(row_frame, values=["All", "1st", "2nd"], width=70)
            cmb_split.set(p.half_month)
            cmb_split.pack(side="left", padx=2)
            
            # 6. Preferred
            ent_pref = ctk.CTkEntry(row_frame, width=120, placeholder_text="Preferred (e.g. 1, 5 AM)")
            ent_pref.insert(0, p.preferred_str)
            ent_pref.pack(side="left", padx=2)
            
            # 7. Avoid
            ent_avoid = ctk.CTkEntry(row_frame, width=120, placeholder_text="Avoid")
            ent_avoid.insert(0, p.avoid_str)
            ent_avoid.pack(side="left", padx=2)
            
            # 8. Override
            btn_over = ctk.CTkButton(row_frame, text="Override" if not p.override_str else "SET", 
                                     width=70, 
                                     fg_color="transparent" if not p.override_str else "#b30000",
                                     border_color="red", border_width=1)
            # Use command wrapper to pass refs
            btn_over.configure(command=lambda x=i, b=btn_over, e=None: self.toggle_override(x, b, e))
            btn_over.pack(side="left", padx=5)
            
            # 9. Delete
            btn_del = ctk.CTkButton(row_frame, text="X", width=30, fg_color="#b30000", command=lambda x=i: self.delete_row(x))
            btn_del.pack(side="left", padx=5)
            
            # Store refs for saving
            self.p_rows.append({
                "active": var_active,
                "name": ent_name,
                "target": ent_tgt,
                "split": cmb_split,
                "pref": ent_pref,
                "avoid": ent_avoid,
                "obj_id": p.id
            })
            
        self.save_state_to_disk()

    def update_physician_objects_from_ui(self):
        # Syncs UI state back to objects
        new_list = []
        # Create map of ID to object to preserve Override strings which aren't in simple entries
        pmap = {p.id: p for p in self.physicians}
        
        for row in self.p_rows:
            orig = pmap.get(row['obj_id'])
            
            name = row['name'].get()
            try:
                tgt = int(row['target'].get())
            except:
                tgt = 0
                
            p = Physician(
                name=name,
                target=tgt,
                active=row['active'].get(),
                half_month=row['split'].get(),
                preferred=row['pref'].get(),
                avoid=row['avoid'].get(),
                override=orig.override_str if orig else ""
            )
            p.id = row['obj_id'] # Keep ID
            new_list.append(p)
        self.physicians = new_list

    # --- TAB 2: Calendar Needs ---
    def setup_tab2(self):
        ctrl = ctk.CTkFrame(self.tab_cal)
        ctrl.pack(pady=10)
        
        now = datetime.datetime.now()
        months = [str(i) for i in range(1, 13)]
        years = [str(now.year + i) for i in range(0, 5)]
        
        self.var_month = ctk.StringVar(value=str(now.month))
        self.var_year = ctk.StringVar(value=str(now.year))
        
        ctk.CTkLabel(ctrl, text="Month:").pack(side="left", padx=5)
        ctk.CTkComboBox(ctrl, values=months, variable=self.var_month, width=60, command=self.build_needs_grid).pack(side="left")
        
        ctk.CTkLabel(ctrl, text="Year:").pack(side="left", padx=5)
        ctk.CTkComboBox(ctrl, values=years, variable=self.var_year, width=70, command=self.build_needs_grid).pack(side="left")
        
        ctk.CTkButton(ctrl, text="Reset to Standards", command=self.reset_needs_std).pack(side="left", padx=20)
        ctk.CTkButton(ctrl, text="Clear All", command=self.clear_needs).pack(side="left")

        self.needs_frame = ctk.CTkScrollableFrame(self.tab_cal)
        self.needs_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.needs_widgets = {} # Map (day, 'AM'/'PM') -> Combobox
        self.build_needs_grid()

    def build_needs_grid(self, _=None):
        for w in self.needs_frame.winfo_children():
            w.destroy()
        self.needs_widgets = {}
        
        y = int(self.var_year.get())
        m = int(self.var_month.get())
        
        # Grid Headers
        days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
        for i, d in enumerate(days):
            ctk.CTkLabel(self.needs_frame, text=d, font=("Arial", 14, "bold")).grid(row=0, column=i, padx=10, pady=5)
            
        cal = calendar.monthcalendar(y, m)
        r = 1
        
        for week in cal:
            has_weekday = False
            for day_idx, day_num in enumerate(week):
                if day_idx >= 5: continue # Skip Sat/Sun
                has_weekday = True
                
                if day_num == 0:
                    ctk.CTkLabel(self.needs_frame, text="").grid(row=r, column=day_idx)
                    continue
                
                # Day Cell
                cell = ctk.CTkFrame(self.needs_frame, border_width=1, border_color="gray")
                cell.grid(row=r, column=day_idx, padx=5, pady=5, sticky="nsew")
                
                ctk.CTkLabel(cell, text=str(day_num), font=("Arial", 12, "bold")).pack(pady=2)
                
                # Needs Defaults logic
                def_am, def_pm = DEFAULT_NEEDS[day_idx]
                
                # AM
                f_am = ctk.CTkFrame(cell, fg_color="transparent")
                f_am.pack(fill="x", padx=2, pady=1)
                ctk.CTkLabel(f_am, text="AM", width=30).pack(side="left")
                cmb_am = ctk.CTkComboBox(f_am, values=["0", "1", "2"], width=55, command=lambda v, c=None: self.color_combo(v, c))
                cmb_am.set(str(def_am))
                cmb_am.pack(side="left")
                self.color_combo(str(def_am), cmb_am) # Init color
                
                # PM
                f_pm = ctk.CTkFrame(cell, fg_color="transparent")
                f_pm.pack(fill="x", padx=2, pady=1)
                ctk.CTkLabel(f_pm, text="PM", width=30).pack(side="left")
                cmb_pm = ctk.CTkComboBox(f_pm, values=["0", "1", "2"], width=55, command=lambda v, c=None: self.color_combo(v, c))
                cmb_pm.set(str(def_pm))
                cmb_pm.pack(side="left")
                self.color_combo(str(def_pm), cmb_pm) # Init color
                
                self.needs_widgets[(day_num, 'AM')] = cmb_am
                self.needs_widgets[(day_num, 'PM')] = cmb_pm
                
                # Pass reference to combo so lambda works later? 
                # Actually lambda captures loop var? No, need binding.
                cmb_am.configure(command=lambda v, w=cmb_am: self.color_combo(v, w))
                cmb_pm.configure(command=lambda v, w=cmb_pm: self.color_combo(v, w))

            if has_weekday:
                r += 1

    def color_combo(self, val, widget):
        if val == "0":
            widget.configure(fg_color="#550000", button_color="#550000")
        else:
            widget.configure(fg_color="#1f6aa5", button_color="#1f6aa5") # Default blueish

    def reset_needs_std(self):
        self.build_needs_grid() # Rebuilds with defaults
        
    def clear_needs(self):
        for w in self.needs_widgets.values():
            w.set("0")
            self.color_combo("0", w)

    def get_needs_data(self):
        data = {}
        for (day, slot), widget in self.needs_widgets.items():
            if day not in data: data[day] = {}
            try:
                data[day][slot] = int(widget.get())
            except:
                data[day][slot] = 0
        return data

    # --- TAB 3: Schedule Output ---
    def setup_tab3(self):
        top = ctk.CTkFrame(self.tab_out)
        top.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkButton(top, text="REGENERATE SCHEDULE", font=("Arial", 14, "bold"), 
                      fg_color="#006400", height=40, command=self.generate_schedule).pack(fill="x")
        
        self.stats_label = ctk.CTkLabel(top, text="Status: Waiting to Generate", text_color="gray")
        self.stats_label.pack(pady=5)
        
        # Split: Calendar on Left, Stats/Warnings on Right
        container = ctk.CTkFrame(self.tab_out, fg_color="transparent")
        container.pack(fill="both", expand=True)
        
        # Calendar Area
        self.out_cal_frame = ctk.CTkScrollableFrame(container)
        self.out_cal_frame.pack(side="left", fill="both", expand=True, padx=5)
        
        # Sidebar
        sidebar = ctk.CTkFrame(container, width=300)
        sidebar.pack(side="right", fill="y", padx=5)
        
        ctk.CTkLabel(sidebar, text="Statistics", font=("Arial", 14, "bold")).pack(pady=5)
        self.stats_text = ctk.CTkTextbox(sidebar, height=300)
        self.stats_text.pack(fill="x", padx=5)
        
        ctk.CTkLabel(sidebar, text="Warnings & Issues", font=("Arial", 14, "bold"), text_color="orange").pack(pady=10)
        self.warn_text = ctk.CTkTextbox(sidebar, height=200, text_color="orange")
        self.warn_text.pack(fill="both", expand=True, padx=5, pady=5)

    def generate_schedule(self):
        # 1. Update Objects
        self.update_physician_objects_from_ui()
        self.save_state_to_disk()
        
        # 2. Get Params
        y = int(self.var_year.get())
        m = int(self.var_month.get())
        needs = self.get_needs_data()
        
        # 3. Run Logic
        logic = SchedulerLogic(self.physicians, y, m, needs)
        logic.run()
        self.scheduler_logic = logic # Store for export
        
        # 4. Render Results
        self.render_output_calendar(logic)
        self.render_stats(logic)

    def render_output_calendar(self, logic):
        for w in self.out_cal_frame.winfo_children():
            w.destroy()
            
        y = logic.year
        m = logic.month
        
        # Grid Headers
        days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
        for i, d in enumerate(days):
            ctk.CTkLabel(self.out_cal_frame, text=d, font=("Arial", 14, "bold")).grid(row=0, column=i, padx=5, pady=5)
            
        cal = calendar.monthcalendar(y, m)
        r = 1
        
        for week in cal:
            has_weekday = False
            for day_idx, day_num in enumerate(week):
                if day_idx >= 5: continue 
                has_weekday = True
                
                if day_num == 0:
                    ctk.CTkLabel(self.out_cal_frame, text="").grid(row=r, column=day_idx)
                    continue
                
                # Cell
                cell = ctk.CTkFrame(self.out_cal_frame, border_width=1, border_color="gray")
                cell.grid(row=r, column=day_idx, padx=2, pady=2, sticky="nsew")
                
                # Header Day
                ctk.CTkLabel(cell, text=str(day_num), font=("Arial", 10, "bold"), text_color="gray").pack(anchor="ne", padx=2)
                
                # Content
                day_sched = logic.schedule.get(day_num, {'AM': [], 'PM': []})
                
                # AM
                if day_sched['AM']:
                    lbl_am = ctk.CTkLabel(cell, text="AM: " + ", ".join(day_sched['AM']), 
                                          text_color="#aaddff", wraplength=120, font=("Arial", 11))
                    lbl_am.pack(anchor="w", padx=2)
                elif logic.daily_needs.get(day_num, {}).get('AM', 0) > 0:
                    ctk.CTkLabel(cell, text="AM: [OPEN]", text_color="red", font=("Arial", 10)).pack(anchor="w")

                # PM
                if day_sched['PM']:
                    lbl_pm = ctk.CTkLabel(cell, text="PM: " + ", ".join(day_sched['PM']), 
                                          text_color="#ffccaa", wraplength=120, font=("Arial", 11))
                    lbl_pm.pack(anchor="w", padx=2)
                elif logic.daily_needs.get(day_num, {}).get('PM', 0) > 0:
                    ctk.CTkLabel(cell, text="PM: [OPEN]", text_color="red", font=("Arial", 10)).pack(anchor="w")
                
            if has_weekday:
                r += 1

    def render_stats(self, logic):
        self.stats_text.delete("0.0", "end")
        self.warn_text.delete("0.0", "end")
        
        # Check if all slots filled
        all_filled = True
        total_slots = 0
        filled_slots = 0
        
        for d, reqs in logic.daily_needs.items():
            total_slots += reqs.get('AM', 0) + reqs.get('PM', 0)
            filled_slots += len(logic.schedule[d]['AM']) + len(logic.schedule[d]['PM'])
        
        if filled_slots < total_slots:
            all_filled = False
            self.stats_label.configure(text=f"MISSING SHIFTS: {filled_slots}/{total_slots}", text_color="red")
        else:
            self.stats_label.configure(text="ALL SHIFTS FILLED ✔", text_color="#00ff00")
            
        # Table
        txt = f"{'Name':<15} {'Tgt':<5} {'Act':<5} {'Net':<5}\n"
        txt += "-"*35 + "\n"
        
        for p in logic.physicians:
            act = len(p.assigned_shifts)
            net = act - p.target
            txt += f"{p.name:<15} {p.target:<5} {act:<5} {net:<5}\n"
            
        self.stats_text.insert("0.0", txt)
        
        # Warnings
        if not all_filled:
            self.warn_text.insert("end", "CRITICAL: Not all shifts are filled.\n")
            
        for w in logic.warnings:
            self.warn_text.insert("end", f"- {w}\n")
            
        if not logic.warnings and all_filled:
            self.warn_text.insert("end", "No issues detected.")

    # --- TAB 4: Import/Export ---
    def setup_tab4(self):
        f = ctk.CTkFrame(self.tab_imp)
        f.pack(expand=True)
        
        ctk.CTkLabel(f, text="Excel Operations", font=("Arial", 20)).pack(pady=20)
        
        ctk.CTkButton(f, text="Export Schedule to Excel", width=200, height=50, command=self.export_excel).pack(pady=10)
        ctk.CTkButton(f, text="Import State from Excel", width=200, height=50, command=self.import_excel).pack(pady=10)

    def export_excel(self):
        if not self.scheduler_logic:
            messagebox.showerror("Error", "Please generate a schedule first.")
            return
            
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not filename: return
        
        try:
            wb = openpyxl.Workbook()
            
            # 1. Calendar Sheet
            ws = wb.active
            ws.title = "Schedule"
            
            # Styles
            bold = Font(bold=True)
            center = Alignment(horizontal="center", vertical="top", wrap_text=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            fill_header = PatternFill("solid", fgColor="DDDDDD")
            
            # Headers
            days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
            for i, d in enumerate(days):
                cell = ws.cell(row=1, column=i+1, value=d)
                cell.font = bold
                cell.fill = fill_header
                cell.alignment = center
                ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = 25
            
            logic = self.scheduler_logic
            cal = calendar.monthcalendar(logic.year, logic.month)
            curr_row = 2
            
            for week in cal:
                # Determine max height for this week row based on resident slots
                # We need a block for AM and PM
                # Structure:
                # [Date]
                # AM: Dr X
                # [Space for Res]
                # [Space for Res]
                # [Space for Res]
                # PM: Dr Y
                # [Space for Res]...
                
                # Simplified: Just write text into cells
                has_weekday = False
                row_cells = []
                
                for day_idx, day_num in enumerate(week):
                    if day_idx >= 5: continue
                    has_weekday = True
                    cell = ws.cell(row=curr_row, column=day_idx+1)
                    cell.border = border
                    cell.alignment = center
                    
                    if day_num == 0:
                        cell.value = ""
                        continue
                        
                    content = f"{day_num}\n"
                    
                    # AM
                    ams = logic.schedule.get(day_num, {}).get('AM', [])
                    content += f"AM: {', '.join(ams)}\n"
                    content += "\n\n\n" # Space for residents
                    
                    # PM
                    pms = logic.schedule.get(day_num, {}).get('PM', [])
                    content += f"PM: {', '.join(pms)}\n"
                    content += "\n\n\n" # Space for residents
                    
                    cell.value = content
                
                if has_weekday:
                    curr_row += 1

            # 2. Settings Sheet (Persistence)
            ws_set = wb.create_sheet("Settings")
            # Save raw JSON string
            self.update_physician_objects_from_ui()
            data = [p.to_dict() for p in self.physicians]
            ws_set.cell(row=1, column=1, value=json.dumps(data))
            
            wb.save(filename)
            messagebox.showinfo("Success", "Export complete.")
            
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def import_excel(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not filename: return
        
        try:
            wb = openpyxl.load_workbook(filename)
            if "Settings" in wb.sheetnames:
                ws = wb["Settings"]
                json_str = ws.cell(row=1, column=1).value
                if json_str:
                    data = json.loads(json_str)
                    self.save_snapshot() # Undo point
                    self.physicians = [Physician.from_dict(d) for d in data]
                    self.refresh_physician_list()
                    messagebox.showinfo("Success", "Physician list restored from Excel.")
                else:
                    messagebox.showerror("Error", "Settings sheet empty.")
            else:
                messagebox.showerror("Error", "No 'Settings' sheet found in this Excel file.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # --- TAB 5: Instructions ---
    def setup_tab5(self):
        txt = ctk.CTkTextbox(self.tab_hlp, font=("Arial", 14))
        txt.pack(fill="both", expand=True, padx=10, pady=10)
        
        info = """
        CLINIC SCHEDULER INSTRUCTIONS
        =============================
        
        1. HOW IT WORKS
        ----------------
        This app generates a monthly schedule using a randomized algorithm while respecting
        hierarchical constraints. The "Generate" button is non-deterministic, meaning you
        can click it multiple times to get different valid schedules.

        2. PRIORITY HIERARCHY
        ---------------------
        The algorithm fills slots in this strict order:
        1. Overrides: (Highest) These are forced, ignoring other rules.
        2. Preferred Dates: Attempted first if target not reached.
        3. Random Fill: Fills remaining target slots from available days.
        4. Overflow (Round Robin): If holes remain after everyone meets their target,
           the app assigns extra shifts to physicians in list order.

        3. MANAGING PHYSICIANS (Tab 1)
        ------------------------------
        - ORDER MATTERS: Use the Up/Down arrows. Doctors at the top get priority
          during the Round Robin phase and filling phase.
        - TIP: Move "hard to schedule" doctors (lots of restrictions) to the top.
        - HALF MONTH: Use "1st" (Days 1-15) or "2nd" (16-End) if a doctor rotates.
        - DATES FORMATTING:
          "12"      -> Both AM and PM on the 12th are considered.
          "12 AM"   -> Only 12th AM.
          "12, 15"  -> Comma separated lists allowed.
        - OVERRIDE: Click the button to set absolute forced dates.

        4. CALENDAR NEEDS (Tab 2)
        -------------------------
        - Define how many doctors are needed for every AM and PM slot.
        - 0 Needs = Clinic Closed (Red background).
        
        5. SAVING
        ---------
        - The list of physicians saves automatically to a local file.
        - You can Export to Excel to share the schedule.
        - The Export also saves your settings, so you can Import that Excel later 
          to restore the exact roster configuration.
        """
        txt.insert("0.0", info)
        txt.configure(state="disabled")

# --- Main Block ---
if __name__ == "__main__":
    app = AppUI()
    app.mainloop()