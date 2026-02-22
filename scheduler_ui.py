import customtkinter as ctk
import os
from tkinter import filedialog, messagebox
import datetime
import calendar
import json
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scheduler_models import Physician
from scheduler_logic import SchedulerLogic
from scheduler_utils import APP_NAME, STATE_FILE, DEFAULT_COLORS, COLOR_PALETTE, DEFAULT_ROSTER_DATA, get_app_path

class ColorPicker(ctk.CTkToplevel):
    def __init__(self, parent, current_color, taken_colors, on_select):
        super().__init__(parent)
        self.title("Select Color")
        self.geometry("400x400")
        self.attributes("-topmost", True)
        self.on_select = on_select
        
        lbl = ctk.CTkLabel(self, text="Choose a color for this physician:", font=("Arial", 14))
        lbl.pack(pady=10)
        
        grid = ctk.CTkFrame(self)
        grid.pack(expand=True, padx=20, pady=20)
        
        palette = list(COLOR_PALETTE)
        for c in DEFAULT_COLORS.values():
            if c not in palette: palette.append(c)

        while len(palette) < len(taken_colors) + 5:
            r = lambda: random.randint(0,255)
            palette.append('#%02X%02X%02X' % (r(),r(),r()))

        r, c = 0, 0
        for color in palette:
            btn = ctk.CTkButton(grid, text="", width=40, height=40, fg_color=color, hover_color=color,
                                command=lambda col=color: self.select(col))
            btn.grid(row=r, column=c, padx=5, pady=5)
            c += 1
            if c > 4:
                c = 0
                r += 1

    def select(self, color):
        self.on_select(color)
        self.destroy()

class AppUI(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Clinic Scheduler V16")
        self.geometry("1400x900")
        
        self.physicians = []
        self.history = [] 
        self.schedule_results = None
        self.scheduler_logic = None
        
        self.tabview = ctk.CTkTabview(self)
        self.tabview.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.tab_emp = self.tabview.add("1. Physicians")
        self.tab_cal = self.tabview.add("2. Needs")
        self.tab_out = self.tabview.add("3. Schedule")
        self.tab_col = self.tabview.add("3.5 Color Schedule")
        self.tab_imp = self.tabview.add("4. Import/Export")
        self.tab_hlp = self.tabview.add("5. Instructions")
        
        self.setup_tab1()
        self.setup_tab2()
        self.setup_tab3()
        self.setup_tab_colored()
        self.setup_tab4()
        self.setup_tab5()
        
        self.load_state_from_disk()
        
    def save_snapshot(self):
        snap = [p.to_dict() for p in self.physicians]
        self.history.append(snap)
        if len(self.history) > 10: self.history.pop(0)

    def undo(self):
        if not self.history: return
        prev_state = self.history.pop()
        self.physicians = [Physician.from_dict(d) for d in prev_state]
        self.refresh_physician_list()

    def get_resource_path(self):
        return os.path.join(get_app_path(), STATE_FILE)

    def save_state_to_disk(self):
        self.update_physician_objects_from_ui()
        data = { "physicians": [p.to_dict() for p in self.physicians] }
        try:
            with open(self.get_resource_path(), 'w') as f: json.dump(data, f, indent=4)
        except: pass

    def load_state_from_disk(self):
        path = self.get_resource_path()
        if os.path.exists(path):
            try:
                with open(path, 'r') as f:
                    data = json.load(f)
                    self.physicians = [Physician.from_dict(x) for x in data.get("physicians", [])]
                self.refresh_physician_list()
                return
            except: pass
        self.reset_to_defaults()

    # --- TAB 1 ---
    def setup_tab1(self):
        # 1. LEGEND / INSTRUCTIONS
        inst_frame = ctk.CTkFrame(self.tab_emp)
        inst_frame.pack(fill="x", padx=5, pady=5)
        
        inst_frame.grid_columnconfigure(0, weight=1)
        inst_frame.grid_columnconfigure(1, weight=1)
        
        # Box 1: Legend (Left)
        leg_frame = ctk.CTkFrame(inst_frame)
        leg_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        
        ctk.CTkLabel(leg_frame, text="LEGEND & KEYS", font=("Arial", 12, "bold"), text_color="#87CEFA").pack(anchor="w", padx=5, pady=2)
        
        leg_box = ctk.CTkTextbox(leg_frame, height=95, font=("Arial", 11), activate_scrollbars=False, fg_color="transparent")
        leg_box.pack(fill="both", expand=True, padx=2, pady=2)
        
        leg_txt = (
            "Active: Uncheck to hide doctor from schedule.\n"
            "Target: Total shifts desired for the month.\n"
            "1st / 2nd: Restricts to early or late month.\n"
            "FullDay?: Check if AM+PM on same day is OK.\n"
            "Pref/Avoid: Request or block specific dates.\n"
            "Override: FORCE a date. Ignores all rules."
        )
        leg_box.insert("0.0", leg_txt)
        
        # Color keywords Blue (#87CEFA)
        kw_list = ["Active:", "Target:", "1st / 2nd:", "FullDay?:", "Pref/Avoid:", "Override:"]
        for kw in kw_list:
            start_idx = "1.0"
            while True:
                pos = leg_box.search(kw, start_idx, stopindex="end")
                if not pos: break
                end_pos = f"{pos}+{len(kw)}c"
                leg_box.tag_add("blue_tag", pos, end_pos)
                start_idx = end_pos
        
        leg_box.tag_config("blue_tag", foreground="#87CEFA")
        leg_box.configure(state="disabled")

        # Box 2: Suggestions (Right)
        sugg_frame = ctk.CTkFrame(inst_frame)
        sugg_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        
        ctk.CTkLabel(sugg_frame, text="SUGGESTIONS", font=("Arial", 12, "bold"), text_color="#00FF00").pack(anchor="w", padx=5, pady=2)
        
        sugg_box = ctk.CTkTextbox(sugg_frame, height=95, font=("Arial", 11), activate_scrollbars=False, fg_color="transparent")
        sugg_box.pack(fill="both", expand=True, padx=2, pady=2)
        
        sugg_txt = (
            "• Acceptable Syntax: 1, 2, 3  |  2AM, 3PM  |  1-5  |  4-8AM\n"
            "• Logic: I suggest keeping Ratio-based Logic ON.\n"
            "• Desperation Stages:\n"
            "  - Stage 0: Safe. Honors all requests. Leaves holes.\n"
            "  - Stage 1: Soft. Recommended balance.\n"
            "  - Stage 2: Hard. Ignores requests to fill slots."
        )
        sugg_box.insert("0.0", sugg_txt)
        
        # Color keywords Green (#00FF00)
        green_kws = ["Acceptable Syntax:", "Logic:", "Desperation Stages:"]
        for kw in green_kws:
            start_idx = "1.0"
            while True:
                pos = sugg_box.search(kw, start_idx, stopindex="end")
                if not pos: break
                end_pos = f"{pos}+{len(kw)}c"
                sugg_box.tag_add("green_tag", pos, end_pos)
                start_idx = end_pos
        
        sugg_box.tag_config("green_tag", foreground="#00FF00")
        sugg_box.configure(state="disabled")

        # 2. CONTROLS
        ctrl_frame = ctk.CTkFrame(self.tab_emp)
        ctrl_frame.pack(fill="x", padx=5, pady=5)
        
        ctk.CTkButton(ctrl_frame, text="Reset Defaults", width=100, fg_color="#555555", command=self.reset_to_defaults).pack(side="left", padx=5)
        ctk.CTkButton(ctrl_frame, text="Undo", width=80, fg_color="#555555", command=self.undo).pack(side="left", padx=5)
        ctk.CTkButton(ctrl_frame, text="+ Add Physician", width=120, command=self.add_blank_physician).pack(side="left", padx=5)
        
        self.lbl_total_target = ctk.CTkLabel(ctrl_frame, text="Total Target: 0", font=("Arial", 16, "bold"), text_color="#00FF00")
        self.lbl_total_target.pack(side="left", padx=20)

        self.btn_clear = ctk.CTkButton(ctrl_frame, text="Clear Dates", width=100, fg_color="#aa0000", hover_color="#880000", command=self.clear_all_dates)
        self.btn_clear.pack(side="right", padx=5)
        
        self.desp_stage = 1
        self.btn_desp = ctk.CTkButton(ctrl_frame, text="Mode: Stage 1 (Fair)", width=150, fg_color="#FFD700", text_color="black", hover_color="#D4B400", command=self.cycle_desperation)
        self.btn_desp.pack(side="right", padx=5)

        self.algo_var = ctk.StringVar(value="Proportional")
        self.switch_algo = ctk.CTkSwitch(ctrl_frame, text="Ratio-Based Logic", variable=self.algo_var, onvalue="Proportional", offvalue="Standard")
        self.switch_algo.pack(side="right", padx=10)

        # 3. HEADERS (Single Row for V13)
        self.header_frame = ctk.CTkFrame(self.tab_emp, height=35, fg_color="#333333")
        self.header_frame.pack(fill="x", padx=5, pady=(5,0))
        
        # V16 Fix: Final Shifts (1cm = ~38px)
        # Active R .25cm -> +10px to Order
        # Name R .25cm -> +10px to Active
        # Target/1st/2nd/FullDay/Pref R 1cm -> +38px to Name
        # Avoid R 1.5cm -> +57px to Preferred
        # Override R 2cm -> +76px to Avoid
        # Color R 2cm -> +76px to Override (implied grouping)
        # Del R 3cm -> +114px to Color
        self.T1_LAYOUT = [
            ("Order", 75, 0),      # +10
            ("Active", 80, 0),     # +10
            ("Name", 216, 0),      # +38
            ("Target", 50, 0),
            ("1st", 40, 0), 
            ("2nd", 50, 0),
            ("FullDay?", 70, 0),   
            ("Preferred", 207, 0), # 150 base + 57 shift
            ("Avoid", 226, 0),     # 150 base + 76 shift
            ("Override", 90, 0),   # Base 90. If Color shifts 2cm too, we don't add padding here? 
                                   # User: "Override to r by 2cm, color to r by 2cm". 
                                   # This implies they move TOGETHER away from Avoid.
                                   # So Avoid gets +76px padding.
                                   # "Del to r by 3cm". This means Del moves away from Color.
            ("Color", 164, 0),     # Base 50 + 114 padding
            ("Del", 40, 0),
            ("", 0, 1)             # SPACER
        ]

        # Configure Header Columns
        for i, (txt, w, wt) in enumerate(self.T1_LAYOUT):
            self.header_frame.grid_columnconfigure(i, weight=wt, minsize=w)
            if txt: 
                # Apply custom label shifts (1cm ~ 38px, .25cm ~ 10px)
                pad_left = 5
                if txt in ["Name", "Target", "1st", "2nd"]: pad_left = 15 
                elif txt == "Preferred": pad_left = 43                
                elif txt == "Avoid": pad_left = 62                    
                elif txt == "Override": pad_left = 15   
                elif txt == "Color": pad_left = 25         # Increased from 15 (Right .25cm)
                elif txt == "Del": pad_left = 11           # Increased from 5 (Right .15cm)
                
                lbl = ctk.CTkLabel(self.header_frame, text=txt, font=("Arial", 12, "bold"), anchor="w")
                lbl.grid(row=0, column=i, padx=(pad_left, 5), pady=5, sticky="ew")

        # 4. ROW LIST
        self.p_scroll = ctk.CTkScrollableFrame(self.tab_emp)
        self.p_scroll.pack(fill="both", expand=True, padx=5, pady=5)
        
        for i, (txt, w, wt) in enumerate(self.T1_LAYOUT):
             self.p_scroll.grid_columnconfigure(i, weight=wt, minsize=w)

        self.p_rows = []

    def reset_to_defaults(self):
        self.save_snapshot()
        self.physicians = []
        for name, target in DEFAULT_ROSTER_DATA:
            color = DEFAULT_COLORS.get(name, "#FFFFFF")
            # Default to FullDay=False for everyone EXCEPT Gandhi and Wesley
            is_fd = (name in ["Gandhi", "Wesley"])
            self.physicians.append(Physician(name, target, active=True, color=color, full_day_ok=is_fd))
        self.refresh_physician_list()

    def add_blank_physician(self):
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        self.physicians.append(Physician("New Doc", 0, color="#FFFFFF"))
        self.refresh_physician_list()

    def clear_all_dates(self):
        if not messagebox.askyesno("Confirm", "Clear dates for ALL physicians?"):
            return
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        for p in self.physicians:
            p.override_str = ""
            p.preferred_str = ""
            p.avoid_str = ""
        self.refresh_physician_list()

    def move_row(self, index, direction):
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        if direction == -1 and index > 0:
            self.physicians[index], self.physicians[index-1] = self.physicians[index-1], self.physicians[index]
        elif direction == 1 and index < len(self.physicians) - 1:
            self.physicians[index], self.physicians[index+1] = self.physicians[index+1], self.physicians[index]
        self.refresh_physician_list()

    def delete_row(self, index):
        if not messagebox.askyesno("Confirm", "Delete this physician?"): return
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        self.physicians.pop(index)
        self.refresh_physician_list()

    def open_color_picker(self, idx):
        self.update_physician_objects_from_ui()
        current_phys = self.physicians[idx]
        used_map = {p.color: p for p in self.physicians if p != current_phys}
        
        def on_color_chosen(new_color):
            current_phys.color = new_color
            self.refresh_physician_list()
            self.save_state_to_disk()

        ColorPicker(self, current_phys.color, list(used_map.keys()), on_color_chosen)

    def toggle_half_month(self, idx, mode):
        row_widgets = self.p_rows[idx]
        if mode == "1st":
            if row_widgets['chk_1st'].get():
                row_widgets['chk_2nd'].set(False)
        else:
            if row_widgets['chk_2nd'].get():
                row_widgets['chk_1st'].set(False)
        self.save_state_to_disk()

    def toggle_override(self, index, btn_ref):
        dialog = ctk.CTkInputDialog(text="Override Dates (e.g. 12 AM, 15 PM):", title="Overrides")
        res = dialog.get_input()
        if res is not None:
            self.physicians[index].override_str = res
            if res.strip():
                btn_ref.configure(text="SET", fg_color="#b30000")
            else:
                btn_ref.configure(text="Set", fg_color="transparent", border_width=1)
            self.save_state_to_disk()

    def refresh_physician_list(self):
        for widget in self.p_scroll.winfo_children(): widget.destroy()
        self.p_rows = []
        total_tgt = 0

        for i, p in enumerate(self.physicians):
            total_tgt += p.target

            f_ord = ctk.CTkFrame(self.p_scroll, fg_color="transparent")
            f_ord.grid(row=i, column=0, padx=2, pady=2)
            ctk.CTkButton(f_ord, text="▲", width=20, command=lambda x=i: self.move_row(x, -1)).pack(side="left")
            ctk.CTkButton(f_ord, text="▼", width=20, command=lambda x=i: self.move_row(x, 1)).pack(side="left")
            
            var_active = ctk.BooleanVar(value=p.active)
            chk_active = ctk.CTkCheckBox(self.p_scroll, text="", variable=var_active, width=20)
            chk_active.grid(row=i, column=1, padx=2)
            
            ent_name = ctk.CTkEntry(self.p_scroll)
            ent_name.insert(0, p.name)
            ent_name.grid(row=i, column=2, padx=2, sticky="ew")
            
            ent_tgt = ctk.CTkEntry(self.p_scroll, width=50)
            ent_tgt.insert(0, str(p.target))
            ent_tgt.grid(row=i, column=3, padx=2)
            # Bind update for total sum
            ent_tgt.bind("<FocusOut>", lambda e: self.update_physician_objects_from_ui())
            
            var_1st = ctk.BooleanVar(value=(p.half_month=="1st"))
            var_2nd = ctk.BooleanVar(value=(p.half_month=="2nd"))
            
            c1 = ctk.CTkCheckBox(self.p_scroll, text="", variable=var_1st, width=20, 
                                 command=lambda x=i: self.toggle_half_month(x, "1st"))
            c1.grid(row=i, column=4, padx=5)
            
            c2 = ctk.CTkCheckBox(self.p_scroll, text="", variable=var_2nd, width=20,
                                 command=lambda x=i: self.toggle_half_month(x, "2nd"))
            c2.grid(row=i, column=5, padx=5)

            # Full Day OK
            var_fd = ctk.BooleanVar(value=p.full_day_ok)
            chk_fd = ctk.CTkCheckBox(self.p_scroll, text="", variable=var_fd, width=20)
            chk_fd.grid(row=i, column=6, padx=5)
            
            # Increased width 1.5x (100 -> 150)
            ent_pref = ctk.CTkEntry(self.p_scroll, width=150)
            ent_pref.insert(0, p.preferred_str)
            ent_pref.grid(row=i, column=7, padx=2, sticky="ew")
            
            ent_avoid = ctk.CTkEntry(self.p_scroll, width=150)
            ent_avoid.insert(0, p.avoid_str)
            ent_avoid.grid(row=i, column=8, padx=2, sticky="ew")
            
            btn_over = ctk.CTkButton(self.p_scroll, text="SET" if p.override_str else "Set", width=50,
                                     fg_color="#b30000" if p.override_str else "transparent",
                                     border_color="red", border_width=1)
            btn_over.configure(command=lambda x=i, b=btn_over: self.toggle_override(x, b))
            btn_over.grid(row=i, column=9, padx=2)
            
            btn_col = ctk.CTkButton(self.p_scroll, text="", width=30, height=20,
                                    fg_color=p.color, border_color="gray", border_width=1,
                                    command=lambda x=i: self.open_color_picker(x))
            btn_col.grid(row=i, column=10, padx=5)

            btn_del = ctk.CTkButton(self.p_scroll, text="X", width=30, fg_color="#b30000", 
                                    command=lambda x=i: self.delete_row(x))
            btn_del.grid(row=i, column=11, padx=15)
            
            self.p_rows.append({
                "active": var_active, "name": ent_name, "target": ent_tgt,
                "chk_1st": var_1st, "chk_2nd": var_2nd, "chk_fd": var_fd,
                "pref": ent_pref, "avoid": ent_avoid, "obj_id": p.id
            })
        
        self.lbl_total_target.configure(text=f"Total Target Sum: {total_tgt}")
        self.save_state_to_disk()

    def update_physician_objects_from_ui(self):
        new_list = []
        pmap = {p.id: p for p in self.physicians}
        total_tgt = 0
        for row in self.p_rows:
            orig = pmap.get(row['obj_id'])
            name = row['name'].get()
            try: tgt = int(row['target'].get())
            except: tgt = 0
            total_tgt += tgt

            hm = "All"
            if row['chk_1st'].get(): hm = "1st"
            elif row['chk_2nd'].get(): hm = "2nd"
            
            p = Physician(
                name=name, target=tgt, active=row['active'].get(),
                half_month=hm, preferred=row['pref'].get(), avoid=row['avoid'].get(),
                full_day_ok=row['chk_fd'].get(),
                override=orig.override_str if orig else "",
                color=orig.color if orig else "#FFFFFF"
            )
            p.id = row['obj_id']
            new_list.append(p)
        self.physicians = new_list
        self.lbl_total_target.configure(text=f"Total Target Sum: {total_tgt}")

    # --- TAB 2 ---
    def setup_tab2(self):
        ctrl = ctk.CTkFrame(self.tab_cal)
        ctrl.pack(pady=10)
        now = datetime.datetime.now()
        months = [str(i) for i in range(1, 13)]
        years = [str(now.year + i) for i in range(0, 5)]
        self.var_month = ctk.StringVar(value=str(now.month))
        self.var_year = ctk.StringVar(value=str(now.year))
        
        # Split Date
        self.var_split = ctk.StringVar(value="Automatic")
        
        ctk.CTkLabel(ctrl, text="Month:").pack(side="left", padx=5)
        ctk.CTkComboBox(ctrl, values=months, variable=self.var_month, width=60, command=self.build_needs_grid).pack(side="left")
        
        ctk.CTkLabel(ctrl, text="Year:").pack(side="left", padx=5)
        ctk.CTkComboBox(ctrl, values=years, variable=self.var_year, width=70, command=self.build_needs_grid).pack(side="left")
        
        ctk.CTkLabel(ctrl, text="2nd Half Starts On:").pack(side="left", padx=10)
        ctk.CTkComboBox(ctrl, values=["Automatic", "14", "15", "16", "17"], variable=self.var_split, width=100).pack(side="left")

        ctk.CTkButton(ctrl, text="Reset to Standards", command=self.reset_needs_std).pack(side="left", padx=20)
        
        self.lbl_total_needs = ctk.CTkLabel(self.tab_cal, text="Total Needs: 0", font=("Arial", 16, "bold"), text_color="#00FF00")
        self.lbl_total_needs.pack(pady=(0, 5))

        self.needs_frame = ctk.CTkScrollableFrame(self.tab_cal)
        self.needs_frame.pack(fill="both", expand=True, padx=10, pady=10)
        self.needs_widgets = {}
        self.needs_widgets = {}
        self.build_needs_grid()
        self.update_total_needs()

    def build_needs_grid(self, _=None):
        for w in self.needs_frame.winfo_children(): w.destroy()
        self.needs_widgets = {}
        y = int(self.var_year.get())
        m = int(self.var_month.get())
        
        days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Actions"]
        for i, d in enumerate(days):
            ctk.CTkLabel(self.needs_frame, text=d, font=("Arial", 14, "bold")).grid(row=0, column=i, padx=5, pady=5)
            self.needs_frame.grid_columnconfigure(i, weight=1)
            
        cal = calendar.monthcalendar(y, m)
        r = 1
        
        DEFAULT_MAP = {
            0: (1, 0), 1: (2, 0), 2: (0, 1), 3: (2, 1), 4: (1, 0), 5: (0, 0), 6: (0, 0)
        }

        for week_idx, week in enumerate(cal):
            week_widgets = []
            
            for day_idx, day_num in enumerate(week):
                if day_num == 0:
                    ctk.CTkLabel(self.needs_frame, text="").grid(row=r, column=day_idx)
                    continue
                
                is_weekend = (day_idx >= 5)
                bg_col = "#444444" if is_weekend else "transparent"
                cell = ctk.CTkFrame(self.needs_frame, border_width=1, border_color="gray", fg_color=bg_col)
                cell.grid(row=r, column=day_idx, padx=5, pady=5, sticky="nsew")
                
                ctk.CTkLabel(cell, text=str(day_num), font=("Arial", 12, "bold")).pack(pady=2)
                
                if not is_weekend:
                    def_am, def_pm = DEFAULT_MAP.get(day_idx, (0,0))
                    
                    # AM
                    f_am = ctk.CTkFrame(cell, fg_color="transparent")
                    f_am.pack(fill="x", padx=2, pady=1)
                    ctk.CTkLabel(f_am, text="AM", width=25, font=("Arial", 10)).pack(side="left")
                    cmb_am = ctk.CTkComboBox(f_am, values=["0", "1", "2"], width=50)
                    cmb_am.set(str(def_am))
                    cmb_am.pack(side="left")
                    self.color_combo(str(def_am), cmb_am)
                    cmb_am.configure(command=lambda v, w=cmb_am: self.color_combo(v, w))
                    
                    # PM
                    f_pm = ctk.CTkFrame(cell, fg_color="transparent")
                    f_pm.pack(fill="x", padx=2, pady=1)
                    ctk.CTkLabel(f_pm, text="PM", width=25, font=("Arial", 10)).pack(side="left")
                    cmb_pm = ctk.CTkComboBox(f_pm, values=["0", "1", "2"], width=50)
                    cmb_pm.set(str(def_pm))
                    cmb_pm.pack(side="left")
                    self.color_combo(str(def_pm), cmb_pm)
                    cmb_pm.configure(command=lambda v, w=cmb_pm: self.color_combo(v, w))
                    
                    self.needs_widgets[(day_num, 'AM')] = cmb_am
                    self.needs_widgets[(day_num, 'PM')] = cmb_pm
                    week_widgets.append(cmb_am)
                    week_widgets.append(cmb_pm)
                else:
                    ctk.CTkLabel(cell, text="Weekend", text_color="gray").pack()

            btn_close = ctk.CTkButton(self.needs_frame, text="Close\nClinic", width=70, fg_color="#800000",
                                      command=lambda w=week_widgets: self.close_week(w))
            btn_close.grid(row=r, column=7, padx=5)
            r += 1

    def color_combo(self, val, widget):
        if val == "0": widget.configure(fg_color="#550000", button_color="#550000")
        else: widget.configure(fg_color="#1f6aa5", button_color="#1f6aa5")
        self.update_total_needs()

    def cycle_desperation(self):
        self.desp_stage = (self.desp_stage + 1) % 3
        
        if self.desp_stage == 0:
            self.btn_desp.configure(text="Desperation: OFF (Safe)", fg_color="#2CC985", text_color="black", hover_color="#22AA70")
        elif self.desp_stage == 1:
            self.btn_desp.configure(text="Desperation: Stage 1 (Soft)", fg_color="#FFD700", text_color="black", hover_color="#D4B400")
        elif self.desp_stage == 2:
            self.btn_desp.configure(text="Desperation: Stage 2 (Hard)", fg_color="#FF8C00", text_color="black", hover_color="#CC7000")

    def reset_needs_std(self):
        self.build_needs_grid()

    def close_week(self, widgets):
        for w in widgets:
            w.set("0")
            self.color_combo("0", w)
        self.update_total_needs()

    def update_total_needs(self):
        total = 0
        for w in self.needs_widgets.values():
            try: total += int(w.get())
            except: pass
        if hasattr(self, 'lbl_total_needs'):
            self.lbl_total_needs.configure(text=f"Total Physician Shifts Needed: {total}")

    def get_needs_data(self):
        data = {}
        for (day, slot), widget in self.needs_widgets.items():
            if day not in data: data[day] = {}
            try: data[day][slot] = int(widget.get())
            except: data[day][slot] = 0
        return data

    # --- TAB 3 & 3.5 ---
    def setup_tab3(self):
        self.setup_sched_tab(self.tab_out, colored_text=False)
    
    def setup_tab_colored(self):
        self.setup_sched_tab(self.tab_col, colored_text=True)

    def setup_sched_tab(self, parent_tab, colored_text):
        top = ctk.CTkFrame(parent_tab)
        top.pack(fill="x", padx=10, pady=5)
        
        btn = ctk.CTkButton(top, text="REGENERATE SCHEDULE", font=("Arial", 14, "bold"), 
                            fg_color="#006400", height=40, command=self.generate_schedule)
        btn.pack(fill="x")
        
        # Calendar Area - Full Width to remove "unused column" appearance
        container = ctk.CTkFrame(parent_tab, fg_color="transparent")
        container.place(relx=0.01, rely=0.05, relwidth=0.98, relheight=0.90)

        # Frozen Headers INSIDE Container (Row 0)
        # ... (headers recreation logic can be simplified if already correct locally, but ensuring strict packing)
        
        # Left Panel (Calendar + Headers)
        left_panel = ctk.CTkFrame(container, fg_color="transparent")
        left_panel.pack(side="left", fill="both", expand=True)

        # Headers
        headers_frame = ctk.CTkFrame(left_panel, height=30, fg_color="transparent")
        headers_frame.pack(fill="x", pady=(0, 2))
        for i in range(5): headers_frame.grid_columnconfigure(i, weight=1, uniform="day_col")
        days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
        for i, d in enumerate(days):
            ctk.CTkLabel(headers_frame, text=d, font=("Arial", 14, "bold")).grid(row=0, column=i, sticky="ew")

        # Scrollable Calendar
        cal_frame = ctk.CTkScrollableFrame(left_panel)
        cal_frame.pack(fill="both", expand=True)
        for i in range(5): cal_frame.grid_columnconfigure(i, weight=1, uniform="day_col")

        if not colored_text: self.out_cal_frame = cal_frame
        else: self.col_cal_frame = cal_frame

        # Right Panel (Stats)
        sidebar = ctk.CTkFrame(container, width=300) 
        sidebar.pack(side="right", fill="y", padx=(10, 0))
        
        ctk.CTkLabel(sidebar, text="Schedule Statistics", font=("Arial", 16, "bold")).pack(pady=5)
        
        # Stats Grid
        stats_grid_frame = ctk.CTkFrame(sidebar, fg_color="transparent")
        stats_grid_frame.pack(fill="x", padx=5)
        
        headers = ["Name", "Target", "Actual", "Net"]
        for i, h in enumerate(headers):
            sticky = "w" if i == 0 else "e"
            ctk.CTkLabel(stats_grid_frame, text=h, font=("Arial", 12, "bold")).grid(row=0, column=i, padx=5, pady=2, sticky=sticky)
            if i > 0: stats_grid_frame.grid_columnconfigure(i, weight=1)
            
        ctk.CTkLabel(sidebar, text="Issues / Deviations", font=("Arial", 14, "bold"), text_color="orange").pack(pady=(10, 0))
        warnings_text = ctk.CTkTextbox(sidebar, height=200, text_color="orange", font=("Consolas", 12))
        warnings_text.pack(fill="both", expand=True, padx=5, pady=5)
        
        stats_label = ctk.CTkLabel(top, text="Status: Waiting", text_color="gray")
        stats_label.pack(pady=5)
        
        # Store refs appropriately so they are unique per tab
        if not colored_text:
            self.stats_grid = stats_grid_frame
            self.warnings_text = warnings_text
            self.stats_label = stats_label
        else:
            self.stats_grid_col = stats_grid_frame
            self.warnings_text_col = warnings_text
            self.stats_label_col = stats_label

    def generate_schedule(self):
        self.update_physician_objects_from_ui()
        self.save_state_to_disk()
        y = int(self.var_year.get())
        m = int(self.var_month.get())
        
        # Logic for Automatic Split
        val = self.var_split.get()
        if val == "Automatic":
            split = 16 
        else:
            try: split = int(val)
            except: split = 16

        needs = self.get_needs_data()
        
        # Auto-Retry Logic (Robustness)
        best_logic = None
        min_open_slots = float('inf')
        
        algo_mode = self.algo_var.get() if hasattr(self, 'algo_var') else "Standard"

        # Try up to 10 times to find a perfect schedule
        for attempt in range(10):
            temp_logic = SchedulerLogic(self.physicians, y, m, needs, split_day=split)
            temp_logic.run(algorithm=algo_mode, desperation_stage=self.desp_stage)
            
            if not temp_logic.has_open_slots():
                best_logic = temp_logic
                print(f"Assigner success on attempt {attempt+1}")
                break
            else:
                # Count open slots to find 'best fit' if we fail all retries
                open_cnt = 0
                for d in range(1, temp_logic.last_day+1):
                     if d in temp_logic.daily_needs:
                         for st in ['AM', 'PM']:
                             nd = temp_logic.daily_needs[d].get(st, 0)
                             act = len(temp_logic.schedule[d][st])
                             if act < nd: open_cnt += (nd - act)
                
                if open_cnt < min_open_slots:
                    min_open_slots = open_cnt
                    best_logic = temp_logic
        
        self.scheduler_logic = best_logic
        
        self.render_calendar_logic(self.out_cal_frame, best_logic, use_colors=False)
        self.render_calendar_logic(self.col_cal_frame, best_logic, use_colors=True)
        
        # Render Stats for BOTH tabs
        self.render_stats(best_logic, self.stats_grid, self.stats_label, self.warnings_text)
        self.render_stats(best_logic, self.stats_grid_col, self.stats_label_col, self.warnings_text_col)

    def render_calendar_logic(self, frame, logic, use_colors):
        for w in frame.winfo_children(): w.destroy()
        
        y, m = logic.year, logic.month
        cal = calendar.monthcalendar(y, m)
        r = 0
        c_map = {p.name: p.color for p in logic.physicians}

        # V8: Uniform Grid Config
        for i in range(5):
            frame.grid_columnconfigure(i, weight=1, uniform="day_col")

        for week in cal:
            has_weekday = False
            for day_idx, day_num in enumerate(week):
                if day_idx >= 5: continue 
                has_weekday = True
                
                # DAY CELL
                if day_num == 0:
                    # Empty filler
                    ctk.CTkFrame(frame, fg_color="transparent", border_width=0).grid(row=r, column=day_idx, sticky="nsew", padx=2, pady=2)
                    continue

                # V18 Fix: Shorter Cells (Height=120) with NO Propagation
                cell = ctk.CTkFrame(frame, border_width=1, border_color="#555555", height=120)
                cell.grid(row=r, column=day_idx, sticky="nsew", padx=2, pady=2)
                cell.pack_propagate(False) # Force height
                cell.grid_propagate(False) # Force height
                
                # V17 Fix: 50/50 Split (Rows 1 and 2 share weight uniformly)
                cell.grid_rowconfigure(0, weight=0, minsize=20)   # Header row fixed small
                cell.grid_rowconfigure(1, weight=1, uniform="split") # AM
                cell.grid_rowconfigure(2, weight=1, uniform="split") # PM
                cell.grid_columnconfigure(0, weight=1)

                # Header for Day Number (Blue, No "D")
                # Removed "ne" anchor and packed? No, use Grid for row 0.
                ctk.CTkLabel(cell, text=f"{day_num}", font=("Arial", 10, "bold"), text_color="#1E90FF").grid(row=0, column=0, sticky="ne", padx=2, pady=0)
                
                day_sched = logic.schedule.get(day_num, {'AM': [], 'PM': []})
                needs_am = logic.daily_needs.get(day_num, {}).get('AM', 0)
                needs_pm = logic.daily_needs.get(day_num, {}).get('PM', 0)
                
                # --- AM Section (Row 1) ---
                bg_am = "#999999" 
                f_am = ctk.CTkFrame(cell, fg_color=bg_am, corner_radius=0)
                f_am.grid(row=1, column=0, sticky="nsew", padx=0, pady=0)
                
                if needs_am > 0:
                    ctk.CTkLabel(f_am, text="AM", font=("Arial", 8, "bold"), text_color="#333333").pack(anchor="n", pady=0) 
                    
                    # Name Container (Centered)
                    name_row = ctk.CTkFrame(f_am, fg_color="transparent")
                    name_row.pack(anchor="center", pady=0)
                    
                    for i, name in enumerate(day_sched['AM']):
                        # Comma?
                        if i > 0:
                            ctk.CTkLabel(name_row, text=", ", text_color="#333333", font=("Arial", 10, "bold")).pack(side="left", padx=0)
                        
                        # Color Logic
                        t_col = "#000055"
                        if use_colors: t_col = c_map.get(name, "black")
                        if hasattr(logic, 'desperation_assignments') and (day_num, 'AM', name) in logic.desperation_assignments:
                             t_col = "#FFD700" # GOLD
                        
                        ctk.CTkLabel(name_row, text=name, text_color=t_col, font=("Arial", 10, "bold")).pack(side="left", padx=0)

                    if len(day_sched['AM']) < needs_am:
                        # Append OPEN if needed
                        open_txt = "[OPEN]" if len(day_sched['AM']) == 0 else ", [OPEN]"
                        ctk.CTkLabel(name_row, text=open_txt, text_color="#aa0000", font=("Arial", 9, "bold")).pack(side="left", padx=0)

                else:
                    ctk.CTkLabel(f_am, text="CLOSED", text_color="#555555", font=("Arial", 8)).place(relx=0.5, rely=0.5, anchor="center")

                # --- PM Section (Row 2) ---
                bg_pm = "#555555" 
                f_pm = ctk.CTkFrame(cell, fg_color=bg_pm, corner_radius=0)
                f_pm.grid(row=2, column=0, sticky="nsew", padx=0, pady=0)

                if needs_pm > 0:
                    ctk.CTkLabel(f_pm, text="PM", font=("Arial", 8, "bold"), text_color="#BBBBBB").pack(anchor="n", pady=0)
                    
                    # Name Container (Centered)
                    name_row = ctk.CTkFrame(f_pm, fg_color="transparent")
                    name_row.pack(anchor="center", pady=0)

                    for i, name in enumerate(day_sched['PM']):
                        # Comma?
                        if i > 0:
                             ctk.CTkLabel(name_row, text=", ", text_color="#BBBBBB", font=("Arial", 10, "bold")).pack(side="left", padx=0)
                        
                        # Color Logic
                        t_col = "#C0C0C0" # SILVER
                        if use_colors: t_col = c_map.get(name, "white")
                        if hasattr(logic, 'desperation_assignments') and (day_num, 'PM', name) in logic.desperation_assignments:
                             t_col = "#FFD700" # GOLD

                        ctk.CTkLabel(name_row, text=name, text_color=t_col, font=("Arial", 10, "bold")).pack(side="left", padx=0)

                    if len(day_sched['PM']) < needs_pm:
                         open_txt = "[OPEN]" if len(day_sched['PM']) == 0 else ", [OPEN]"
                         ctk.CTkLabel(name_row, text=open_txt, text_color="#FF5555", font=("Arial", 9, "bold")).pack(side="left", padx=0)
                else:
                    ctk.CTkLabel(f_pm, text="CLOSED", text_color="#333333").place(relx=0.5, rely=0.5, anchor="center")

            # Move to next row after week is done
            if has_weekday: r += 1

    def render_stats(self, logic, grid_frame, status_lbl, warn_box):
        for w in grid_frame.winfo_children():
            if int(w.grid_info()["row"]) > 0: w.destroy()
        
        r = 1
        total_slots = 0
        filled_slots = 0
        for d, reqs in logic.daily_needs.items():
            total_slots += reqs.get('AM', 0) + reqs.get('PM', 0)
            filled_slots += len(logic.schedule.get(d, {}).get('AM', [])) + len(logic.schedule.get(d, {}).get('PM', []))
        
        if filled_slots < total_slots:
            status_lbl.configure(text=f"MISSING SHIFTS: {filled_slots}/{total_slots}", text_color="red")
        else:
            status_lbl.configure(text="ALL CLINIC SHIFTS FILLED ✔", text_color="#00ff00")

        for p in logic.physicians:
            act = len(p.assigned_shifts)
            net = act - p.target
            
            if net > 0: net_col = "yellow"; net_str = f"+{net}"
            elif net < 0: net_col = "#00BFFF"; net_str = f"{net}" 
            else: net_col = "#00FF00"; net_str = "0" 
            
            ctk.CTkLabel(grid_frame, text=p.name).grid(row=r, column=0, sticky="w", padx=5)
            ctk.CTkLabel(grid_frame, text=str(p.target)).grid(row=r, column=1, sticky="e", padx=5)
            ctk.CTkLabel(grid_frame, text=str(act)).grid(row=r, column=2, sticky="e", padx=5)
            ctk.CTkLabel(grid_frame, text=net_str, text_color=net_col).grid(row=r, column=3, sticky="e", padx=5)
            r += 1

        warn_box.delete("0.0", "end")
        if not logic.warnings:
            warn_box.insert("end", "No issues or deviations found.")
        else:
            for w in logic.warnings:
                warn_box.insert("end", f"- {w}\n")

    # --- TAB 4 ---
    def setup_tab4(self):
        f = ctk.CTkFrame(self.tab_imp)
        f.pack(expand=True)
        ctk.CTkLabel(f, text="Excel Operations", font=("Arial", 20)).pack(pady=20)
        ctk.CTkButton(f, text="Export Schedule (Template Format)", width=250, height=50, command=self.export_excel).pack(pady=10)
        ctk.CTkButton(f, text="Import State from Excel", width=250, height=50, command=self.import_excel).pack(pady=10)

    def export_excel(self):
        if not self.scheduler_logic: 
            messagebox.showwarning("Warning", "Please generate a schedule first (Tab 3).")
            return
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not filename: return
        
        try:
            wb = openpyxl.Workbook()
            # SHEET 1: Printable Template
            ws = wb.active
            ws.title = f"Schedule {self.var_month.get()}-{self.var_year.get()}"
            
            # --- STYLES ---
            title_font = Font(name='Calibri', size=16, bold=True)
            bold_font = Font(name='Calibri', size=11, bold=True)
            clinic_lbl_font = Font(name='Calibri', size=11, bold=True)
            
            center_align = Alignment(horizontal="center", vertical="center")
            
            fill_black = PatternFill("solid", fgColor="000000")
            
            # Map physician colors (Strip to fix "name " space mismatch)
            c_map = {p.name.strip(): p.color.replace("#","") for p in self.scheduler_logic.physicians}
            
            # --- LAYOUT SETUP ---
            ws.column_dimensions['A'].width = 3 
            
            col_width = 20
            gps = ['B', 'C', 'D', 'E', 'F']
            gps2 = ['H', 'I', 'J', 'K', 'L']
            
            for c in gps + gps2:
                ws.column_dimensions[c].width = col_width
                
            ws.column_dimensions['G'].width = 12 
            # MOVED: Black fill applied at end to match calendar height

            # --- HEADERS ---
            ws.merge_cells("B1:F1")
            ws["B1"] = f"Outpatient Center-LBJ Group A {calendar.month_name[int(self.var_month.get())]} {self.var_year.get()}"
            ws["B1"].font = title_font
            ws["B1"].alignment = center_align
            
            ws.merge_cells("H1:L1")
            ws["H1"] = f"Outpatient Center-LBJ Group B {calendar.month_name[int(self.var_month.get())]} {self.var_year.get()}"
            ws["H1"].font = title_font
            ws["H1"].alignment = center_align

            # Day Names
            days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
            for i, d in enumerate(days):
                col_let = get_column_letter(i+2) # Start B
                ws[f"{col_let}3"] = d
                ws[f"{col_let}3"].font = bold_font
                ws[f"{col_let}3"].alignment = center_align
                
                col_let_b = get_column_letter(i+8) # Start H
                ws[f"{col_let_b}3"] = d
                ws[f"{col_let_b}3"].font = bold_font
                ws[f"{col_let_b}3"].alignment = center_align

            logic = self.scheduler_logic
            cal = calendar.monthcalendar(logic.year, logic.month)
            current_row = 4
            
            for week in cal:
                has_weekday = any(d != 0 and idx < 5 for idx, d in enumerate(week))
                if not has_weekday: continue
                
                start_row_of_block = current_row
                
                for day_idx in range(5):
                    day_num = week[day_idx]
                    if day_num == 0: continue
                    
                    col_a = day_idx + 2
                    col_b = day_idx + 8
                    
                    ws.cell(row=current_row, column=col_a, value=day_num).font = bold_font
                    ws.cell(row=current_row, column=col_a).alignment = center_align
                    ws.cell(row=current_row, column=col_b, value=day_num).font = bold_font
                    ws.cell(row=current_row, column=col_b).alignment = center_align
                    
                    sched = logic.schedule.get(day_num, {'AM': [], 'PM': []})
                    
                    # --- AM ---
                    am_docs = sched['AM']
                    if am_docs:
                        ws.cell(row=current_row+1, column=col_a, value="AM CLINIC").font = clinic_lbl_font
                        ws.cell(row=current_row+1, column=col_b, value="AM CLINIC").font = clinic_lbl_font
                        
                        primary_col = col_a
                        secondary_col = col_b
                        
                        if day_idx == 0 or day_idx == 4: # Mon, Fri
                            primary_col = col_b
                            secondary_col = col_a
                        
                        doc1 = am_docs[0]
                        ws.cell(row=current_row+2, column=primary_col, value=doc1).font = bold_font
                        if doc1.strip() in c_map:
                             ws.cell(row=current_row+2, column=primary_col).fill = PatternFill("solid", fgColor=c_map[doc1.strip()])
                        
                        overflow = am_docs[1:]
                        if overflow:
                            txt = " / ".join(overflow)
                            existing = ws.cell(row=current_row+2, column=secondary_col).value
                            ws.cell(row=current_row+2, column=secondary_col, value=f"{existing} / {txt}" if existing else txt)
                            ws.cell(row=current_row+2, column=secondary_col).font = bold_font
                            # Fix 1: Apply color for 2nd attending
                            first_overflow = overflow[0].strip()
                            if first_overflow in c_map:
                                 ws.cell(row=current_row+2, column=secondary_col).fill = PatternFill("solid", fgColor=c_map[first_overflow])

                    # --- PM ---
                    pm_docs = sched['PM']
                    if pm_docs:
                        ws.cell(row=current_row+7, column=col_a, value="PM CLINIC").font = clinic_lbl_font
                        ws.cell(row=current_row+7, column=col_b, value="PM CLINIC").font = clinic_lbl_font
                    
                        primary_col = col_b
                        secondary_col = col_a
                        
                        if day_idx == 2 or day_idx == 3: # Wed, Thu
                             primary_col = col_a
                             secondary_col = col_b
                        
                        doc1 = pm_docs[0]
                        ws.cell(row=current_row+8, column=primary_col, value=doc1).font = bold_font
                        if doc1.strip() in c_map:
                             ws.cell(row=current_row+8, column=primary_col).fill = PatternFill("solid", fgColor=c_map[doc1.strip()])
                        
                        overflow = pm_docs[1:]
                        if overflow:
                            txt = " / ".join(overflow)
                            existing = ws.cell(row=current_row+8, column=secondary_col).value
                            ws.cell(row=current_row+8, column=secondary_col, value=f"{existing} / {txt}" if existing else txt)
                            ws.cell(row=current_row+8, column=secondary_col).font = bold_font
                            # Fix 1: Apply color for 2nd attending PM
                            first_overflow = overflow[0].strip()
                            if first_overflow in c_map:
                                 ws.cell(row=current_row+8, column=secondary_col).fill = PatternFill("solid", fgColor=c_map[first_overflow])

                # --- BORDERS ---
                for day_idx in range(5):
                    if week[day_idx] == 0: continue
                    
                    col_a_idx = day_idx + 2
                    col_b_idx = day_idx + 8
                    
                    for r_off in range(13):
                        r = start_row_of_block + r_off
                        cell = ws.cell(row=r, column=col_a_idx)
                        
                        l = Side(style='medium')
                        r_side = Side(style='medium')
                        t = Side(style='medium') if r_off == 0 else None
                        b = Side(style='medium') if r_off == 12 else None
                        
                        # Fix 2: Date Cell Borders (row 0) - Add Bottom Border
                        if r_off == 0:
                            b = Side(style='medium')
                            
                        # Fix 3: Thin Grey Border above PM (Row 7 is PM Header)
                        # So Row 6 is empty space before PM? Or Row 7 *is* PM line?
                        # user: "use a thin grey border just above PM clinic"
                        # PM Clinic text is at r_off = 7. So top of 7.
                        if r_off == 7:
                            t = Side(style='thin', color="808080")
                        
                        cell.border = Border(left=l, right=r_side, top=t, bottom=b)
                        
                    for r_off in range(13):
                        r = start_row_of_block + r_off
                        cell = ws.cell(row=r, column=col_b_idx)
                        
                        l = Side(style='medium')
                        r_side = Side(style='medium')
                        t = Side(style='medium') if r_off == 0 else None
                        b = Side(style='medium') if r_off == 12 else None
                        
                        # Fix 2: Date Cell Borders
                        if r_off == 0:
                            b = Side(style='medium')

                        # Fix 3: Thin Grey Border above PM
                        if r_off == 7:
                            t = Side(style='thin', color="808080")
                        
                        cell.border = Border(left=l, right=r_side, top=t, bottom=b)

                current_row += 13

            # --- SAVE CONFIG ---
            
            # Apply Divider Black Fill (Dynamic Height)
            for r in range(1, current_row):
                ws[f"G{r}"].fill = fill_black
                
            self.update_physician_objects_from_ui()
            data_list = [p.to_dict() for p in self.physicians]
            json_str = json.dumps(data_list)
            ws["A80"] = json_str
            ws["A80"].font = Font(color="FFFFFF") 
            ws["A80"].alignment = Alignment(wrap_text=False) # Clipped

            # SHEET 2: Visual Calendar
            ws_cal = wb.create_sheet("Calendar & Settings")
            ws_cal.column_dimensions['A'].width = 15
            ws_cal.column_dimensions['B'].width = 15
            ws_cal.column_dimensions['C'].width = 15
            ws_cal.column_dimensions['D'].width = 15
            ws_cal.column_dimensions['E'].width = 15
            
            ws_cal["A1"] = f"Visual Schedule: {self.var_month.get()}-{self.var_year.get()}"
            ws_cal["A1"].font = Font(bold=True, size=16)
            
            days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
            for i, d in enumerate(days):
                c = ws_cal.cell(row=2, column=i+1, value=d)
                c.font = bold_font
                c.alignment = center_align
                c.fill = PatternFill("solid", fgColor="DDDDDD")
            
            curr_row = 3
            for week in cal:
                has_weekday = any(d != 0 and idx < 5 for idx, d in enumerate(week))
                if not has_weekday: continue
                
                for day_idx in range(5):
                    day_num = week[day_idx]
                    cell = ws_cal.cell(row=curr_row, column=day_idx+1)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    if day_num == 0: continue
                    
                    sched = logic.schedule.get(day_num, {'AM': [], 'PM': []})
                    txt = f"{day_num}\n\nAM: {', '.join(sched['AM'])}\n\nPM: {', '.join(sched['PM'])}"
                    cell.value = txt
                curr_row += 1
            
            ws_cal["A80"] = json_str
            
            wb.save(filename)
            messagebox.showinfo("Success", "Export complete.")
            
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def import_excel(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not filename: return
        try:
            wb = openpyxl.load_workbook(filename)
            # Look for settings in 'Calendar & Settings' cell A80, or legacy 'Settings' A1
            json_str = None
            if "Calendar & Settings" in wb.sheetnames:
                json_str = wb["Calendar & Settings"]["A80"].value
            elif "Settings" in wb.sheetnames:
                json_str = wb["Settings"]["A1"].value
            
            if json_str:
                data = json.loads(json_str)
                self.save_snapshot()
                self.physicians = [Physician.from_dict(d) for d in data]
                self.refresh_physician_list()
                messagebox.showinfo("Success", "Restored.")
            else:
                messagebox.showerror("Error", "No settings found in A80 of Sheet 2.")
        except Exception as e: messagebox.showerror("Error", str(e))

    def export_to_excel(self):
        try:
            filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
            if not filename: return
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Schedule"
            
            # --- STYLES ---
            bold_font = Font(bold=True, size=11)
            title_font = Font(bold=True, size=14)
            clinic_lbl_font = Font(size=11, bold=True)
            center_align = Alignment(horizontal='center', vertical='center')
            fill_black = PatternFill("solid", fgColor="000000")
            c_map = {p.name.strip(): p.color.replace("#","") for p in self.scheduler_logic.physicians}

            # --- LAYOUT SETUP ---
            ws.column_dimensions['A'].width = 3 
            
            col_width = 20
            gps = ['B', 'C', 'D', 'E', 'F']
            gps2 = ['H', 'I', 'J', 'K', 'L']
            
            for c in gps + gps2:
                ws.column_dimensions[c].width = col_width
                
            ws.column_dimensions['G'].width = 12 

            # --- HEADERS ---
            ws.merge_cells("B1:F1")
            ws["B1"] = f"Outpatient Center-LBJ Group A {calendar.month_name[int(self.var_month.get())]} {self.var_year.get()}"
            ws["B1"].font = title_font
            ws["B1"].alignment = center_align
            
            ws.merge_cells("H1:L1")
            ws["H1"] = f"Outpatient Center-LBJ Group B {calendar.month_name[int(self.var_month.get())]} {self.var_year.get()}"
            ws["H1"].font = title_font
            ws["H1"].alignment = center_align

            # Day Names
            days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
            for i, d in enumerate(days):
                col_let = get_column_letter(i+2) # Start B
                ws[f"{col_let}3"] = d
                ws[f"{col_let}3"].font = bold_font
                ws[f"{col_let}3"].alignment = center_align
                
                col_let_b = get_column_letter(i+8) # Start H
                ws[f"{col_let_b}3"] = d
                ws[f"{col_let_b}3"].font = bold_font
                ws[f"{col_let_b}3"].alignment = center_align

            logic = self.scheduler_logic
            cal = calendar.monthcalendar(logic.year, logic.month)
            current_row = 4
            
            for week in cal:
                has_weekday = any(d != 0 and idx < 5 for idx, d in enumerate(week))
                if not has_weekday: continue
                
                start_row_of_block = current_row
                
                for day_idx in range(5):
                    day_num = week[day_idx]
                    if day_num == 0: continue
                    
                    col_a = day_idx + 2
                    col_b = day_idx + 8
                    
                    ws.cell(row=current_row, column=col_a, value=day_num).font = bold_font
                    ws.cell(row=current_row, column=col_a).alignment = center_align
                    ws.cell(row=current_row, column=col_b, value=day_num).font = bold_font
                    ws.cell(row=current_row, column=col_b).alignment = center_align
                    
                    sched = logic.schedule.get(day_num, {'AM': [], 'PM': []})
                    
                    # --- Helper to split docs ---
                    def get_group_assignments(day_idx, docs, is_pm):
                        # Returns (GroupA_Docs, GroupB_Docs)
                        if not docs: return ([], [])
                        
                        gA = []
                        gB = []
                        
                        # Logic:
                        # AM: Mon(0)/Fri(4) -> Primary=B, Sec=A. Else Primary=A, Sec=B.
                        # PM: Wed(2)/Thu(3) -> Primary=A, Sec=B. Else Primary=B, Sec=A.
                        
                        # Primary: The first doctor in the list.
                        # Secondary: The overflow (rest of the list).
                        
                        primary = docs[0]
                        secondary = docs[1:]
                        
                        primary_is_A = True # Default
                        
                        if not is_pm:
                            # AM Logic
                            if day_idx == 0 or day_idx == 4: # Mon, Fri
                                primary_is_A = False # Primary is B
                        else:
                            # PM Logic
                            # Code says: if Wed/Thu: Primary=A. Else Primary=B.
                            if day_idx == 2 or day_idx == 3: # Wed, Thu
                                primary_is_A = True
                            else:
                                primary_is_A = False
                        
                        if primary_is_A:
                            gA.append(primary)
                            gB.extend(secondary)
                        else:
                            gB.append(primary)
                            gA.extend(secondary)
                            
                        return gA, gB

                    # --- AM ---
                    am_A, am_B = get_group_assignments(day_idx, sched['AM'], is_pm=False)
                    
                    if am_A:
                        ws.cell(row=current_row+1, column=col_a, value="AM CLINIC").font = clinic_lbl_font
                        # Render A
                        doc1 = am_A[0]
                        ws.cell(row=current_row+2, column=col_a, value=doc1).font = bold_font
                        if doc1.strip() in c_map:
                             ws.cell(row=current_row+2, column=col_a).fill = PatternFill("solid", fgColor=c_map[doc1.strip()])
                        
                        overflow = am_A[1:]
                        if overflow:
                             # Should effectively not happen if logic is 1 per group?
                             # But if overflow logic allows multiple in A (not typical for this specific splitting, but possible if > 2 docs total)
                             # Original code put all overflow in secondary slot.
                             # Here, gA.extend(secondary) puts ALL overflow in A if Primary is B? 
                             # Wait, standard logic: Primary=1 doc. Secondary=All others.
                             # So if Primary=B, then gB has 1, gA has Rest.
                             # If gA has multiple, they are slash separated.
                             txt = " / ".join(overflow)
                             existing = ws.cell(row=current_row+2, column=col_a).value
                             ws.cell(row=current_row+2, column=col_a, value=f"{existing} / {txt}" if existing else txt)
                             
                             # Color for 2nd doc in same cell? Not possible with cell fill. 
                             # Original logic colored the CELL based on the primary occupant of that cell.
                             # If A has 2 people (overflow), we use color of first person in A.
                             pass

                    if am_B:
                        ws.cell(row=current_row+1, column=col_b, value="AM CLINIC").font = clinic_lbl_font
                        # Render B
                        doc1 = am_B[0]
                        ws.cell(row=current_row+2, column=col_b, value=doc1).font = bold_font
                        if doc1.strip() in c_map:
                             ws.cell(row=current_row+2, column=col_b).fill = PatternFill("solid", fgColor=c_map[doc1.strip()])
                        
                        overflow = am_B[1:]
                        if overflow:
                             txt = " / ".join(overflow)
                             existing = ws.cell(row=current_row+2, column=col_b).value
                             ws.cell(row=current_row+2, column=col_b, value=f"{existing} / {txt}" if existing else txt)

                    # --- PM ---
                    pm_A, pm_B = get_group_assignments(day_idx, sched['PM'], is_pm=True)
                    
                    if pm_A:
                        ws.cell(row=current_row+7, column=col_a, value="PM CLINIC").font = clinic_lbl_font
                        doc1 = pm_A[0]
                        ws.cell(row=current_row+8, column=col_a, value=doc1).font = bold_font
                        if doc1.strip() in c_map:
                             ws.cell(row=current_row+8, column=col_a).fill = PatternFill("solid", fgColor=c_map[doc1.strip()])
                        overflow = pm_A[1:]
                        if overflow:
                             txt = " / ".join(overflow)
                             existing = ws.cell(row=current_row+8, column=col_a).value
                             ws.cell(row=current_row+8, column=col_a, value=f"{existing} / {txt}" if existing else txt)

                    if pm_B:
                        ws.cell(row=current_row+7, column=col_b, value="PM CLINIC").font = clinic_lbl_font
                        doc1 = pm_B[0]
                        ws.cell(row=current_row+8, column=col_b, value=doc1).font = bold_font
                        if doc1.strip() in c_map:
                             ws.cell(row=current_row+8, column=col_b).fill = PatternFill("solid", fgColor=c_map[doc1.strip()])
                        overflow = pm_B[1:]
                        if overflow:
                             txt = " / ".join(overflow)
                             existing = ws.cell(row=current_row+8, column=col_b).value
                             ws.cell(row=current_row+8, column=col_b, value=f"{existing} / {txt}" if existing else txt)

                # --- BORDERS ---
                for day_idx in range(5):
                    if week[day_idx] == 0: continue
                    
                    col_a_idx = day_idx + 2
                    col_b_idx = day_idx + 8
                    
                    for r_off in range(13):
                        r = start_row_of_block + r_off
                        cell = ws.cell(row=r, column=col_a_idx)
                        
                        l = Side(style='medium')
                        r_side = Side(style='medium')
                        t = Side(style='medium') if r_off == 0 else None
                        b = Side(style='medium') if r_off == 12 else None
                        
                        # Fix 2: Date Cell Borders (row 0) - Add Bottom Border
                        if r_off == 0:
                            b = Side(style='medium')
                            
                        # Fix 3: Thin Grey Border above PM (Row 7 is PM Header)
                        if r_off == 7:
                            t = Side(style='thin', color="808080")
                        
                        cell.border = Border(left=l, right=r_side, top=t, bottom=b)
                        
                    for r_off in range(13):
                        r = start_row_of_block + r_off
                        cell = ws.cell(row=r, column=col_b_idx)
                        
                        l = Side(style='medium')
                        r_side = Side(style='medium')
                        t = Side(style='medium') if r_off == 0 else None
                        b = Side(style='medium') if r_off == 12 else None
                        
                        # Fix 2: Date Cell Borders
                        if r_off == 0:
                            b = Side(style='medium')

                        # Fix 3: Thin Grey Border above PM
                        if r_off == 7:
                            t = Side(style='thin', color="808080")
                        
                        cell.border = Border(left=l, right=r_side, top=t, bottom=b)

                current_row += 13

            # --- SAVE CONFIG ---
            
            # Apply Divider Black Fill (Dynamic Height)
            for r in range(1, current_row):
                ws[f"G{r}"].fill = fill_black
                
            self.update_physician_objects_from_ui()
            data_list = [p.to_dict() for p in self.physicians]
            json_str = json.dumps(data_list)
            ws["A80"] = json_str
            ws["A80"].font = Font(color="FFFFFF") 
            ws["A80"].alignment = Alignment(wrap_text=False) # Clipped

            # SHEET 2: Visual Calendar
            ws_cal = wb.create_sheet("Calendar & Settings")
            ws_cal.column_dimensions['A'].width = 15
            ws_cal.column_dimensions['B'].width = 15
            ws_cal.column_dimensions['C'].width = 15
            ws_cal.column_dimensions['D'].width = 15
            ws_cal.column_dimensions['E'].width = 15
            
            ws_cal["A1"] = f"Visual Schedule: {self.var_month.get()}-{self.var_year.get()}"
            ws_cal["A1"].font = Font(bold=True, size=16)
            
            days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
            for i, d in enumerate(days):
                c = ws_cal.cell(row=2, column=i+1, value=d)
                c.font = bold_font
                c.alignment = center_align
                c.fill = PatternFill("solid", fgColor="DDDDDD")
            
            curr_row = 3
            for week in cal:
                has_weekday = any(d != 0 and idx < 5 for idx, d in enumerate(week))
                if not has_weekday: continue
                
                for day_idx in range(5):
                    day_num = week[day_idx]
                    cell = ws_cal.cell(row=curr_row, column=day_idx+1)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    if day_num == 0: continue
                    
                    sched = logic.schedule.get(day_num, {'AM': [], 'PM': []})
                    txt = f"{day_num}\n\nAM: {', '.join(sched['AM'])}\n\nPM: {', '.join(sched['PM'])}"
                    cell.value = txt
                curr_row += 1
            
            ws_cal["A80"] = json_str
            
            wb.save(filename)
            messagebox.showinfo("Success", "Export complete.")
            
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def export_legacy_excel(self):
        try:
            filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
            if not filename: return
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Schedule"
            
            # --- STYLES ---
            bold_font = Font(bold=True, size=11)
            title_font = Font(bold=True, size=14)
            center_align = Alignment(horizontal='center', vertical='center')
            right_align = Alignment(horizontal='right')
            
            col_pairs_a = [(1,2), (3,4), (5,6), (7,8), (9,10)] # A-J
            col_pairs_b = [(12,13), (14,15), (16,17), (18,19), (20,21)] # L-U
            
            for c1, c2 in col_pairs_a + col_pairs_b:
                ws.column_dimensions[get_column_letter(c1)].width = 16
                ws.column_dimensions[get_column_letter(c2)].width = 4
                
            ws.column_dimensions['K'].width = 8 # Divider
            for r in range(1, 200): ws[f"K{r}"].fill = PatternFill("solid", fgColor="000000")

            # --- HEADERS ---
            ws.merge_cells("A1:J1")
            ws["A1"] = f"Outpatient Center-LBJ Group A {calendar.month_name[int(self.var_month.get())]} {self.var_year.get()}"
            ws["A1"].font = title_font
            ws["A1"].alignment = center_align
            
            ws.merge_cells("L1:U1")
            ws["L1"] = f"Outpatient Center-LBJ Group B {calendar.month_name[int(self.var_month.get())]} {self.var_year.get()}"
            ws["L1"].font = title_font
            ws["L1"].alignment = center_align

            days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
            
            # Draw Days Headers
            for i, d in enumerate(days):
                # Group A
                c1_idx = col_pairs_a[i][0]
                c2_idx = col_pairs_a[i][1]
                ws.merge_cells(start_row=3, start_column=c1_idx, end_row=3, end_column=c2_idx)
                cell = ws.cell(row=3, column=c1_idx, value=d)
                cell.font = bold_font
                cell.alignment = center_align
                for cx in range(c1_idx, c2_idx+1):
                     ws.cell(row=3, column=cx).border = Border(top=Side(style='thick'), bottom=Side(style='thick'), left=Side(style='thick') if cx==c1_idx else None, right=Side(style='thick') if cx==c2_idx else None)
                
                # Group B
                c1_idx_b = col_pairs_b[i][0]
                c2_idx_b = col_pairs_b[i][1]
                ws.merge_cells(start_row=3, start_column=c1_idx_b, end_row=3, end_column=c2_idx_b)
                cell = ws.cell(row=3, column=c1_idx_b, value=d)
                cell.font = bold_font
                cell.alignment = center_align
                for cx in range(c1_idx_b, c2_idx_b+1):
                     ws.cell(row=3, column=cx).border = Border(top=Side(style='thick'), bottom=Side(style='thick'), left=Side(style='thick') if cx==c1_idx_b else None, right=Side(style='thick') if cx==c2_idx_b else None)

            logic = self.scheduler_logic
            cal = calendar.monthcalendar(logic.year, logic.month)
            c_map = {p.name.strip(): p.color.replace("#","") for p in self.physicians}
            
            current_row = 4
            
            for week in cal:
                has_weekday = any(d != 0 and idx < 5 for idx, d in enumerate(week))
                if not has_weekday: continue
                
                start_r = current_row
                end_r = current_row + 11 # 0 to 11 is 12 rows
                
                for day_idx in range(5):
                    day_num = week[day_idx]
                    if day_num == 0: continue
                    
                    idx_a_text, idx_a_date = col_pairs_a[day_idx]
                    idx_b_text, idx_b_date = col_pairs_b[day_idx]
                    
                    sched = logic.schedule.get(day_num, {'AM': [], 'PM': []})
                    
                    # --- FILL Logic ---
                    def fill_day_block(c_txt, c_date, shifts, is_group_b):
                        # Row 1: AM Header + Date
                        if shifts['AM']:
                            ws.cell(row=start_r, column=c_txt, value="AM CLINIC").font = Font(size=10, bold=True)
                        
                        ws.cell(row=start_r, column=c_date, value=day_num).font = Font(size=11, bold=True)
                        ws.cell(row=start_r, column=c_date).alignment = right_align
                        
                        # Row 2: AM Name
                        if shifts['AM']:
                            doc = shifts['AM'][0]
                            cell = ws.cell(row=start_r+1, column=c_txt, value=doc)
                            cell.font = bold_font
                            ws.merge_cells(start_row=start_r+1, start_column=c_txt, end_row=start_r+1, end_column=c_date)
                            if doc.strip() in c_map:
                                cell.fill = PatternFill("solid", fgColor=c_map[doc.strip()])
                        
                        # Row 7: PM Header
                        if shifts['PM']:
                            ws.cell(row=start_r+6, column=c_txt, value="PM CLINIC").font = Font(size=10, bold=True)
                            
                        # Row 8: PM Name
                        if shifts['PM']:
                            doc = shifts['PM'][0]
                            cell = ws.cell(row=start_r+7, column=c_txt, value=doc)
                            cell.font = bold_font
                            ws.merge_cells(start_row=start_r+7, start_column=c_txt, end_row=start_r+7, end_column=c_date)
                            if doc.strip() in c_map:
                                cell.fill = PatternFill("solid", fgColor=c_map[doc.strip()])

                    am_all = sched['AM']
                    pm_all = sched['PM']
                    
                    shifts_A = {'AM':[], 'PM':[]}
                    shifts_B = {'AM':[], 'PM':[]}
                    
                    am_prim_is_B = (day_idx == 0 or day_idx == 4) 
                    if am_all:
                        p1 = am_all[0]
                        others = am_all[1:]
                        if am_prim_is_B:
                            shifts_B['AM'].append(p1)
                            if others: shifts_A['AM'].extend(others)
                        else:
                            shifts_A['AM'].append(p1)
                            if others: shifts_B['AM'].extend(others)

                    pm_prim_is_A = (day_idx == 2 or day_idx == 3)
                    if pm_all:
                        p1 = pm_all[0]
                        others = pm_all[1:]
                        if pm_prim_is_A:
                            shifts_A['PM'].append(p1)
                            if others: shifts_B['PM'].extend(others)
                        else:
                            shifts_B['PM'].append(p1)
                            if others: shifts_A['PM'].extend(others) 

                    fill_day_block(idx_a_text, idx_a_date, shifts_A, is_group_b=False)
                    fill_day_block(idx_b_text, idx_b_date, shifts_B, is_group_b=True)
                    
                    def draw_block_border(c1, c2):
                        for r_x in range(start_r, end_r+1):
                            for c_x in range(c1, c2+1):
                                t = Side(style='thick') if r_x == start_r else None
                                b = Side(style='thick') if r_x == end_r else None
                                l = Side(style='thick') if c_x == c1 else None
                                r_s = Side(style='thick') if c_x == c2 else None
                                
                                existing = ws.cell(row=r_x, column=c_x).border
                                new_border = Border(
                                    top=t if t else existing.top,
                                    bottom=b if b else existing.bottom,
                                    left=l if l else existing.left,
                                    right=r_s if r_s else existing.right
                                )
                                ws.cell(row=r_x, column=c_x).border = new_border
                                
                    draw_block_border(idx_a_text, idx_a_date)
                    draw_block_border(idx_b_text, idx_b_date)

                current_row += 12 

            for r in range(1, current_row): ws[f"K{r}"].fill = PatternFill("solid", fgColor="000000")

            wb.save(filename)
            messagebox.showinfo("Success", "Legacy Export complete.")
            
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def setup_tab4(self):
        # Import/Export Tab
        for w in self.tab_imp.winfo_children(): w.destroy()
        frame = ctk.CTkFrame(self.tab_imp)
        frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(frame, text="Data Management", font=("Arial", 20, "bold")).pack(pady=20)
        
        ctk.CTkButton(frame, text="IMPORT Schedule from Excel", command=self.import_excel, height=50, width=300).pack(pady=10)
        ctk.CTkButton(frame, text="EXPORT Schedule to Excel (Standard)", command=self.export_to_excel, height=50, width=300, fg_color="green").pack(pady=10)
        ctk.CTkButton(frame, text="EXPORT Schedule to Excel (Legacy Format)", command=self.export_legacy_excel, height=50, width=300, fg_color="#555555").pack(pady=10)
        
        ctk.CTkLabel(frame, text="Note: Export includes settings in hidden cell A80.", text_color="gray").pack(pady=20)

    def setup_tab5(self):
        txt = ctk.CTkTextbox(self.tab_hlp, font=("Arial", 14), activate_scrollbars=True)
        txt.pack(fill="both", expand=True, padx=10, pady=10)
        
        info = """CLINIC SCHEDULER - USER GUIDE

==================================================
              1. BASICS
==================================================
Welcome to the Clinic Scheduler! This tool is designed to take the headache out of scheduling by automatically balancing workloads and honoring doctor preferences.

HOW TO RUN A BASIC SCHEDULE:
1. Enter Doctor Information (Tab 1)
   - Add your physicians and set their "Target" shifts for the month.
   - Enter their preferences, avoidances, or if they only work early/late in the month.
2. Define Clinic Needs (Tab 2)
   - Specify how many doctors are needed each morning (AM) and afternoon (PM).
   - Use "Close Clinic" to mark weeks with no activity (e.g., holidays).
3. Generate the Schedule (Tab 3 & 3.5)
   - Click 'REGENERATE SCHEDULE'. The software will instantly assign shifts.
   - Don't like the result? Just click 'REGENERATE SCHEDULE' again! Every click explores a different viable combination.
4. Export and Share (Tab 4)
   - Once you are happy with the visual calendar, go to Tab 4 to export the schedule directly into an Excel template.

==================================================
              2. SETTINGS
==================================================
A comprehensive guide to every option available in the UI.

TAB 1: PHYSICIANS
- Active (Checkbox): Uncheck to temporarily remove a doctor from scheduling consideration.
- Target: The strict total number of shifts this doctor needs to work this month.
- 1st / 2nd (Checkboxes): Restricts the doctor to only work during the first half (1st) or second half (2nd) of the month.
- FullDay? (Checkbox): If checked, the algorithm is allowed to assign this doctor an AM and PM shift on the exact same day. If unchecked, they will only work half-days.
- Preferred & Avoid & Override Dates: 
  - Acceptable Syntax: You can use spaces, commas, or both to separate dates.
  - Single Dates: "12, 15" or "12 15"
  - Date Ranges: "20-25" (Includes the 20th through the 25th)
  - AM/PM Specific: "12AM, 15PM" or "1-5AM" (Restricts only that half of the day)
  - Combinations: "1, 3-5, 12PM" 
  - (Preferred dates are soft requests. Avoid/Override are strict constraints.)

TAB 1 GLOBALS:
- Algorithm Mode ("Ratio-Based Logic" vs "Standard") & Auto-Balance: 
  - Standard: A purely mathematical approach to filling the calendar linearly.
  - Ratio-Based (Recommended): A highly intelligent mode that scales down targets to calculate the true needed "ratio" of shifts before beginning, ensuring a smoother spread across the month and eliminating end-of-month clumping.
  - Auto-Balance Limits: Check this to automatically spread any remaining needed shifts among available doctors evenly when the clinic needs exceed the sum of doctor targets.
- Desperation Level: Controls how strictly the algorithm adheres to rules.
  - OFF (Stage 0): Perfect compliance. Will leave a shift empty rather than break a rule.
  - FAIR (Stage 1 - RECOMMENDED): Bends minor rules (like preference density) to ensure the clinic is staffed. 
  - HIGH (Stage 2): Breaks major rules (like 'Avoid' dates) to force staffing. Keep this OFF unless completely desperate.

TAB 2: NEEDS
- AM/PM Number Inputs: Set the required number of coverage spots per half-day.
- First/Second Half Divider: Sets the specific date that divides the "1st" and "2nd" half of the month for doctors using those checkboxes (e.g., setting it to 15 means the 1st half is 1-15, and the 2nd half is 16+).
- Close Clinic (Button): Zeros out all requirements for a specific week, effectively blocking any scheduled shifts for that time.

TAB 3 & 3.5: SCHEDULE GENERATION
- REGENERATE SCHEDULE (Button): Deletes the old calendar and starts a fresh algorithmic generation.

TAB 4: DATA MANAGEMENT
- IMPORT: Load a previously saved schedule configuration from an Excel file.
- EXPORT (Standard): Save the calendar in the standard printable format.
- EXPORT (Legacy): Save a heavily color-coded legacy format.

==================================================
              3. LOGIC
==================================================
How does the algorithmic brain actually work?

THE CORE ENGINE:
1. Setup Phase: The system reads global needs and individual constraints. It calculates the exact number of shifts everyone needs to work. 
2. Override Insertion: "Override" dates are injected immediately, bypassing all checks.
3. Scoring Phase: The system creates a pool of "Available Candidates" for every single shift slot in the month.
4. Sorting by Rarity: It orders the days based on how difficult they are to fill. Days with fewest available doctors get scheduled first.
5. Drafting Phase: Doctors draft shifts. Priority is given to whoever is furthest from their Target.

FAIRNESS & DIFFICULTY ADAPTATION:
The scheduler understands "Shift Difficulty." 
- Example: If Dr. Alpha is only available early in the month, and Dr. Beta is fully open, the system recognizes early slots as "High Difficulty" to fill. It guarantees Dr. Alpha gets those spots, preserving Dr. Beta's flexible availability for the end of the month.

CONFLICT RESOLUTION:
If a slot cannot be filled, the system consults the "Desperation Tracker." 
- Stage 0: Perfect compliance. No rules broken.
- Stage 1: Minor rules bent. (e.g., pushing doctors closer together than they prefer).
- Stage 2: Major rules bent. (e.g., overriding 'Avoid' dates).

ERROR RECOVERY:
Instead of crashing, if the scheduler hits a dead end where a shift is fundamentally impossible to cover, it marks the slot as "OPEN" in bright red, saving the rest of the schedule so you can manually intervene.

==================================================
              4. PRO-TIPS
==================================================
- Supplemental Files: The software creates and updates background files (like `physician_state.json`) in the same folder to save your roster and settings. Do not delete these files unless you intentionally want to clear all your saved data!
- "Click Until It Clicks": If a generated schedule doesn't visually look pleasing to you, just hit regenerate! The randomizer weights guarantee you'll see a variety of solid options.
- The FullDay Hack: If you are desperately short on doctors, temporarily turning on 'FullDay?' for a few flexible staff members is the fastest way to solve scheduling blockages.
- Overriding: Use Overrides sparingly. If everyone gets an override, the system's ability to balance workloads effectively collapses.
- Excel Editing: Never edit the hidden configuration string (Row 80) in the exported Excel sheets; it breaks the ability to re-import settings!
"""
        txt.insert("0.0", info)
        txt.configure(state="disabled")
