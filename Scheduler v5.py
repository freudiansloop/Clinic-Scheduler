import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import calendar
import datetime
import random
import json
import os
import sys
import copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Configuration & Constants ---
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

APP_NAME = "Clinic Physician Scheduler Pro"
STATE_FILE = "physician_state.json"

# specific default colors
DEFAULT_COLORS = {
    "Gandhi": "#87CEFA",    # Light Blue
    "Wesley": "#CC5500",    # Burnt Orange
    "Aisenberg": "#A9A9A9", # Dark Gray (adjusted for visibility)
    "Govindu": "#C8A2C8",   # Lilac
    "Reymunde": "#800080",  # Purple
    "Koney": "#98FF98",     # Mint
    "Lee": "#F0E68C",       # Khaki
    "Huq": "#4C6A92",       # Light Navy
    "Rendon": "#FFC0CB",    # Pink
    "Bhandari": "#FFFF00",  # Highlighter Yellow
    "Dash": "#008000",      # Green
    "Khaja": "#FFDAB9"      # Light Orange
}

# Fallback palette
COLOR_PALETTE = [
    "#FF0000", "#00FF00", "#0000FF", "#FFFF00", "#00FFFF", "#FF00FF", 
    "#C0C0C0", "#800000", "#808000", "#008000", "#800080", "#008080", 
    "#000080", "#FF4500", "#DA70D6", "#FA8072", "#20B2AA", "#778899", 
    "#B0C4DE", "#FFFFE0", "#FFD700", "#ADFF2F", "#7FFFD4", "#FF69B4"
]

DEFAULT_ROSTER_DATA = [
    ("Gandhi", 8), ("Wesley", 8), ("Khaja", 2), ("Rendon", 2),
    ("Reymunde", 2), ("Dash", 2), ("Govindu", 2), ("Lee", 2),
    ("Koney", 2), ("Aisenberg", 2), ("Bhandari", 2), ("Huq", 2)
]

# --- Helper Functions ---

def get_app_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def parse_date_input(text, year, month):
    if not text.strip():
        return []
    parts = [p.strip() for p in text.split(',')]
    result = []
    for p in parts:
        p = p.upper()
        try:
            day_str = ''.join(filter(str.isdigit, p))
            if not day_str: continue
            day = int(day_str)
            _, last_day = calendar.monthrange(year, month)
            if day < 1 or day > last_day: continue

            if "AM" in p:
                result.append({'day': day, 'type': 'AM'})
            elif "PM" in p:
                result.append({'day': day, 'type': 'PM'})
            else:
                result.append({'day': day, 'type': 'AM'})
                result.append({'day': day, 'type': 'PM'})
        except ValueError:
            continue
    return result

# --- Classes ---

class Physician:
    def __init__(self, name, target, active=True, half_month="All", preferred="", avoid="", override="", color="#FFFFFF"):
        self.id = str(random.getrandbits(32))
        self.name = name
        self.target = int(target)
        self.active = active
        self.half_month = half_month # "All", "1st", "2nd"
        self.preferred_str = preferred
        self.avoid_str = avoid
        self.override_str = override
        self.color = color
        self.assigned_shifts = [] 
    
    def to_dict(self):
        return {
            "name": self.name,
            "target": self.target,
            "active": self.active,
            "half_month": self.half_month,
            "preferred": self.preferred_str,
            "avoid": self.avoid_str,
            "override": self.override_str,
            "color": self.color
        }

    @classmethod
    def from_dict(cls, data):
        return cls(
            name=data.get("name", ""),
            target=data.get("target", 0),
            active=data.get("active", True),
            half_month=data.get("half_month", "All"),
            preferred=data.get("preferred", ""),
            avoid=data.get("avoid", ""),
            override=data.get("override", ""),
            color=data.get("color", "#FFFFFF")
        )

class SchedulerLogic:
    def __init__(self, physicians, year, month, daily_needs, split_day=16):
        self.physicians = [p for p in physicians if p.active]
        self.year = year
        self.month = month
        self.daily_needs = daily_needs
        self.split_day = split_day
        self.schedule = {} 
        self.logs = []
        self.warnings = [] 
        
        _, self.last_day = calendar.monthrange(year, month)
        for d in range(1, self.last_day + 1):
            self.schedule[d] = {'AM': [], 'PM': []}
            
    def is_weekend(self, day):
        weekday = calendar.weekday(self.year, self.month, day)
        return weekday >= 5

    def get_valid_days(self, physician):
        valid = []
        for d in range(1, self.last_day + 1):
            if self.is_weekend(d): continue
            
            # Use dynamic split day
            if physician.half_month == "1st" and d >= self.split_day: continue
            if physician.half_month == "2nd" and d < self.split_day: continue
            
            valid.append(d)
        return valid

    def can_assign(self, physician, day, shift_type, check_avoid=True):
        if check_avoid:
            avoids = parse_date_input(physician.avoid_str, self.year, self.month)
            for avoid in avoids:
                if avoid['day'] == day and avoid['type'] == shift_type:
                    self.warnings.append(f"Avoidance Conflict: {physician.name} avoided D{day} {shift_type} but was placed via override.")
                    return False
        
        for assigned_d, assigned_t in physician.assigned_shifts:
            if assigned_d == day and assigned_t == shift_type:
                return False
        return True

    def run(self):
        self.warnings = []
        for p in self.physicians:
            p.assigned_shifts = []

        # 1. Overrides
        for p in self.physicians:
            overrides = parse_date_input(p.override_str, self.year, self.month)
            for req in overrides:
                day, s_type = req['day'], req['type']
                if day > self.last_day: continue
                
                can_assign_result = self.can_assign(p, day, s_type, check_avoid=True)
                if not can_assign_result:
                    self.warnings.append(f"OVERRIDE WARNING: {p.name} Override on D{day} {s_type} conflicts with their Avoidance. Assigned anyway.")
                    
                needed = self.daily_needs.get(day, {}).get(s_type, 0)
                current = len(self.schedule[day][s_type])
                
                if current < needed:
                    self.schedule[day][s_type].append(p.name)
                    p.assigned_shifts.append((day, s_type))
                else:
                    self.warnings.append(f"OVERRIDE FAILED: {p.name} on D{day} {s_type} - Clinic slot is not needed (Needs: {needed}).")

        # 2. Preferences
        for p in self.physicians:
            prefs = parse_date_input(p.preferred_str, self.year, self.month)
            for req in prefs:
                day, s_type = req['day'], req['type']
                if day > self.last_day: continue

                if len(p.assigned_shifts) >= p.target:
                    self.warnings.append(f"Preference Ignored: {p.name} on D{day} {s_type} - Target of {p.target} shifts already met.")
                    continue
                
                needed = self.daily_needs.get(day, {}).get(s_type, 0)
                current = len(self.schedule[day][s_type])
                
                if current < needed and self.can_assign(p, day, s_type, check_avoid=True):
                    self.schedule[day][s_type].append(p.name)
                    p.assigned_shifts.append((day, s_type))
                else:
                    self.warnings.append(f"Preference Failed: {p.name} on D{day} {s_type} - Slot unavailable (full, avoided, or half-month restriction).")

        # 3. Round Robin (With Shuffle for Non-Determinism)
        phys_idx = 0 
        max_iterations = len(self.physicians) * self.last_day * 2 * 3 

        for i in range(max_iterations):
            eligible_physicians = [p for p in self.physicians if len(p.assigned_shifts) < p.target]
            
            open_slots_exist = False
            for d in range(1, self.last_day + 1):
                if d in self.daily_needs:
                    for s_type in ['AM', 'PM']:
                        needed = self.daily_needs[d].get(s_type, 0)
                        current = len(self.schedule[d][s_type])
                        if current < needed:
                            open_slots_exist = True
                            break
                if open_slots_exist: break
                
            if not eligible_physicians or not open_slots_exist:
                break 

            p_index = phys_idx % len(eligible_physicians)
            p = eligible_physicians[p_index]
            
            found_slot = False
            valid_days = self.get_valid_days(p)
            
            # --- Non-Deterministic Fix ---
            # Instead of iterating sorted days, get all possible open slots for this doc
            # and pick one randomly.
            possible_slots = []
            for day in valid_days:
                for s_type in ['AM', 'PM']:
                    needed = self.daily_needs.get(day, {}).get(s_type, 0)
                    current = len(self.schedule[day][s_type])
                    if current < needed:
                         if self.can_assign(p, day, s_type, check_avoid=True):
                             possible_slots.append((day, s_type))
            
            if possible_slots:
                # Random choice makes it non-deterministic
                pick_day, pick_type = random.choice(possible_slots)
                self.schedule[pick_day][pick_type].append(p.name)
                p.assigned_shifts.append((pick_day, pick_type))
                phys_idx += 1
                found_slot = True
            
            if not found_slot:
                phys_idx += 1 

        # 4. Audits
        for p in self.physicians:
            act = len(p.assigned_shifts)
            if act < p.target:
                self.warnings.append(f"TARGET UNMET: {p.name} needed {p.target}, assigned {act}.")

        for d in range(1, self.last_day + 1):
            if d in self.daily_needs:
                for s_type in ['AM', 'PM']:
                    needed = self.daily_needs[d].get(s_type, 0)
                    current = len(self.schedule[d][s_type])
                    if current < needed:
                        self.warnings.append(f"CRITICAL: Unfilled Clinic Slot on Day {d} {s_type} ({needed - current} still needed).")

class ColorPicker(ctk.CTkToplevel):
    def __init__(self, parent, current_color, taken_colors, on_select):
        super().__init__(parent)
        self.title("Select Color")
        self.geometry("400x400")
        self.attributes("-topmost", True)
        self.on_select = on_select
        
        lbl = ctk.CTkLabel(self, text="Choose a color for this physician:", font=("Arial", 14))
        lbl.pack(pady=10)
        
        grid = ctk.CTkFrame(self)
        grid.pack(expand=True, padx=20, pady=20)
        
        palette = list(COLOR_PALETTE)
        for c in DEFAULT_COLORS.values():
            if c not in palette: palette.append(c)

        while len(palette) < len(taken_colors) + 5:
            r = lambda: random.randint(0,255)
            palette.append('#%02X%02X%02X' % (r(),r(),r()))

        r, c = 0, 0
        for color in palette:
            btn = ctk.CTkButton(grid, text="", width=40, height=40, fg_color=color, hover_color=color,
                                command=lambda col=color: self.select(col))
            btn.grid(row=r, column=c, padx=5, pady=5)
            c += 1
            if c > 4:
                c = 0
                r += 1

    def select(self, color):
        self.on_select(color)
        self.destroy()

class AppUI(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1400x900")
        
        self.physicians = []
        self.history = [] 
        self.schedule_results = None
        self.scheduler_logic = None
        
        self.tabview = ctk.CTkTabview(self)
        self.tabview.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.tab_emp = self.tabview.add("1. Physicians")
        self.tab_cal = self.tabview.add("2. Needs")
        self.tab_out = self.tabview.add("3. Schedule")
        self.tab_col = self.tabview.add("3.5 Color Schedule")
        self.tab_imp = self.tabview.add("4. Import/Export")
        self.tab_hlp = self.tabview.add("5. Instructions")
        
        self.setup_tab1()
        self.setup_tab2()
        self.setup_tab3()
        self.setup_tab_colored()
        self.setup_tab4()
        self.setup_tab5()
        
        self.load_state_from_disk()
        
    def save_snapshot(self):
        snap = [p.to_dict() for p in self.physicians]
        self.history.append(snap)
        if len(self.history) > 10: self.history.pop(0)

    def undo(self):
        if not self.history: return
        prev_state = self.history.pop()
        self.physicians = [Physician.from_dict(d) for d in prev_state]
        self.refresh_physician_list()

    def get_resource_path(self):
        return os.path.join(get_app_path(), STATE_FILE)

    def save_state_to_disk(self):
        self.update_physician_objects_from_ui()
        data = { "physicians": [p.to_dict() for p in self.physicians] }
        try:
            with open(self.get_resource_path(), 'w') as f: json.dump(data, f, indent=4)
        except: pass

    def load_state_from_disk(self):
        path = self.get_resource_path()
        if os.path.exists(path):
            try:
                with open(path, 'r') as f:
                    data = json.load(f)
                    self.physicians = [Physician.from_dict(x) for x in data.get("physicians", [])]
                self.refresh_physician_list()
                return
            except: pass
        self.reset_to_defaults()

    # --- TAB 1 ---
    def setup_tab1(self):
        # Instructions
        inst_frame = ctk.CTkFrame(self.tab_emp, height=40)
        inst_frame.pack(fill="x", padx=5, pady=5)
        inst_lbl = ctk.CTkLabel(inst_frame, text="INSTRUCTIONS: Enter targets as integers. Use '12' or '15 AM, 16 PM' format for dates. Order list by priority.", font=("Arial", 12))
        inst_lbl.pack(pady=5)

        ctrl_frame = ctk.CTkFrame(self.tab_emp)
        ctrl_frame.pack(fill="x", padx=5, pady=5)
        
        ctk.CTkButton(ctrl_frame, text="Reset Defaults", fg_color="#555555", command=self.reset_to_defaults).pack(side="left", padx=5)
        ctk.CTkButton(ctrl_frame, text="Undo", fg_color="#555555", command=self.undo).pack(side="left", padx=5)
        ctk.CTkButton(ctrl_frame, text="Add Physician", command=self.add_blank_physician).pack(side="left", padx=5)
        ctk.CTkButton(ctrl_frame, text="Clear All Dates", fg_color="#b30000", command=self.clear_all_dates).pack(side="right", padx=20)

        self.header_frame = ctk.CTkFrame(self.tab_emp, height=30)
        self.header_frame.pack(fill="x", padx=5, pady=(5,0))
        
        headers = ["Order", "Active", "Name", "Target", "1st Half", "2nd Half", "Preferred", "Avoid", "Override", "Color", "Del"]
        self.header_frame.grid_columnconfigure(2, weight=1) 
        self.header_frame.grid_columnconfigure(6, weight=1) 
        self.header_frame.grid_columnconfigure(7, weight=1) 
        
        for i, h in enumerate(headers):
            lbl = ctk.CTkLabel(self.header_frame, text=h, font=("Arial", 12, "bold"))
            lbl.grid(row=0, column=i, padx=2, sticky="ew")

        self.p_scroll = ctk.CTkScrollableFrame(self.tab_emp)
        self.p_scroll.pack(fill="both", expand=True, padx=5, pady=5)
        self.p_scroll.grid_columnconfigure(2, weight=1)
        self.p_scroll.grid_columnconfigure(6, weight=1)
        self.p_scroll.grid_columnconfigure(7, weight=1)

        # Footer for Sum
        self.footer_frame = ctk.CTkFrame(self.tab_emp, height=30)
        self.footer_frame.pack(fill="x", padx=5, pady=5)
        self.lbl_total_target = ctk.CTkLabel(self.footer_frame, text="Total Target: 0", font=("Arial", 14, "bold"))
        self.lbl_total_target.pack()

        self.p_rows = []

    def reset_to_defaults(self):
        self.save_snapshot()
        self.physicians = []
        for name, target in DEFAULT_ROSTER_DATA:
            color = DEFAULT_COLORS.get(name, "#FFFFFF")
            self.physicians.append(Physician(name, target, active=True, color=color))
        self.refresh_physician_list()

    def add_blank_physician(self):
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        self.physicians.append(Physician("New Doc", 0, color="#FFFFFF"))
        self.refresh_physician_list()

    def clear_all_dates(self):
        if not messagebox.askyesno("Confirm", "Clear dates for ALL physicians?"):
            return
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        for p in self.physicians:
            p.override_str = ""
            p.preferred_str = ""
            p.avoid_str = ""
        self.refresh_physician_list()

    def move_row(self, index, direction):
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        if direction == -1 and index > 0:
            self.physicians[index], self.physicians[index-1] = self.physicians[index-1], self.physicians[index]
        elif direction == 1 and index < len(self.physicians) - 1:
            self.physicians[index], self.physicians[index+1] = self.physicians[index+1], self.physicians[index]
        self.refresh_physician_list()

    def delete_row(self, index):
        if not messagebox.askyesno("Confirm", "Delete this physician?"): return
        self.save_snapshot()
        self.update_physician_objects_from_ui()
        self.physicians.pop(index)
        self.refresh_physician_list()

    def open_color_picker(self, idx):
        self.update_physician_objects_from_ui()
        current_phys = self.physicians[idx]
        used_map = {p.color: p for p in self.physicians if p != current_phys}
        
        def on_color_chosen(new_color):
            current_phys.color = new_color
            self.refresh_physician_list()
            self.save_state_to_disk()

        ColorPicker(self, current_phys.color, list(used_map.keys()), on_color_chosen)

    def toggle_half_month(self, idx, mode):
        row_widgets = self.p_rows[idx]
        if mode == "1st":
            if row_widgets['chk_1st'].get():
                row_widgets['chk_2nd'].set(False)
        else:
            if row_widgets['chk_2nd'].get():
                row_widgets['chk_1st'].set(False)
        self.save_state_to_disk()

    def toggle_override(self, index, btn_ref):
        dialog = ctk.CTkInputDialog(text="Override Dates (e.g. 12 AM, 15 PM):", title="Overrides")
        res = dialog.get_input()
        if res is not None:
            self.physicians[index].override_str = res
            if res.strip():
                btn_ref.configure(text="SET", fg_color="#b30000")
            else:
                btn_ref.configure(text="Set", fg_color="transparent", border_width=1)
            self.save_state_to_disk()

    def refresh_physician_list(self):
        for widget in self.p_scroll.winfo_children(): widget.destroy()
        self.p_rows = []
        total_tgt = 0

        for i, p in enumerate(self.physicians):
            total_tgt += p.target

            f_ord = ctk.CTkFrame(self.p_scroll, fg_color="transparent")
            f_ord.grid(row=i, column=0, padx=2, pady=2)
            ctk.CTkButton(f_ord, text="▲", width=20, command=lambda x=i: self.move_row(x, -1)).pack(side="left")
            ctk.CTkButton(f_ord, text="▼", width=20, command=lambda x=i: self.move_row(x, 1)).pack(side="left")
            
            var_active = ctk.BooleanVar(value=p.active)
            chk_active = ctk.CTkCheckBox(self.p_scroll, text="", variable=var_active, width=20)
            chk_active.grid(row=i, column=1, padx=2)
            
            ent_name = ctk.CTkEntry(self.p_scroll)
            ent_name.insert(0, p.name)
            ent_name.grid(row=i, column=2, padx=2, sticky="ew")
            
            ent_tgt = ctk.CTkEntry(self.p_scroll, width=50)
            ent_tgt.insert(0, str(p.target))
            ent_tgt.grid(row=i, column=3, padx=2)
            # Bind update for total sum
            ent_tgt.bind("<FocusOut>", lambda e: self.update_physician_objects_from_ui())
            
            var_1st = ctk.BooleanVar(value=(p.half_month=="1st"))
            var_2nd = ctk.BooleanVar(value=(p.half_month=="2nd"))
            
            c1 = ctk.CTkCheckBox(self.p_scroll, text="", variable=var_1st, width=20, 
                                 command=lambda x=i: self.toggle_half_month(x, "1st"))
            c1.grid(row=i, column=4, padx=5)
            
            c2 = ctk.CTkCheckBox(self.p_scroll, text="", variable=var_2nd, width=20,
                                 command=lambda x=i: self.toggle_half_month(x, "2nd"))
            c2.grid(row=i, column=5, padx=5)

            ent_pref = ctk.CTkEntry(self.p_scroll)
            ent_pref.insert(0, p.preferred_str)
            ent_pref.grid(row=i, column=6, padx=2, sticky="ew")
            
            ent_avoid = ctk.CTkEntry(self.p_scroll)
            ent_avoid.insert(0, p.avoid_str)
            ent_avoid.grid(row=i, column=7, padx=2, sticky="ew")
            
            btn_over = ctk.CTkButton(self.p_scroll, text="SET" if p.override_str else "Set", width=50,
                                     fg_color="#b30000" if p.override_str else "transparent",
                                     border_color="red", border_width=1)
            btn_over.configure(command=lambda x=i, b=btn_over: self.toggle_override(x, b))
            btn_over.grid(row=i, column=8, padx=2)
            
            btn_col = ctk.CTkButton(self.p_scroll, text="", width=30, height=20,
                                    fg_color=p.color, border_color="gray", border_width=1,
                                    command=lambda x=i: self.open_color_picker(x))
            btn_col.grid(row=i, column=9, padx=5)

            btn_del = ctk.CTkButton(self.p_scroll, text="X", width=30, fg_color="#b30000", 
                                    command=lambda x=i: self.delete_row(x))
            btn_del.grid(row=i, column=10, padx=15)
            
            self.p_rows.append({
                "active": var_active, "name": ent_name, "target": ent_tgt,
                "chk_1st": var_1st, "chk_2nd": var_2nd,
                "pref": ent_pref, "avoid": ent_avoid, "obj_id": p.id
            })
        
        self.lbl_total_target.configure(text=f"Total Target Sum: {total_tgt}")
        self.save_state_to_disk()

    def update_physician_objects_from_ui(self):
        new_list = []
        pmap = {p.id: p for p in self.physicians}
        total_tgt = 0
        for row in self.p_rows:
            orig = pmap.get(row['obj_id'])
            name = row['name'].get()
            try: tgt = int(row['target'].get())
            except: tgt = 0
            total_tgt += tgt

            hm = "All"
            if row['chk_1st'].get(): hm = "1st"
            elif row['chk_2nd'].get(): hm = "2nd"
            
            p = Physician(
                name=name, target=tgt, active=row['active'].get(),
                half_month=hm, preferred=row['pref'].get(), avoid=row['avoid'].get(),
                override=orig.override_str if orig else "",
                color=orig.color if orig else "#FFFFFF"
            )
            p.id = row['obj_id']
            new_list.append(p)
        self.physicians = new_list
        self.lbl_total_target.configure(text=f"Total Target Sum: {total_tgt}")

    # --- TAB 2 ---
    def setup_tab2(self):
        ctrl = ctk.CTkFrame(self.tab_cal)
        ctrl.pack(pady=10)
        now = datetime.datetime.now()
        months = [str(i) for i in range(1, 13)]
        years = [str(now.year + i) for i in range(0, 5)]
        self.var_month = ctk.StringVar(value=str(now.month))
        self.var_year = ctk.StringVar(value=str(now.year))
        
        # Split Date
        self.var_split = ctk.StringVar(value="16")
        
        ctk.CTkLabel(ctrl, text="Month:").pack(side="left", padx=5)
        ctk.CTkComboBox(ctrl, values=months, variable=self.var_month, width=60, command=self.build_needs_grid).pack(side="left")
        
        ctk.CTkLabel(ctrl, text="Year:").pack(side="left", padx=5)
        ctk.CTkComboBox(ctrl, values=years, variable=self.var_year, width=70, command=self.build_needs_grid).pack(side="left")
        
        ctk.CTkLabel(ctrl, text="2nd Half Starts On:").pack(side="left", padx=10)
        ctk.CTkComboBox(ctrl, values=["14", "15", "16", "17"], variable=self.var_split, width=60).pack(side="left")

        ctk.CTkButton(ctrl, text="Reset to Standards", command=self.reset_needs_std).pack(side="left", padx=20)

        self.needs_frame = ctk.CTkScrollableFrame(self.tab_cal)
        self.needs_frame.pack(fill="both", expand=True, padx=10, pady=10)
        self.needs_widgets = {}
        self.build_needs_grid()

    def build_needs_grid(self, _=None):
        for w in self.needs_frame.winfo_children(): w.destroy()
        self.needs_widgets = {}
        y = int(self.var_year.get())
        m = int(self.var_month.get())
        
        days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Actions"]
        for i, d in enumerate(days):
            ctk.CTkLabel(self.needs_frame, text=d, font=("Arial", 14, "bold")).grid(row=0, column=i, padx=5, pady=5)
            self.needs_frame.grid_columnconfigure(i, weight=1)
            
        cal = calendar.monthcalendar(y, m)
        r = 1
        
        DEFAULT_MAP = {
            0: (1, 0), 1: (2, 0), 2: (0, 1), 3: (2, 1), 4: (1, 0), 5: (0, 0), 6: (0, 0)
        }

        for week_idx, week in enumerate(cal):
            week_widgets = []
            
            for day_idx, day_num in enumerate(week):
                if day_num == 0:
                    ctk.CTkLabel(self.needs_frame, text="").grid(row=r, column=day_idx)
                    continue
                
                is_weekend = (day_idx >= 5)
                bg_col = "#444444" if is_weekend else "transparent"
                cell = ctk.CTkFrame(self.needs_frame, border_width=1, border_color="gray", fg_color=bg_col)
                cell.grid(row=r, column=day_idx, padx=5, pady=5, sticky="nsew")
                
                ctk.CTkLabel(cell, text=str(day_num), font=("Arial", 12, "bold")).pack(pady=2)
                
                if not is_weekend:
                    def_am, def_pm = DEFAULT_MAP.get(day_idx, (0,0))
                    
                    # AM
                    f_am = ctk.CTkFrame(cell, fg_color="transparent")
                    f_am.pack(fill="x", padx=2, pady=1)
                    ctk.CTkLabel(f_am, text="AM", width=25, font=("Arial", 10)).pack(side="left")
                    cmb_am = ctk.CTkComboBox(f_am, values=["0", "1", "2"], width=50)
                    cmb_am.set(str(def_am))
                    cmb_am.pack(side="left")
                    self.color_combo(str(def_am), cmb_am)
                    cmb_am.configure(command=lambda v, w=cmb_am: self.color_combo(v, w))
                    
                    # PM
                    f_pm = ctk.CTkFrame(cell, fg_color="transparent")
                    f_pm.pack(fill="x", padx=2, pady=1)
                    ctk.CTkLabel(f_pm, text="PM", width=25, font=("Arial", 10)).pack(side="left")
                    cmb_pm = ctk.CTkComboBox(f_pm, values=["0", "1", "2"], width=50)
                    cmb_pm.set(str(def_pm))
                    cmb_pm.pack(side="left")
                    self.color_combo(str(def_pm), cmb_pm)
                    cmb_pm.configure(command=lambda v, w=cmb_pm: self.color_combo(v, w))
                    
                    self.needs_widgets[(day_num, 'AM')] = cmb_am
                    self.needs_widgets[(day_num, 'PM')] = cmb_pm
                    week_widgets.append(cmb_am)
                    week_widgets.append(cmb_pm)
                else:
                    ctk.CTkLabel(cell, text="Weekend", text_color="gray").pack()

            btn_close = ctk.CTkButton(self.needs_frame, text="Close\nClinic", width=70, fg_color="#800000",
                                      command=lambda w=week_widgets: self.close_week(w))
            btn_close.grid(row=r, column=7, padx=5)
            r += 1

    def color_combo(self, val, widget):
        if val == "0": widget.configure(fg_color="#550000", button_color="#550000")
        else: widget.configure(fg_color="#1f6aa5", button_color="#1f6aa5")

    def reset_needs_std(self):
        self.build_needs_grid()

    def close_week(self, widgets):
        for w in widgets:
            w.set("0")
            self.color_combo("0", w)

    def get_needs_data(self):
        data = {}
        for (day, slot), widget in self.needs_widgets.items():
            if day not in data: data[day] = {}
            try: data[day][slot] = int(widget.get())
            except: data[day][slot] = 0
        return data

    # --- TAB 3 & 3.5 ---
    def setup_tab3(self):
        self.setup_sched_tab(self.tab_out, colored_text=False)
    
    def setup_tab_colored(self):
        self.setup_sched_tab(self.tab_col, colored_text=True)

    def setup_sched_tab(self, parent_tab, colored_text):
        top = ctk.CTkFrame(parent_tab)
        top.pack(fill="x", padx=10, pady=5)
        
        btn = ctk.CTkButton(top, text="REGENERATE SCHEDULE", font=("Arial", 14, "bold"), 
                            fg_color="#006400", height=40, command=self.generate_schedule)
        btn.pack(fill="x")
        
        # Frozen Headers Frame
        headers_frame = ctk.CTkFrame(parent_tab)
        headers_frame.pack(fill="x", padx=10, pady=(5,0))
        
        # Calendar Area
        container = ctk.CTkFrame(parent_tab, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=5, pady=5)
        
        cal_frame = ctk.CTkScrollableFrame(container)
        cal_frame.pack(side="left", fill="both", expand=True)
        
        # Configure columns for expansion
        for i in range(5): 
            headers_frame.grid_columnconfigure(i, weight=1)
            cal_frame.grid_columnconfigure(i, weight=1)
        
        # Render Frozen Headers
        days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
        for i, d in enumerate(days):
            ctk.CTkLabel(headers_frame, text=d, font=("Arial", 14, "bold")).grid(row=0, column=i, sticky="ew")

        if not colored_text:
            self.out_cal_frame = cal_frame
        else:
            self.col_cal_frame = cal_frame
            
        sidebar = ctk.CTkFrame(container, width=350)
        sidebar.pack(side="right", fill="y", padx=5)
        
        ctk.CTkLabel(sidebar, text="Schedule Statistics", font=("Arial", 16, "bold")).pack(pady=5)
        
        # Stats Grid
        stats_grid_frame = ctk.CTkFrame(sidebar, fg_color="transparent")
        stats_grid_frame.pack(fill="x", padx=5)
        
        headers = ["Name", "Target", "Actual", "Net"]
        for i, h in enumerate(headers):
            sticky = "w" if i == 0 else "e"
            ctk.CTkLabel(stats_grid_frame, text=h, font=("Arial", 12, "bold")).grid(row=0, column=i, padx=5, pady=2, sticky=sticky)
            if i > 0: stats_grid_frame.grid_columnconfigure(i, weight=1)
            
        ctk.CTkLabel(sidebar, text="Issues / Deviations", font=("Arial", 14, "bold"), text_color="orange").pack(pady=(10, 0))
        warnings_text = ctk.CTkTextbox(sidebar, height=200, text_color="orange")
        warnings_text.pack(fill="both", expand=True, padx=5, pady=5)
        
        stats_label = ctk.CTkLabel(top, text="Status: Waiting", text_color="gray")
        stats_label.pack(pady=5)
        
        # Store refs appropriately so they are unique per tab
        if not colored_text:
            self.stats_grid = stats_grid_frame
            self.warnings_text = warnings_text
            self.stats_label = stats_label
        else:
            self.stats_grid_col = stats_grid_frame
            self.warnings_text_col = warnings_text
            self.stats_label_col = stats_label

    def generate_schedule(self):
        self.update_physician_objects_from_ui()
        self.save_state_to_disk()
        y = int(self.var_year.get())
        m = int(self.var_month.get())
        try: split = int(self.var_split.get())
        except: split = 16

        needs = self.get_needs_data()
        
        logic = SchedulerLogic(self.physicians, y, m, needs, split_day=split)
        logic.run()
        self.scheduler_logic = logic
        
        self.render_calendar_logic(self.out_cal_frame, logic, use_colors=False)
        self.render_calendar_logic(self.col_cal_frame, logic, use_colors=True)
        
        # Render Stats for BOTH tabs
        self.render_stats(logic, self.stats_grid, self.stats_label, self.warnings_text)
        self.render_stats(logic, self.stats_grid_col, self.stats_label_col, self.warnings_text_col)

    def render_calendar_logic(self, frame, logic, use_colors):
        for w in frame.winfo_children(): w.destroy()
        y, m = logic.year, logic.month
        
        # Headers are now frozen outside, so we just render grid
        cal = calendar.monthcalendar(y, m)
        r = 0
        
        c_map = {p.name: p.color for p in logic.physicians}

        for week in cal:
            has_weekday = False
            for day_idx, day_num in enumerate(week):
                if day_idx >= 5: continue 
                has_weekday = True
                
                # Cell
                cell = ctk.CTkFrame(frame, border_width=1, border_color="#555555") 
                cell.grid(row=r, column=day_idx, padx=1, pady=1, sticky="nsew")
                
                if day_num == 0:
                    cell.configure(fg_color="transparent", border_width=0)
                    continue
                
                ctk.CTkLabel(cell, text=f"D{day_num}", font=("Arial", 10, "bold"), text_color="gray").pack(anchor="ne", padx=2)
                
                day_sched = logic.schedule.get(day_num, {'AM': [], 'PM': []})
                needs_am = logic.daily_needs.get(day_num, {}).get('AM', 0)
                needs_pm = logic.daily_needs.get(day_num, {}).get('PM', 0)
                
                # AM Half (#505050 is Light Gray)
                f_am = ctk.CTkFrame(cell, fg_color="#505050" if needs_am > 0 else "transparent", border_width=0)
                f_am.pack(fill="both", expand=True, padx=1, pady=(1, 0))
                
                if needs_am > 0:
                     ctk.CTkLabel(f_am, text="AM:", font=("Arial", 8, "bold"), text_color="silver").pack(anchor="w", padx=2)
                     for name in day_sched['AM']:
                         col = c_map.get(name, "white") if use_colors else "#aaddff"
                         ctk.CTkLabel(f_am, text=name, text_color=col, font=("Arial", 9)).pack(anchor="w", padx=2)
                     if len(day_sched['AM']) < needs_am:
                         ctk.CTkLabel(f_am, text=f"[OPEN]", text_color="#ff5555", font=("Arial", 9)).pack(anchor="w", padx=2)
                else:
                     ctk.CTkLabel(f_am, text="CLOSED", text_color="gray", font=("Arial", 9)).pack(anchor="center")

                # DIVIDER
                ctk.CTkLabel(cell, text="---", text_color="gray", font=("Arial", 8)).pack(pady=0)

                # PM Half (#2b2b2b is Dark Gray)
                f_pm = ctk.CTkFrame(cell, fg_color="#2b2b2b" if needs_pm > 0 else "transparent", border_width=0)
                f_pm.pack(fill="both", expand=True, padx=1, pady=(0, 1))

                if needs_pm > 0:
                     ctk.CTkLabel(f_pm, text="PM:", font=("Arial", 8, "bold"), text_color="gray").pack(anchor="w", padx=2)
                     for name in day_sched['PM']:
                         col = c_map.get(name, "white") if use_colors else "#ffccaa"
                         ctk.CTkLabel(f_pm, text=name, text_color=col, font=("Arial", 9)).pack(anchor="w", padx=2)
                     if len(day_sched['PM']) < needs_pm:
                         ctk.CTkLabel(f_pm, text=f"[OPEN]", text_color="#ff5555", font=("Arial", 9)).pack(anchor="w", padx=2)
                else:
                     ctk.CTkLabel(f_pm, text="CLOSED", text_color="gray", font=("Arial", 9)).pack(anchor="center")
            
            if has_weekday: r += 1

    def render_stats(self, logic, grid_frame, status_lbl, warn_box):
        for w in grid_frame.winfo_children():
            if int(w.grid_info()["row"]) > 0: w.destroy()
        
        r = 1
        total_slots = 0
        filled_slots = 0
        for d, reqs in logic.daily_needs.items():
            total_slots += reqs.get('AM', 0) + reqs.get('PM', 0)
            filled_slots += len(logic.schedule.get(d, {}).get('AM', [])) + len(logic.schedule.get(d, {}).get('PM', []))
        
        if filled_slots < total_slots:
            status_lbl.configure(text=f"MISSING SHIFTS: {filled_slots}/{total_slots}", text_color="red")
        else:
            status_lbl.configure(text="ALL CLINIC SHIFTS FILLED ✔", text_color="#00ff00")

        for p in logic.physicians:
            act = len(p.assigned_shifts)
            net = act - p.target
            
            if net > 0: net_col = "yellow"; net_str = f"+{net}"
            elif net < 0: net_col = "#00BFFF"; net_str = f"{net}" 
            else: net_col = "#00FF00"; net_str = "0" 
            
            ctk.CTkLabel(grid_frame, text=p.name).grid(row=r, column=0, sticky="w", padx=5)
            ctk.CTkLabel(grid_frame, text=str(p.target)).grid(row=r, column=1, sticky="e", padx=5)
            ctk.CTkLabel(grid_frame, text=str(act)).grid(row=r, column=2, sticky="e", padx=5)
            ctk.CTkLabel(grid_frame, text=net_str, text_color=net_col).grid(row=r, column=3, sticky="e", padx=5)
            r += 1

        warn_box.delete("0.0", "end")
        if not logic.warnings:
            warn_box.insert("end", "No issues or deviations found.")
        else:
            for w in logic.warnings:
                warn_box.insert("end", f"- {w}\n")

    # --- TAB 4 ---
    def setup_tab4(self):
        f = ctk.CTkFrame(self.tab_imp)
        f.pack(expand=True)
        ctk.CTkLabel(f, text="Excel Operations", font=("Arial", 20)).pack(pady=20)
        ctk.CTkButton(f, text="Export Schedule (Template Format)", width=250, height=50, command=self.export_excel).pack(pady=10)
        ctk.CTkButton(f, text="Import State from Excel", width=250, height=50, command=self.import_excel).pack(pady=10)

    def export_excel(self):
        if not self.scheduler_logic: 
            messagebox.showwarning("Warning", "Please generate a schedule first (Tab 3).")
            return
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not filename: return
        
        try:
            wb = openpyxl.Workbook()
            # SHEET 1: Printable Template
            ws = wb.active
            ws.title = f"Schedule {self.var_month.get()}-{self.var_year.get()}"
            
            # Styles
            bold_font = Font(bold=True, size=11)
            header_font = Font(bold=True, size=14)
            center_align = Alignment(horizontal="center", vertical="center")
            border_medium = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
            fill_black = PatternFill("solid", fgColor="000000")
            
            # Map physician colors to fills
            c_map = {p.name: p.color.replace("#","") for p in self.scheduler_logic.physicians}
            
            # Headers
            ws.merge_cells("A1:E1")
            ws["A1"] = f"Outpatient Center-LBJ Group A {calendar.month_name[int(self.var_month.get())]} {self.var_year.get()}"
            ws["A1"].font = header_font
            ws["A1"].alignment = center_align
            
            ws.merge_cells("G1:K1")
            ws["G1"] = f"Outpatient Center-LBJ Group B {calendar.month_name[int(self.var_month.get())]} {self.var_year.get()}"
            ws["G1"].font = header_font
            ws["G1"].alignment = center_align

            # Black Bar
            ws.column_dimensions['F'].width = 4
            for r in range(1, 100):
                ws[f"F{r}"].fill = fill_black

            # Day Headers
            days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
            for i, d in enumerate(days):
                col_let = get_column_letter(i+1)
                ws[f"{col_let}3"] = d
                ws[f"{col_let}3"].font = bold_font
                ws[f"{col_let}3"].alignment = center_align
                ws.column_dimensions[col_let].width = 25
                
                col_let_b = get_column_letter(i+7)
                ws[f"{col_let_b}3"] = d
                ws[f"{col_let_b}3"].font = bold_font
                ws[f"{col_let_b}3"].alignment = center_align
                ws.column_dimensions[col_let_b].width = 25
            
            logic = self.scheduler_logic
            cal = calendar.monthcalendar(logic.year, logic.month)
            current_row = 4
            
            for week in cal:
                has_weekday = any(d != 0 and idx < 5 for idx, d in enumerate(week))
                if not has_weekday: continue
                
                for day_idx in range(5):
                    day_num = week[day_idx]
                    if day_num == 0: continue
                    
                    # Date Header (with box border later)
                    ws.cell(row=current_row, column=day_idx+1, value=day_num).font = bold_font
                    ws.cell(row=current_row, column=day_idx+1).alignment = center_align
                    ws.cell(row=current_row, column=day_idx+7, value=day_num).font = bold_font
                    ws.cell(row=current_row, column=day_idx+7).alignment = center_align
                    
                    sched = logic.schedule.get(day_num, {'AM': [], 'PM': []})
                    
                    # --- AM ---
                    am_docs = sched['AM']
                    if am_docs:
                        primary_group = 'A'
                        if day_idx == 0 or day_idx == 4: primary_group = 'B'
                        
                        doc1 = am_docs[0] if len(am_docs) > 0 else ""
                        overflow = am_docs[1:]

                        ws.cell(row=current_row+1, column=day_idx+1, value="AM CLINIC").font = Font(size=9)
                        ws.cell(row=current_row+1, column=day_idx+7, value="AM CLINIC").font = Font(size=9)
                        
                        target_col = day_idx + 1 if primary_group == 'A' else day_idx + 7
                        cell = ws.cell(row=current_row+2, column=target_col, value=doc1)
                        cell.font = bold_font
                        if doc1 in c_map:
                             cell.fill = PatternFill("solid", fgColor=c_map[doc1])

                        # Overflow goes to B
                        if overflow:
                            txt = "/ ".join(overflow)
                            cell_b = ws.cell(row=current_row+2, column=day_idx+7)
                            curr = cell_b.value
                            cell_b.value = f"{curr} / {txt}" if curr else txt
                            cell_b.font = bold_font
                    
                    # --- PM ---
                    pm_docs = sched['PM']
                    if pm_docs:
                        primary_group = 'A'
                        if day_idx == 2 or day_idx == 3: primary_group = 'A'
                        
                        doc1 = pm_docs[0] if len(pm_docs) > 0 else ""
                        overflow = pm_docs[1:]

                        ws.cell(row=current_row+7, column=day_idx+1, value="PM CLINIC").font = Font(size=9)
                        ws.cell(row=current_row+7, column=day_idx+7, value="PM CLINIC").font = Font(size=9)
                        
                        target_col = day_idx + 1 if primary_group == 'A' else day_idx + 7
                        cell = ws.cell(row=current_row+8, column=target_col, value=doc1)
                        cell.font = bold_font
                        if doc1 in c_map:
                             cell.fill = PatternFill("solid", fgColor=c_map[doc1])

                        if overflow:
                            txt = "/ ".join(overflow)
                            cell_b = ws.cell(row=current_row+8, column=day_idx+7)
                            curr = cell_b.value
                            cell_b.value = f"{curr} / {txt}" if curr else txt
                            cell_b.font = bold_font

                # Borders for Day Blocks (A border around the whole 13-row day block)
                for r_off in range(13):
                    r = current_row + r_off
                    for c_idx in range(1, 6): # A-E
                        side_l = Side(style='medium') if c_idx==1 else Side(style='thin')
                        side_r = Side(style='medium') if c_idx==5 else Side(style='thin')
                        side_t = Side(style='medium') if r_off==0 else Side(style='thin')
                        side_b = Side(style='medium') if r_off==12 else Side(style='thin')
                        ws.cell(row=r, column=c_idx).border = Border(left=side_l, right=side_r, top=side_t, bottom=side_b)

                    for c_idx in range(7, 12): # G-K
                        side_l = Side(style='medium') if c_idx==7 else Side(style='thin')
                        side_r = Side(style='medium') if c_idx==11 else Side(style='thin')
                        side_t = Side(style='medium') if r_off==0 else Side(style='thin')
                        side_b = Side(style='medium') if r_off==12 else Side(style='thin')
                        ws.cell(row=r, column=c_idx).border = Border(left=side_l, right=side_r, top=side_t, bottom=side_b)

                current_row += 13 

            # SHEET 2: Visual Calendar + Hidden Settings
            ws_cal = wb.create_sheet("Calendar & Settings")
            ws_cal.column_dimensions['A'].width = 15
            ws_cal.column_dimensions['B'].width = 15
            ws_cal.column_dimensions['C'].width = 15
            ws_cal.column_dimensions['D'].width = 15
            ws_cal.column_dimensions['E'].width = 15
            
            ws_cal["A1"] = f"Visual Schedule: {self.var_month.get()}-{self.var_year.get()}"
            ws_cal["A1"].font = Font(bold=True, size=16)
            
            days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
            for i, d in enumerate(days):
                c = ws_cal.cell(row=2, column=i+1, value=d)
                c.font = bold_font
                c.alignment = center_align
                c.fill = PatternFill("solid", fgColor="DDDDDD")
            
            curr_row = 3
            for week in cal:
                has_weekday = any(d != 0 and idx < 5 for idx, d in enumerate(week))
                if not has_weekday: continue
                
                for day_idx in range(5):
                    day_num = week[day_idx]
                    cell = ws_cal.cell(row=curr_row, column=day_idx+1)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = border_medium
                    
                    if day_num == 0: continue
                    
                    sched = logic.schedule.get(day_num, {'AM': [], 'PM': []})
                    txt = f"{day_num}\n\nAM: {', '.join(sched['AM'])}\n\nPM: {', '.join(sched['PM'])}"
                    cell.value = txt
                curr_row += 1
            
            # HIDDEN SETTINGS in A80
            self.update_physician_objects_from_ui()
            data = [p.to_dict() for p in self.physicians]
            setting_cell = ws_cal.cell(row=80, column=1, value=json.dumps(data))
            setting_cell.alignment = Alignment(wrap_text=False) # Clip logic
            # Hide row 80? Optional, but putting it far down is usually enough
            
            wb.save(filename)
            messagebox.showinfo("Success", "Export complete.")
            
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def import_excel(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not filename: return
        try:
            wb = openpyxl.load_workbook(filename)
            # Look for settings in 'Calendar & Settings' cell A80, or legacy 'Settings' A1
            json_str = None
            if "Calendar & Settings" in wb.sheetnames:
                json_str = wb["Calendar & Settings"]["A80"].value
            elif "Settings" in wb.sheetnames:
                json_str = wb["Settings"]["A1"].value
            
            if json_str:
                data = json.loads(json_str)
                self.save_snapshot()
                self.physicians = [Physician.from_dict(d) for d in data]
                self.refresh_physician_list()
                messagebox.showinfo("Success", "Restored.")
            else:
                messagebox.showerror("Error", "No settings found in A80 of Sheet 2.")
        except Exception as e: messagebox.showerror("Error", str(e))

    def setup_tab5(self):
        txt = ctk.CTkTextbox(self.tab_hlp, font=("Arial", 14))
        txt.pack(fill="both", expand=True, padx=10, pady=10)
        info = """CLINIC SCHEDULER PRO - USER GUIDE

1. QUICK TIPS
- Order Matters: Doctors at the top of the list (Tab 1) get priority for scheduling.
- Use Checkboxes: Use '1st Half' / '2nd Half' checkboxes instead of manual dates where possible.
- Match Totals: Ensure the total number of shifts required (Tab 2) matches the total targets set for doctors (Tab 1).

2. TAB EXPLANATIONS
- Tab 1 (Physicians): Manage roster, set targets, and assign specific days off. 'Override' forces a specific date assignment.
- Tab 2 (Needs): Define how many doctors are needed for every AM and PM slot.
    - Set '2nd Half Starts On' to adjust when the monthly split occurs (default 16th).
- Tab 3 (Schedule): View the final schedule.
    - Light Gray = AM
    - Dark Gray = PM
- Tab 3.5 (Color): View schedule using assigned physician colors.

3. EXPORTING
- The Export function creates an Excel file with two sheets:
    - Sheet 1: A formatted 'Group A / Group B' template ready for printing.
    - Sheet 2: A simple visual calendar for reference.
    - NOTE: Your settings are saved invisibly in Sheet 2 (Cell A80). You can Import this file to restore your roster.

----------------------------------------------------------------------------------

DEEP DIVE: LOGIC & SETTINGS

A. ALGORITHM PHASES
1. Overrides: This is the "Nuclear Option". If you put dates in the Override box, the doctor IS working that day. It ignores everything else (Needs, Avoids, Targets). Use this for fixed admin days.
2. Preferences: The system tries to give doctors their preferred dates, but ONLY if:
   a) The clinic actually needs a doctor that shift.
   b) The doctor hasn't hit their target yet.
   c) It doesn't conflict with their Avoid dates.
3. Round Robin (Randomized): This fills the rest of the schedule.
   - It looks at everyone who hasn't met their target.
   - It picks the highest priority doctor available.
   - It finds all valid open slots for them.
   - It picks one RANDOMLY (this ensures if you click Regenerate, you get different results).

B. EDGE CASES
- "Critical: Unfilled Slot": This means you need e.g. 40 shifts covered (Tab 2), but your doctors' targets (Tab 1) only add up to 38. You need to increase targets or add a doctor.
- "Target Unmet": You have too many doctors and not enough open slots. Decrease targets or open more clinic slots.
- "2nd Half Split": If a doctor is "1st Half Only", they will never be scheduled on or after the split date defined in Tab 2.
"""
        txt.insert("0.0", info)
        txt.configure(state="disabled")

if __name__ == "__main__":
    app = AppUI()
    app.mainloop()