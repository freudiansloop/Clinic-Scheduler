import json
import calendar
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def generate_export_excel(data, schedule_results, temp_path):
    """
    Generates the Excel schedule mimicking the v1.6.0 template.
    Returns the path to the temporary generated file.
    """
    wb = Workbook()
    ws = wb.active
    
    year = int(data["needs"].get("year", 2024))
    month = int(data["needs"].get("month", 3))
    ws.title = f"Schedule {month}-{year}"
    
    # --- STYLES ---
    title_font = Font(name='Calibri', size=16, bold=True)
    bold_font = Font(name='Calibri', size=11, bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    fill_black = PatternFill("solid", fgColor="000000")
    
    c_map = {p["name"].strip(): p.get("color", "#FFFFFF").replace("#", "") for p in data["physicians"]}
    
    # --- LAYOUT SETUP ---
    ws.column_dimensions['A'].width = 3 
    col_width = 20
    gps = ['B', 'C', 'D', 'E', 'F']
    gps2 = ['H', 'I', 'J', 'K', 'L']
    for c in gps + gps2:
        ws.column_dimensions[c].width = col_width
    ws.column_dimensions['G'].width = 12 
    
    # --- HEADERS ---
    ws.merge_cells("B1:F1")
    ws["B1"] = f"Outpatient Center-LBJ Group A {calendar.month_name[month]} {year}"
    ws["B1"].font = title_font
    ws["B1"].alignment = center_align
    
    ws.merge_cells("H1:L1")
    ws["H1"] = f"Outpatient Center-LBJ Group B {calendar.month_name[month]} {year}"
    ws["H1"].font = title_font
    ws["H1"].alignment = center_align

    days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    for i, d in enumerate(days):
        col_let = get_column_letter(i+2) # Start B
        ws[f"{col_let}3"] = d
        ws[f"{col_let}3"].font = bold_font
        ws[f"{col_let}3"].alignment = center_align
        
        col_let_b = get_column_letter(i+8) # Start H
        ws[f"{col_let_b}3"] = d
        ws[f"{col_let_b}3"].font = bold_font
        ws[f"{col_let_b}3"].alignment = center_align

    cal = calendar.monthcalendar(year, month)
    current_row = 4
    
    schedule = schedule_results.get("schedule", {}) if schedule_results else {}
    
    for week in cal:
        has_weekday = any(d != 0 and idx < 5 for idx, d in enumerate(week))
        if not has_weekday: continue
        
        start_row_of_block = current_row
        
        for day_idx in range(5):
            day_num = week[day_idx]
            if day_num == 0: continue
            
            col_a = day_idx + 2
            col_b = day_idx + 8
            
            ws.cell(row=current_row, column=col_a, value=day_num).font = bold_font
            ws.cell(row=current_row, column=col_a).alignment = center_align
            ws.cell(row=current_row, column=col_b, value=day_num).font = bold_font
            ws.cell(row=current_row, column=col_b).alignment = center_align
            
            sched = schedule.get(str(day_num), {'AM': [], 'PM': []})
            
            # --- Helper to split docs ---
            def get_group_assignments(day_idx, docs, is_pm):
                if not docs: return ([], [])
                gA = []
                gB = []
                primary = docs[0]
                secondary = docs[1:]
                
                primary_is_A = True
                if not is_pm:
                    if day_idx == 0 or day_idx == 4: # Mon, Fri
                        primary_is_A = False
                else:
                    if day_idx == 2 or day_idx == 3: # Wed, Thu
                        primary_is_A = True
                    else:
                        primary_is_A = False
                
                if primary_is_A:
                    gA.append(primary)
                    gB.extend(secondary)
                else:
                    gB.append(primary)
                    gA.extend(secondary)
                    
                return gA, gB

            # --- AM ---
            am_A, am_B = get_group_assignments(day_idx, sched['AM'], is_pm=False)
            
            if am_A:
                ws.cell(row=current_row+1, column=col_a, value="AM CLINIC").font = bold_font
                doc1 = am_A[0]
                ws.cell(row=current_row+2, column=col_a, value=doc1).font = bold_font
                if doc1.strip() in c_map:
                     ws.cell(row=current_row+2, column=col_a).fill = PatternFill("solid", fgColor=c_map[doc1.strip()])
                overflow = am_A[1:]
                if overflow:
                     txt = " / ".join(overflow)
                     ws.cell(row=current_row+2, column=col_a, value=f"{doc1} / {txt}")

            if am_B:
                ws.cell(row=current_row+1, column=col_b, value="AM CLINIC").font = bold_font
                doc1 = am_B[0]
                ws.cell(row=current_row+2, column=col_b, value=doc1).font = bold_font
                if doc1.strip() in c_map:
                     ws.cell(row=current_row+2, column=col_b).fill = PatternFill("solid", fgColor=c_map[doc1.strip()])
                overflow = am_B[1:]
                if overflow:
                     txt = " / ".join(overflow)
                     ws.cell(row=current_row+2, column=col_b, value=f"{doc1} / {txt}")

            # --- PM ---
            pm_A, pm_B = get_group_assignments(day_idx, sched['PM'], is_pm=True)
            
            if pm_A:
                ws.cell(row=current_row+7, column=col_a, value="PM CLINIC").font = bold_font
                doc1 = pm_A[0]
                ws.cell(row=current_row+8, column=col_a, value=doc1).font = bold_font
                if doc1.strip() in c_map:
                     ws.cell(row=current_row+8, column=col_a).fill = PatternFill("solid", fgColor=c_map[doc1.strip()])
                overflow = pm_A[1:]
                if overflow:
                     txt = " / ".join(overflow)
                     ws.cell(row=current_row+8, column=col_a, value=f"{doc1} / {txt}")

            if pm_B:
                ws.cell(row=current_row+7, column=col_b, value="PM CLINIC").font = bold_font
                doc1 = pm_B[0]
                ws.cell(row=current_row+8, column=col_b, value=doc1).font = bold_font
                if doc1.strip() in c_map:
                     ws.cell(row=current_row+8, column=col_b).fill = PatternFill("solid", fgColor=c_map[doc1.strip()])
                overflow = pm_B[1:]
                if overflow:
                     txt = " / ".join(overflow)
                     ws.cell(row=current_row+8, column=col_b, value=f"{doc1} / {txt}")

        # --- BORDERS ---
        from openpyxl.styles import Border, Side
        for day_idx in range(5):
            if week[day_idx] == 0: continue
            
            col_a_idx = day_idx + 2
            col_b_idx = day_idx + 8
            
            for r_off in range(13):
                r = start_row_of_block + r_off
                cell_a = ws.cell(row=r, column=col_a_idx)
                cell_b = ws.cell(row=r, column=col_b_idx)
                
                l = Side(style='medium')
                r_side = Side(style='medium')
                t = Side(style='medium') if r_off == 0 else None
                b = Side(style='medium') if r_off == 12 else None
                
                if r_off == 0:
                    b = Side(style='medium')
                if r_off == 7:
                    t = Side(style='thin', color="808080")
                
                border = Border(left=l, right=r_side, top=t, bottom=b)
                cell_a.border = border
                cell_b.border = border

        current_row += 13
        
    # Fill center column G gap with black block for week divider visual
    for r in range(4, current_row):
        ws[f"G{r}"].fill = fill_black

    # --- SAVE EMBEDDED SETTINGS DATA TO CELL A80 OF NEW SHEET ---
    ws2 = wb.create_sheet("Calendar & Settings")
    ws2["A1"] = "WARNING: DO NOT EDIT THIS SHEET. IT CONTAINS THE STATE DICTIONARY FOR THE GENERATOR."
    ws2["A80"] = json.dumps(data)

    wb.save(temp_path)
    return temp_path

def parse_import_excel(filepath):
    """
    Parses an uploaded excel schedule template to extract the state JSON.
    """
    wb = load_workbook(filepath)
    json_str = None
    if "Calendar & Settings" in wb.sheetnames:
        json_str = wb["Calendar & Settings"]["A80"].value
    elif "Settings" in wb.sheetnames:
        json_str = wb["Settings"]["A1"].value
    
    if json_str:
        return json.loads(json_str)
    return None
